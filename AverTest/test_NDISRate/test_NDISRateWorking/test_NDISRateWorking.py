import datetime
import math
import re
from selenium.webdriver.support.select import Select
import time
import openpyxl
from datetime import datetime,date
import datetime as datetime
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from sys import platform
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import os
import random
import string


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("email").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("password").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_NDISRateWorking"
  description = "This test scenario is to verify the Working of NDIS Rate process"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_NDISRateWorking"
  global Exe
  Exe="Yes"
  Directory = 'test_NDISRate/'



  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'D:/AVER/AverTest/' + Directory

  MachineName = os.getenv('COMPUTERNAME')
  print(MachineName)
  if MachineName=="DESKTOP-JLLTS65":
      path=path.replace('D:', 'C:')

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      if platform == "linux" or platform == "linux2":
          driver = webdriver.Chrome(executable_path="/home/legion/office 1wayit/AVER/AverTest/chrome/chromedriverLinux1")
      elif platform == "win32" or platform == "win64":
          if MachineName == "DESKTOP-JLLTS65":
              driver = webdriver.Chrome(executable_path="C:/AVER/AverTest/chrome/chromedriver.exe")
          else:
            driver = webdriver.Chrome(executable_path="D:/AVER/AverTest/chrome/chromedriver.exe")

      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://averreplica.1wayit.com/login")
      enter_username("admin@averplanning.com")
      enter_password("admin786")
      driver.find_element_by_xpath("//button[@type='submit']").click()

  yield
  if Exe == "Yes":
      time_change = datetime.timedelta(hours=5)
      new_time = datetime.datetime.now() + time_change
      ctReportHeader = new_time.strftime("%d %B %Y %I %M%p")

      ct = new_time.strftime("%d_%B_%Y_%I_%M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      #driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    global select
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 3
        LONG_TIMEOUT = 60
        LOADING_ELEMENT_XPATH = "//body[@class='sidebar-xs loader_overlay']"

        try:
            print()
            # ---------------------------Verify NDIS Rate Page-----------------------------
            PageName = "NDIS Rate page"
            Ptitle1 = ""
            try:
                driver.find_element_by_xpath("//i[@class='icon-paragraph-justify3']/parent::a").click()
                time.sleep(2)
                driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[6]/a").click()
                time.sleep(2)

                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break
                time.sleep(2)
                TestResult.append(PageName + " opened successfully")
                TestResultStatus.append("Pass")
            except Exception as ee:
                print(ee)
                TestResult.append(PageName + " is not able to open")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify NDIS upload functionality-----------------------------

            import csv
            RegGroupNumberxl = []
            SupportItemNamexl = []
            SupportItemNumberxl = []
            Unitxl = []

            xcelFileName = "NDISData"
            locx1 = (path + 'NDISData/' + xcelFileName + '.csv')
            file = open(locx1)
            csvreader = csv.reader(file)
            header = next(csvreader)
            #print(header)
            TestResult.append("Reading NDIS rates csv file")
            TestResultStatus.append("Pass")

            for row in csvreader:
                #print(row[0])
                SupportItemNumberxl.append(row[0])
                #print(row[1])
                SupportItemNamexl.append(row[1])
                #print(row[2])
                RegGroupNumberxl.append(row[2])
                #print(row[6])
                Unitxl.append(row[6])
            file.close()

            print(RegGroupNumberxl)
            print(SupportItemNamexl)
            print(SupportItemNumberxl)
            print(Unitxl)

            NewDate = date.today()
            NewDate = NewDate.strftime("%d-%m-%Y")
            EffectiveDate=NewDate
            EffectiveDate="11-03-2022"
            PageName="NDIS upload button"
            try:
                driver.find_element_by_xpath("//a[text()='Upload']").click()
                TestResult.append(PageName + " clicked successfully")
                TestResultStatus.append("Pass")
                time.sleep(2)
                driver.find_element_by_xpath("//input[@name='effective_date']").send_keys(EffectiveDate)
                time.sleep(2)
                ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                time.sleep(2)
                driver.find_element_by_xpath("//input[@name='ndis_rate_file']").send_keys(locx1)
                time.sleep(2)
                driver.find_element_by_xpath("//button[text()='Save']").click()
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break

            except Exception as ee:
                print(ee)
                TestResult.append(PageName + " is not able to click")
                TestResultStatus.append("Fail")
                time.sleep(2)

            TValuseDict = {}
            BodyColsList1 = []
            BodyColsList2 = []
            BodyColsList3 = []
            BodyColsList4 = []
            BodyColsList5 = []
            BodyColsList6 = []
            try:
                print()
                try:
                    try:
                        for tc in range(1,7):
                            THead = driver.find_element_by_xpath(
                                "//table[@id='ndis_rate_table_data']/thead/tr/th[" + str(tc) + "]").text
                            TValuseDict[THead]=None
                    except Exception as d:
                        pass
                    try:
                        BodyRows = driver.find_elements_by_xpath("//table[@id='ndis_rate_table_data']/tbody/tr")
                        BodyRows = len(BodyRows)
                        #print(BodyRows)
                        for i in range(BodyRows):
                            #print("i is "+str(i))
                            for i1 in range(1,7):
                                #print("i1 is " + str(i1))
                                keyData = driver.find_element_by_xpath(
                                    "//table[@id='ndis_rate_table_data']/thead/tr[1]/th[" + str(i1) + "]").text
                                #print(keyData)
                                BodyColsText = driver.find_element_by_xpath("//table[@id='ndis_rate_table_data']/tbody/tr["+str(i+1)+"]/td["+str(i1)+"]").text
                                #print(BodyColsText)
                                if keyData=="#":
                                    BodyColsList1.append(BodyColsText)
                                    TValuseDict[keyData] = BodyColsList1
                                elif  keyData=="Registration Group Number":
                                    BodyColsList2.append(BodyColsText)
                                    TValuseDict[keyData] = BodyColsList2
                                elif  keyData=="Service Support Item":
                                    BodyColsList3.append(BodyColsText)
                                    TValuseDict[keyData] = BodyColsList3
                                elif  keyData=="Support Item Number":
                                    BodyColsList4.append(BodyColsText)
                                    TValuseDict[keyData] = BodyColsList4
                                elif  keyData=="Unit of Measure":
                                    BodyColsList5.append(BodyColsText)
                                    TValuseDict[keyData] = BodyColsList5
                                elif  keyData=="Effective Date":
                                    BodyColsList6.append(BodyColsText)
                                    TValuseDict[keyData] = BodyColsList6

                    except Exception:
                        pass

                except Exception as ee:
                    print(ee)
            except Exception:
                pass
            print(TValuseDict)

            if (TValuseDict["Registration Group Number"] == RegGroupNumberxl):
                print("Registration Group Numbers are matching")
                TestResult.append("Registration Group Numbers are matching")
                TestResultStatus.append("Pass")
            else:
                print("Registration Group Numbers are not matching")
                TestResult.append("Registration Group Numbers are not matching")
                TestResultStatus.append("Fail")

            if (TValuseDict["Service Support Item"] == SupportItemNamexl):
                print("Service Support Item names are matching")
                TestResult.append("Service Support Item names are matching")
                TestResultStatus.append("Pass")
            else:
                print("Service Support Item names are not matching")
                TestResult.append("Service Support Item names are not matching")
                TestResultStatus.append("Fail")

            if (TValuseDict["Support Item Number"] == SupportItemNumberxl):
                print("Support Item Numbers are matching")
                TestResult.append("Support Item Numbers are matching")
                TestResultStatus.append("Pass")
            else:
                print("Support Item Numbers are not matching")
                TestResult.append("Support Item Numbers are not matching")
                TestResultStatus.append("Fail")

            if (TValuseDict["Unit of Measure"] == Unitxl):
                print("Unit of Measures are matching")
                TestResult.append("Unit of Measures are matching")
                TestResultStatus.append("Pass")
            else:
                print("Unit of Measures are not matching")
                TestResult.append("Unit of Measures are not matching")
                TestResultStatus.append("Fail")

            for i2 in range(1,BodyRows+1):
                driver.find_element_by_xpath("//table[@id='ndis_rate_table_data']/tbody/tr[" + str(i2) + "]/td[last()]/a[1]").click()
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break
                TestResult.append(
                    "Below NDIS Rate details are found for record no.["+str(i2)+"]:")
                TestResultStatus.append("Pass")

                try:
                    print()
                    RegistrationGroupName = driver.find_element_by_xpath("//form[@class='frm_viw_data mb-4']/div[1]/div/span").text
                    print("Registration Group Name is "+RegistrationGroupName)
                    TestResult.append("Registration Group Name is "+RegistrationGroupName)
                    TestResultStatus.append("Pass")
                    time.sleep(1)

                    SupportCategoryName = driver.find_element_by_xpath("//form[@class='frm_viw_data mb-4']/div[2]/div/span").text
                    print("Support Category Name is "+SupportCategoryName)
                    TestResult.append(
                        "Support Category Name is "+SupportCategoryName)
                    TestResultStatus.append("Pass")
                    time.sleep(1)

                    SupportItemNumber = driver.find_element_by_xpath("//form[@class='frm_viw_data mb-4']/div[3]/div/span").text
                    print("Support Item Number is "+SupportItemNumber)
                    TestResult.append(
                        "Support Item Number is "+SupportItemNumber)
                    TestResultStatus.append("Pass")
                    time.sleep(1)

                    SupportItemName = driver.find_element_by_xpath("//form[@class='frm_viw_data mb-4']/div[4]/div/span").text
                    print("Support Item Name is "+SupportItemName)
                    TestResult.append(
                        "Support Item Name is "+SupportItemName)
                    TestResultStatus.append("Pass")
                    time.sleep(1)

                    Unit = driver.find_element_by_xpath("//form[@class='frm_viw_data mb-4']/div[5]/div/span").text
                    print("Unit is "+Unit)
                    TestResult.append(
                        "Unit is "+Unit)
                    TestResultStatus.append("Pass")
                    time.sleep(1)

                    Quote = driver.find_element_by_xpath("//form[@class='frm_viw_data mb-4']/div[6]/div/span").text
                    print("Quote is "+Quote)
                    TestResult.append(
                        "Quote is "+Quote)
                    TestResultStatus.append("Pass")
                    time.sleep(1)

                except Exception:
                    pass
                print()
                print()
                driver.find_element_by_xpath(
                    "//a[text()='Back']").click()
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break


        except Exception as err:
            print(err)
            TestResult.append("NDIS Rate is not working correctly. Below error found\n"+str(err))
            TestResultStatus.append("Fail")
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


