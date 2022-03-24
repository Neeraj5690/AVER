import datetime
import math
import re
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from sys import platform

from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import os


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("email").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("password").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_HomePageWorking"
  description = "This test scenario is to verify working of Elements present at Homepage"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_ElementsPresent"
  global Exe
  Exe="Yes"
  Directory = 'test_Homepage/'
  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'D:/AVER/AverTest/' + Directory

  MachineName = os.getenv('COMPUTERNAME')
  if MachineName == "DESKTOP-JLLTS65":
      path = path.replace('D:', 'C:')

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      if platform == "linux" or platform == "linux2":
          driver = webdriver.Chrome(executable_path="/home/legion/office 1wayit/AVER/AverTest/chrome/chromedriverLinux")
      elif platform == "win32" or platform == "win64":
          if MachineName == "DESKTOP-JLLTS65":
              driver = webdriver.Chrome(executable_path="C:/AVER/AverTest/chrome/chromedriver.exe")
          else:
              driver = webdriver.Chrome(executable_path="D:/AVER/AverTest/chrome/chromedriver.exe")

      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://averreplica.1wayit.com/login")
      enter_username("admin@averplanning.com")
      enter_password("admin786")
      driver.find_element_by_xpath("//button[@type='submit']").click()

  yield
  if Exe == "Yes":
      time_change = datetime.timedelta(hours=5)
      new_time = datetime.datetime.now() + time_change
      ctReportHeader = new_time.strftime("%d %B %Y %I %M%p")

      ct = new_time.strftime("%d_%B_%Y_%I_%M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 2
        LONG_TIMEOUT = 60
        LOADING_ELEMENT_XPATH = "//div[@class='main-loader LoaderImageLogo']"
        try:
            #---------------------------Verify Login page elements-----------------------------
            PageName="Logo at left panel"
            Ptitle1="rounded-circle"
            try:
                PageTitle1 = driver.find_element_by_xpath("//img[@class='rounded-circle ']").get_attribute('class')
                assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + "  is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName +" is not present")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            #---------------------------------------------------------------------------------

            #---------------------------Verify Hamburger icon-----------------------------
            PageName = "Hamburger menu icon"
            Ptitle1 = ""
            try:
                driver.find_element_by_xpath("//i[@class='icon-paragraph-justify3']/parent::a").click()
                time.sleep(2)
                driver.find_element_by_xpath("//i[@class='icon-paragraph-justify3']/parent::a").click()
                TestResult.append(PageName + "  is present and able to click")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            #---------------------------------------------------------------------------------

            #---------------------------Verify Top Search box-----------------------------
            PageName = "Top Search box"
            Ptitle1 = "Search here"
            try:
                PageTitle1 = driver.find_element_by_xpath("//li[@class='globallisrch']/div/form/input").get_attribute('placeholder')
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not able to open"
                driver.find_element_by_xpath("//li[@class='globallisrch']/div/form/input").send_keys("test")
                TestResult.append(PageName + "  is present and user is able to send inputs")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            #---------------------------------------------------------------------------------

            #---------------------------Verify Calendar icon-----------------------------
            time.sleep(2)
            PageName = "Calendar icon"
            Ptitle1 = ""
            try:
                driver.find_element_by_xpath("//a[@title='Calendar']").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until(EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                TestResult.append(PageName + "  is present and user is able to click")
                TestResultStatus.append("Pass")
            except Exception as qq:
                print(qq)
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[1]/a/i").click()
            time.sleep(TimeSpeed)
            #---------------------------------------------------------------------------------

            #---------------------------Verify My profile drop down-----------------------------
            time.sleep(2)
            PageName = "My profile drop down"
            try:
                driver.find_element_by_xpath("//a[@data-toggle='dropdown']").click()
                time.sleep(2)
                driver.find_element_by_xpath("//a[@data-toggle='dropdown']").click()
                time.sleep(2)
                driver.find_element_by_xpath("//a[@href='https://averreplica.1wayit.com/admin-profile']").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until(EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                TestResult.append(PageName + "  is present and user is able to click")
                TestResultStatus.append("Pass")
            except Exception as aaq:
                print(aaq)
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            time.sleep(2)
            time.sleep(TimeSpeed)
            driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[1]/a/i").click()
            time.sleep(TimeSpeed)

            # ---------------------------Verify Pagination clicks-----------------------------
            RecordsPerPage=10
            try:
                for i1 in range(4):
                    print()
                    try:
                        if i1==0:
                            TestResult.append(
                                "By default [ 10 ] no. of records per page is selected")
                            TestResultStatus.append("Pass")
                        if i1>0:
                            select = Select(driver.find_element_by_xpath("//div[@class='table_data']/div/div[1]/label/select"))
                            select.select_by_index(i1)
                            time.sleep(3)

                        RecordsPerPage = driver.find_element_by_xpath(
                            "//div[@class='table_data']/div/div[1]/label/span/span[1]/span/span[1]").text
                        RecordsPerPage = int(RecordsPerPage)
                        print(RecordsPerPage)
                    except Exception as ww:
                        print(ww)
                        TestResult.append(
                            "For table [ Tracking ] Pagination something went wrong. Below error found\n"+str(ww))
                        TestResultStatus.append("Fail")

                    time.sleep(2)
                    TotalItem = driver.find_element_by_xpath("//div[@id='tracking_listing_info']").text
                    print(TotalItem)

                    if TotalItem!="":
                        ShowingError = "Showing 1 to " + str(RecordsPerPage)
                        if ShowingError not in TotalItem:
                            print("ShowingError found in " + TotalItem)
                            TestResult.append(
                                "For table [ Tracking ] Pagination footer info is wrong. It is showing " + TotalItem + " when selecting pagination for " + str(
                                    RecordsPerPage))
                            TestResultStatus.append("Fail")

                        substr = "of"
                        x = TotalItem.split(substr)
                        string_name = x[0]
                        TotalItemAfterOf = x[1]
                        abc = ""
                        countspace = 0
                        for element in range(0, len(string_name)):
                            if string_name[(len(string_name) - 1) - element] == " ":
                                countspace = countspace + 1
                                if countspace == 2:
                                    break
                            else:
                                abc = abc + string_name[(len(string_name) - 1) - element]
                        abc = abc[::-1]
                        TotalItemBeforeOf = abc
                        TotalItemAfterOf = TotalItemAfterOf.split(" ")
                        TotalItemAfterOf=TotalItemAfterOf[1]
                        TotalItemAfterOf = re.sub('[^A-Za-z0-9]+', '', TotalItemAfterOf)

                        TotalItemAfterOf = int(TotalItemAfterOf)
                        TotalPages = TotalItemAfterOf/RecordsPerPage
                        NumberOfPages = math.ceil(float(TotalPages))
                        print(NumberOfPages)
                        try:
                            ItemLenght=driver.find_elements_by_xpath("//table[@id='tracking_listing']/tbody/tr")
                            #print("ItemLenght is "+str(len(ItemLenght)))
                            if len(ItemLenght)!=RecordsPerPage:
                                TestResult.append("No of records for [ " + str(RecordsPerPage) + " ] is not correct. "+str(len(ItemLenght))+" records were found per page")
                                TestResultStatus.append("Fail")
                        except Exception:
                            pass

                        for i in range(NumberOfPages):
                            if i==NumberOfPages-1:
                                TestResult.append("Pagination for [ "+str(RecordsPerPage)+" ] no. of records is successfully verified")
                                TestResultStatus.append("Pass")
                                break
                            driver.find_element_by_xpath("//div[@class='dataTables_paginate paging_simple_numbers']/a[2]").click()
                            time.sleep(0.5)

                        if i != NumberOfPages - 1:
                            TestResult.append(
                                "Pagination for [ " + str(RecordsPerPage) + " ] no. of records is not working correctly")
                            TestResultStatus.append("Fail")
                        driver.refresh()
                        try:
                            WebDriverWait(driver, SHORT_TIMEOUT
                                          ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                            WebDriverWait(driver, LONG_TIMEOUT
                                          ).until(EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                        except TimeoutException:
                            pass
                    else:
                        print("Blank")
                        TestResult.append(
                            "After selecting Pagination for [ " + str(RecordsPerPage) + " ] no. of records, no data found")
                        TestResultStatus.append("Pass")

            except Exception as aq:
                print(aq)
                TestResult.append("Pagination is not working properly\n"+str(aq))
                TestResultStatus.append("Fail")
            # ---------------------------------------------------------------------------------

        except Exception as err:
            print(err)
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


