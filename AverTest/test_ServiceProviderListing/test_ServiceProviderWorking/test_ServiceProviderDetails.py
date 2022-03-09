import datetime
import math
import random
import re
import os
import time

import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from sys import platform

from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait


@allure.step("Entering username ")
def enter_username(username):
    driver.find_element_by_id("email").send_keys(username)


@allure.step("Entering password ")
def enter_password(password):
    driver.find_element_by_id("password").send_keys(password)


@pytest.fixture()
def test_setup():
    global driver
    global TestName
    global description
    global TestResult
    global TestResultStatus
    global TestDirectoryName
    global path
    global ClickCounter

    TestName = "test_ServiceProviderDetails"
    description = "This test scenario is to verify the details of service provider and reimburse client present in application"
    TestResult = []
    TestResultStatus = []
    TestFailStatus = []
    FailStatus = "Pass"
    TestDirectoryName = "test_ServiceProviderWorking"
    global Exe
    Exe = "Yes"
    Directory = 'test_ServiceProviderListing/'
    if platform == "linux" or platform == "linux2":
        path = '/home/legion/office 1wayit/AVER/AverTest/' + Directory
    elif platform == "win32" or platform == "win64":
        path = 'D:/AVER/AverTest/' + Directory

    MachineName = os.getenv('COMPUTERNAME')
    if MachineName == "DESKTOP-JLLTS65":
        path = path.replace('D:', 'C:')

    ExcelFileName = "Execution"
    locx = (path + 'Executiondir/' + ExcelFileName + '.xlsx')
    wbx = openpyxl.load_workbook(locx)
    sheetx = wbx.active

    for ix in range(1, 100):
        if sheetx.cell(ix, 1).value == None:
            break
        else:
            if sheetx.cell(ix, 1).value == TestName:
                if sheetx.cell(ix, 2).value == "No":
                    Exe = "No"
                elif sheetx.cell(ix, 2).value == "Yes":
                    Exe = "Yes"

    if Exe == "Yes":
        if platform == "linux" or platform == "linux2":
            driver = webdriver.Chrome(
                executable_path="/home/legion/office 1wayit/AVER/AverTest/chrome/chromedriverLinux")
        elif platform == "win32" or platform == "win64":
            if MachineName == "DESKTOP-JLLTS65":
                driver = webdriver.Chrome(executable_path="C:/AVER/AverTest/chrome/chromedriver.exe")
            else:
                driver = webdriver.Chrome(executable_path="D:/AVER/AverTest/chrome/chromedriver.exe")

        driver.implicitly_wait(10)
        driver.maximize_window()
        driver.get("https://averreplica.1wayit.com/login")
        enter_username("admin@averplanning.com")
        enter_password("admin786")
        driver.find_element_by_xpath("//button[@type='submit']").click()

    yield
    if Exe == "Yes":
        ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
        time_change = datetime.timedelta(hours=5)
        new_time = datetime.datetime.now() + time_change
        ctReportHeader = new_time.strftime("%d %B %Y %I %M%p")
        ct1 = new_time.strftime("%d_%B_%Y_%I_%M%p")

        class PDF(FPDF):
            def header(self):
                self.image(path + 'EmailReportContent/logo.png', 10, 8, 33)
                self.set_font('Arial', 'B', 15)
                self.cell(73)
                self.set_text_color(0, 0, 0)
                self.cell(35, 10, ' Test Report ', 1, 1, 'B')
                self.set_font('Arial', 'I', 10)
                self.cell(150)
                self.cell(30, 10, ctReportHeader, 0, 0, 'C')
                self.ln(20)

            def footer(self):
                self.set_y(-15)
                self.set_font('Arial', 'I', 8)
                self.set_text_color(0, 0, 0)
                self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

        pdf = PDF()
        pdf.alias_nb_pages()
        pdf.add_page()
        pdf.set_font('Times', '', 12)
        pdf.cell(0, 10, "Test Case Name:  " + TestName, 0, 1)
        pdf.multi_cell(0, 10, "Description:  " + description, 0, 1)

        for i1 in range(len(TestResult)):
            pdf.set_fill_color(255, 255, 255)
            pdf.set_text_color(0, 0, 0)
            if (TestResultStatus[i1] == "Fail"):
                # print("Fill Red color")
                pdf.set_text_color(255, 0, 0)
                TestFailStatus.append("Fail")
            TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
            pdf.multi_cell(0, 7, str(i1 + 1) + ")  " + TestName1, 0, 1, fill=True)
            TestFailStatus.append("Pass")
        pdf.output(TestName + "_" + ct1 + ".pdf", 'F')

        # -----------To check if any failed Test case present-------------------
        for io in range(len(TestResult)):
            if TestFailStatus[io] == "Fail":
                FailStatus = "Fail"
        # ---------------------------------------------------------------------

        # -----------To add test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path + 'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        print()
        check = TestName
        PdfName = TestName + "_" + ct + ".pdf"
        checkcount = 0

        for i in range(1, 100):
            if sheet.cell(i, 1).value == None:
                if checkcount == 0:
                    sheet.cell(row=i, column=1).value = check
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
                wb.save(loc)
                break
            else:
                if sheet.cell(i, 1).value == check:
                    if checkcount == 0:
                        sheet.cell(row=i, column=2).value = PdfName
                        sheet.cell(row=i, column=3).value = TestDirectoryName
                        sheet.cell(row=i, column=4).value = description
                        sheet.cell(row=i, column=5).value = FailStatus
                        checkcount = 1
        # ----------------------------------------------------------------------------

        # ---------------------To add Test name in Execution sheet--------------------
        ExcelFileName1 = "Execution"
        loc1 = (path + 'Executiondir/' + ExcelFileName1 + '.xlsx')
        wb1 = openpyxl.load_workbook(loc1)
        sheet1 = wb1.active
        checkcount1 = 0

        for ii1 in range(1, 100):
            if sheet1.cell(ii1, 1).value == None:
                if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
                wb1.save(loc1)
                break
            else:
                if sheet1.cell(ii1, 1).value == check:
                    if checkcount1 == 0:
                        sheet1.cell(row=ii1, column=1).value = check
                        checkcount1 = 1
        # -----------------------------------------------------------------------------

        driver.quit()


@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 1
        LONG_TIMEOUT = 100
        LOADING_ELEMENT_XPATH = "//body[@class='sidebar-xs loader_overlay']"
        UName = "admin@averplanning.com"
        PName = "admin786"
        SpDict = {}
        try:

            # ---------------------------Verify Service provider listing icon click-----------------------------
            try:
                print()
                # ---------------------------Service provider listing icon click-----------------------------
                PageName = "Service provider listing icon"
                Ptitle1 = ""
                try:
                   #
                    driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[7]/a").click()
                    time.sleep(2)
                    #driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[8]/ul/li/a").click()

                    for load in range(LONG_TIMEOUT):
                        try:
                            if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                time.sleep(0.5)
                        except Exception:
                            break

                    time.sleep(2)
                    TestResult.append(PageName + " is clicked")
                    TestResultStatus.append("Pass")
                except Exception as ee:
                    print(ee)
                    TestResult.append(PageName + " is not clicked")
                    TestResultStatus.append("Fail")
                print()
                time.sleep(TimeSpeed)
                # ---------------------------------------------------------------------------------
                try:
                    SpPresentxl = "False"
                    xcelFileName = "NewSpRefData"
                    locx1 = (path + 'SpRefData/' + xcelFileName + '.xlsx')
                    wbx1 = openpyxl.load_workbook(locx1)
                    sheetx1 = wbx1.active

                    for i_ref in range(1, 10):
                        if sheetx1.cell(i_ref, 1).value != None:
                            Namexl = sheetx1.cell(i_ref, 1).value
                            SpPresentxl = "True"
                            break

                        else:
                            SpPresentxl = "False"
                            pass

                    if SpPresentxl == "False":
                        TestResult.append("Service provider to verify is not present in reference sheet, it will be re-verified once create service provider job run")
                        TestResultStatus.append("Pass")
                    elif SpPresentxl == "True":
                        TestResult.append(
                            "Service provider is present in reference sheet, searching it in service provider listing table in application")
                        TestResultStatus.append("Pass")

                        check=0
                        #--------------Need to add pagination concept--------------------
                        PaginationClick=3
                        for sp1 in range(PaginationClick):
                            SpTableData = driver.find_elements_by_xpath("//table[@id='table_data']/tbody/tr/td[2]/a")
                            LengthOfSpData = len(SpTableData)
                            print(LengthOfSpData)
                            for sp in range(1,LengthOfSpData+1):
                                SpName = driver.find_element_by_xpath("//table[@id='table_data']/tbody/tr["+str(sp)+"]/td[2]/a").text
                                try:
                                    if Namexl == SpName:
                                        check=1
                                        try:
                                            text1=driver.find_element_by_xpath(
                                                        "//table[@id='table_data']/tbody/tr[" + str(sp) + "]/td[2]/a").text
                                            if text1=="":
                                                text1 = "Blank"
                                        except Exception:
                                            text1 = "Blank"

                                        try:
                                            text2 = driver.find_element_by_xpath(
                                                "//table[@id='table_data']/tbody/tr[" + str(sp) + "]/td[3]").text
                                            if text2=="":
                                                text2 = "Blank"
                                        except Exception:
                                            text2 = "Blank"

                                        try:
                                            text3 = driver.find_element_by_xpath(
                                                "//table[@id='table_data']/tbody/tr[" + str(sp) + "]/td[4]").text
                                            if text3=="":
                                                text3 = "Blank"
                                        except Exception:
                                            text3 = "Blank"

                                        try:
                                            text4 = driver.find_element_by_xpath(
                                                "//table[@id='table_data']/tbody/tr[" + str(sp) + "]/td[5]").text
                                            if text4=="":
                                                text4 = "Blank"
                                        except Exception:
                                            text4 = "Blank"
                                        try:
                                            text5 = driver.find_element_by_xpath(
                                                "//table[@id='table_data']/tbody/tr[" + str(sp) + "]/td[6]").text
                                            if text5=="":
                                                text5 = "Blank"
                                        except Exception:
                                            text5 = "Blank"
                                        try:
                                            text6 = driver.find_element_by_xpath(
                                                "//table[@id='table_data']/tbody/tr[" + str(sp) + "]/td[7]").text
                                            if text6=="":
                                                text6 = "Blank"
                                        except Exception:
                                            text6 = "Blank"

                                        try:
                                            text7 = driver.find_element_by_xpath(
                                                "//table[@id='table_data']/tbody/tr[" + str(sp) + "]/td[8]").text
                                            if text7=="":
                                                text7 = "Blank"
                                        except Exception:
                                            text7 = "Blank"

                                        TestResult.append(
                                            "Service provider is present in table in application. Here are the details\nName : "+text1+"\nService Type : "+text2+"\nAbn : "+text4+"\nAccount Name : "+text5+"\nBsb : "+text6+"\nAccount Number : "+text7)
                                        TestResultStatus.append("Pass")

                                        SpDict["Name"]=text1
                                        SpDict["Service Type"] = text2
                                        SpDict["Franchise"] = text3
                                        SpDict["Abn"] = text4
                                        SpDict["Account Name"] = text5
                                        SpDict["Bsb"] = text6
                                        SpDict["Account Number"] = text7

                                        button = driver.find_element_by_xpath(
                                            "//a[text()='"+SpName+"']")
                                        driver.execute_script("arguments[0].click();", button)
                                        for load in range(LONG_TIMEOUT):
                                            try:
                                                if driver.find_element_by_xpath(
                                                        LOADING_ELEMENT_XPATH).is_displayed() == True:
                                                    time.sleep(0.5)
                                            except Exception:
                                                break

                                        TestResult.append("Checking Service Provide view page details")
                                        TestResultStatus.append("Pass")

                                        #-------Verifying Service provider name---------
                                        try:
                                            Name = driver.find_element_by_xpath("//form[@class='frm_viw_data']/div[1]/div/span").text
                                            print(Name)
                                            print(SpDict["Name"])
                                            if Name != SpDict["Name"]:
                                                print("Name does not match")
                                                TestResult.append("Name does not match")
                                                TestResultStatus.append("Fail")
                                            elif Name == SpDict["Name"]:
                                                print("Name matched")
                                                TestResult.append("Name matched")
                                                TestResultStatus.append("Pass")
                                            print()
                                        except Exception as Ex1:
                                            TestResult.append("Name details are not able to find. Below error found\n"+str(Ex1))
                                            TestResultStatus.append("Fail")

                                        # -------Verifying Service provider Abn---------
                                        try:
                                            Abn = driver.find_element_by_xpath(
                                                "//form[@class='frm_viw_data']/div[3]/div/span").text
                                            print(Abn)
                                            print(SpDict["Abn"])
                                            if Abn != SpDict["Abn"]:
                                                print("Abn does not match")
                                                TestResult.append("Abn does not match")
                                                TestResultStatus.append("Fail")
                                            elif Abn == SpDict["Abn"]:
                                                print("Abn matched")
                                                TestResult.append("Abn matched")
                                                TestResultStatus.append("Pass")
                                            print()
                                        except Exception as Ex2:
                                            TestResult.append("Abn details are not able to find. Below error found\n"+str(Ex2))
                                            TestResultStatus.append("Fail")

                                        # -------Verifying Service provider AccountName---------
                                        try:
                                            AccountName = driver.find_element_by_xpath(
                                                "//form[@class='frm_viw_data']/div[5]/div/span").text
                                            print(AccountName)
                                            print(SpDict["Account Name"])
                                            if AccountName != SpDict["Account Name"]:
                                                print("Account Name does not match")
                                                TestResult.append("Account Name does not match")
                                                TestResultStatus.append("Fail")
                                            elif AccountName == SpDict["Account Name"]:
                                                print("Account Name matched")
                                                TestResult.append("Account Name matched")
                                                TestResultStatus.append("Pass")
                                            print()
                                        except Exception as Ex3:
                                            TestResult.append("Account Name details are not able to find. Below error found\n"+str(Ex3))
                                            TestResultStatus.append("Fail")

                                        # -------Verifying Service provider Bsb---------
                                        try:
                                            Bsb = driver.find_element_by_xpath(
                                                "//form[@class='frm_viw_data']/div[4]/div/span").text
                                            print(Bsb)
                                            print(SpDict["Bsb"])
                                            if Bsb != SpDict["Bsb"]:
                                                print("Bsb does not match")
                                                TestResult.append("Bsb does not match")
                                                TestResultStatus.append("Fail")
                                            elif Bsb == SpDict["Bsb"]:
                                                print("Bsb matched")
                                                TestResult.append("Bsb matched")
                                                TestResultStatus.append("Pass")
                                            print()
                                        except Exception as Ex4:
                                            TestResult.append("Bsb details are not able to find. Below error found\n"+str(Ex4))
                                            TestResultStatus.append("Fail")

                                        # -------Verifying Service provider AccountNumber---------
                                        try:
                                            AccountNumber = driver.find_element_by_xpath(
                                                "//form[@class='frm_viw_data']/div[6]/div/span").text
                                            print(AccountNumber)
                                            print(SpDict["Account Number"])
                                            if AccountNumber != SpDict["Account Number"]:
                                                print("Account Number does not match")
                                                TestResult.append("Account Number does not match")
                                                TestResultStatus.append("Fail")
                                            elif AccountNumber == SpDict["Account Number"]:
                                                print("Account Number matched")
                                                TestResult.append("Account Number matched")
                                                TestResultStatus.append("Pass")
                                            print()
                                        except Exception as Ex5:
                                            TestResult.append("Account Number details are not able to find. Below error found\n"+str(Ex5))
                                            TestResultStatus.append("Fail")

                                        break
                                except Exception as q:
                                    print(q)

                            print()
                            print("sp1 is "+str(sp1))
                            print(PaginationClick)
                            if sp1>=1 and sp1<(PaginationClick-2):
                                print("Yesssssss")
                                print("sp1 is " + str(sp1))
                                print(PaginationClick)
                                try:
                                    driver.find_element_by_xpath("//div[@id='table_data_paginate']/a[2]").click()
                                    time.sleep(2)
                                except Exception as Ex6:
                                    TestResult.append(
                                        "Pagination next icon is not able to click. Below error found\n" + str(Ex6))
                                    TestResultStatus.append("Fail")

                        if check==0:
                            TestResult.append(
                                "Service provider [ "+Namexl +" ] not found in service provider listing table in application")
                            TestResultStatus.append("Fail")
                            sheetx1.cell(i_ref, 1).value= None
                            wbx1.save(locx1)


                except Exception:
                    pass
            except Exception:
                TestResult.append("Not able to verify service provider details")
                TestResultStatus.append("Fail")
        except Exception:
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path + 'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


