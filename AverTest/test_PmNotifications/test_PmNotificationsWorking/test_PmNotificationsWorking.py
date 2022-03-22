import datetime
import math
import re
import time
import openpyxl
from datetime import datetime,date
import datetime as datetime
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from sys import platform
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import pyperclip
import random
import os


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("email").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("password").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_PmNotificationsWorking"
  description = "This test scenario is to verify the Working of Elements at PM Notifications page"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_PmNotificationsWorking"
  global Exe
  Exe="Yes"
  Directory = 'test_PmNotifications/'
  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'D:/AVER/AverTest/' + Directory

  MachineName = os.getenv('COMPUTERNAME')
  if MachineName == "DESKTOP-JLLTS65":
      path = path.replace('D:', 'C:')

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      if platform == "linux" or platform == "linux2":
          driver = webdriver.Chrome(executable_path="/home/legion/office 1wayit/AVER/AverTest/chrome/chromedriverLinux1")
      elif platform == "win32" or platform == "win64":
          if MachineName == "DESKTOP-JLLTS65":
              driver = webdriver.Chrome(executable_path="C:/AVER/AverTest/chrome/chromedriver.exe")
          else:
              driver = webdriver.Chrome(executable_path="D:/AVER/AverTest/chrome/chromedriver.exe")

      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://averreplica.1wayit.com/login")
      enter_username("admin@averplanning.com")
      enter_password("admin786")
      driver.find_element_by_xpath("//button[@type='submit']").click()

  yield
  if Exe == "Yes":
      time_change = datetime.timedelta(hours=5)
      new_time = datetime.datetime.now() + time_change
      ctReportHeader = new_time.strftime("%d %B %Y %I %M%p")

      ct = new_time.strftime("%d_%B_%Y_%I_%M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 3
        LONG_TIMEOUT = 60
        LOADING_ELEMENT_XPATH = "//body[@class='sidebar-xs loader_overlay']"
        try:
            # ---------------------------Verify Client Listing icon click-----------------------------
            PageName = "Client Listing icon"
            try:
                driver.find_element_by_xpath("//i[@class='icon-paragraph-justify3']/parent::a").click()
                time.sleep(2)
                driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[3]/a").click()
                time.sleep(2)
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break
                time.sleep(2)
                TestResult.append(PageName + " is present in left menu and able to click")
                TestResultStatus.append("Pass")
            except Exception as ee:
                print(ee)
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)

            # ---------------------------Fetching all Clients-----------------------------
            ClientList=[]
            try:
                TotalItem = driver.find_element_by_xpath("//div[@id='table_data_info']").text
                print(TotalItem)
                substr = "of"
                x = TotalItem.split(substr)
                string_name = x[0]
                TotalItemAfterOf = x[1]
                abc = ""
                countspace = 0
                for element in range(0, len(string_name)):
                    if string_name[(len(string_name) - 1) - element] == " ":
                        countspace = countspace + 1
                        if countspace == 2:
                            break
                    else:
                        abc = abc + string_name[(len(string_name) - 1) - element]
                abc = abc[::-1]
                TotalItemBeforeOf = abc
                TotalItemAfterOf = TotalItemAfterOf.split(" ")
                TotalItemAfterOf = TotalItemAfterOf[1]
                TotalItemAfterOf = re.sub('[^A-Za-z0-9]+', '', TotalItemAfterOf)
                print(TotalItemAfterOf)

                TotalItemAfterOf = int(TotalItemAfterOf)
                RecordsPerPage = 50
                TotalPages = TotalItemAfterOf / RecordsPerPage
                NumberOfPages = math.ceil(float(TotalPages))
                print(NumberOfPages)

                for i in range(NumberOfPages):
                    NOfRecords = driver.find_elements_by_xpath("//table[@id='table_data']/tbody/tr")
                    print(len(NOfRecords))
                    for i1 in range(len(NOfRecords)):
                        CName = driver.find_element_by_xpath("//table[@id='table_data']/tbody/tr["+str(i1+1)+"]/td[2]").text
                        ClientList.append(CName)
                    if NumberOfPages>1:
                        driver.find_element_by_xpath("//a[@id='table_data_next']").click()
                        i1=0
                        for load in range(LONG_TIMEOUT):
                            try:
                                if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                    time.sleep(0.5)
                            except Exception:
                                break
            except Exception as pmm:
                print(pmm)
                pass

            # ---------------------------Verify PM Notifications icon click-----------------------------
            print(ClientList)

            #ClientList=['avneet1', 'BitsAver', 'BitsInGlass1', 'Carter', 'check', 'fee', 'Fee', 'FName', 'FNameA', 'gagan', 'hdnvn', 'hxnjx', 'ithhr', 'jeudh', 'jsisn', 'kjryw', 'kmbgj', 'liqcl', 'lpazv', 'mandeep', 'Mary', 'Master Terry', 'Miss Annie', 'Mr Brenton', 'mrinh', 'nfyci', 'no email client', 'Note', 'nwwjz', 'otgpw', 'ovljn', 'piteo', 'pljqp', 'qbpkw', 'rdmue', 'Reilly', 'Reilly', 'Rose-Marie', 'rzdul', 'schedule', 'sumreet', 'Sumreet', 'SUNIL', 'tas client', 'test', 'tlaee', 'uegpt', 'vfcex', 'wvxcg', 'yad', 'ycfkd', 'ydcip', 'yycpf', 'zdpqy', 'zqato']
            for Clclick in range (len(ClientList)):
                print(Clclick)
                NameToOpen = ClientList[Clclick]
                print(NameToOpen)
                try:
                    driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[3]/a").click()
                    time.sleep(2)
                    for load in range(LONG_TIMEOUT):
                        try:
                            if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                time.sleep(0.5)
                        except Exception:
                            break
                    driver.find_element_by_xpath("//input[@id='searchFilter']").send_keys(NameToOpen)
                    time.sleep(2)
                    driver.find_element_by_xpath("//button[@id='searchBtn']").click()
                    for load in range(LONG_TIMEOUT):
                        try:
                            if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                time.sleep(0.5)
                        except Exception:
                            break
                    driver.find_element_by_xpath("//table[@id='table_data']/tbody/tr[1]/td[2]/a").click()
                    for load in range(LONG_TIMEOUT):
                        try:
                            if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                time.sleep(0.5)
                        except Exception:
                            break
                    try:
                        button = driver.find_element_by_xpath("//div[@id='alert_modal']/div/div/div/button")
                        driver.execute_script("arguments[0].click();", button)
                    except Exception as dd:
                        print(dd)
                        pass
                    driver.find_element_by_xpath("//tbody/tr/td[@class='ServiceBookingTHwidth']/p/a").click()
                    for load in range(LONG_TIMEOUT):
                        try:
                            if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                time.sleep(0.5)
                        except Exception:
                            break
                    try:
                        forZeroBreak = 0
                        AllocatedAmount = driver.find_element_by_xpath("//tr[@class='LighBluetr yellowTbodyBorder p-0 ']/td[2]/span").text
                        print(AllocatedAmount)
                        time.sleep(2)
                        RemainingAmount = driver.find_element_by_xpath("//tr[@class='LighBluetr yellowTbodyBorder p-0 ']/td[4]/span").text
                        print(RemainingAmount)


                        for char in AllocatedAmount:
                            AllocatedAmount = AllocatedAmount.replace(',', "")
                            temp = re.findall(r'\d+', AllocatedAmount)
                            res = list(map(int, temp))
                            try:
                                AllocatedAmountFound = res[0]
                                AllocatedAmountFound = float(AllocatedAmountFound)
                            except Exception:
                                AllocatedAmountFound = 0.0

                            print(AllocatedAmountFound)
                            if AllocatedAmountFound > 0.0:
                                forZeroBreak = 1
                                break
                            print(AllocatedAmountFound)

                        for char1 in RemainingAmount:
                            RemainingAmount = RemainingAmount.replace(',', "")
                            temp = re.findall(r'\d+', RemainingAmount)
                            res = list(map(int, temp))
                            try:
                                RemainingAmountFound = res[0]
                                RemainingAmountFound = float(RemainingAmountFound)
                            except Exception:
                                RemainingAmountFound = 0.0

                            print(RemainingAmountFound)
                            if RemainingAmountFound > 0.0:
                                forZeroBreak = 1
                                break
                            print(RemainingAmountFound)
                        PercentageAmount = AllocatedAmountFound*50/100
                        print(PercentageAmount)
                        if RemainingAmountFound<PercentageAmount:
                            print("Yes")
                            TestResult.append("Suitable Client found to verify PM Notification. Client name is below\n"+NameToOpen)
                            TestResultStatus.append("Pass")

                            driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[10]/a").click()
                            TestResult.append(
                                "Navigated to PM Notification page successfully")
                            TestResultStatus.append("Pass")

                            time.sleep(2)
                            for load in range(LONG_TIMEOUT):
                                try:
                                    if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                        time.sleep(0.5)
                                except Exception:
                                    break
                            driver.find_element_by_xpath("//a[text()='Back']/parent::div/span[1]/div/button").click()
                            time.sleep(2)
                            driver.find_element_by_xpath("//a[text()='Back']/parent::div/span[1]/div/button").click()
                            time.sleep(2)
                            driver.find_element_by_xpath("//a[text()='Back']/parent::div/span[1]/div/div/div/input").send_keys(NameToOpen)
                            time.sleep(2)
                            ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                            time.sleep(2)
                            ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                            time.sleep(2)
                            driver.find_element_by_xpath("//a[text()='Back']/parent::div/span[1]/div/button").click()
                            TestResult.append(
                                "Client searched to find notification Information")
                            TestResultStatus.append("Pass")

                            FoundClient=driver.find_element_by_xpath("//table[@id='notification_table_data']/tbody/tr[1]/td[2]").text
                            print(FoundClient)
                            FoundNotification = driver.find_element_by_xpath(
                                "//table[@id='notification_table_data']/tbody/tr[1]/td[4]").text
                            print(FoundNotification)
                            if NameToOpen in  FoundClient:
                                print("Client Notifiaction found")
                                TestResult.append(
                                    "Notification Information (as given below) found successfully\n"+FoundNotification)
                                TestResultStatus.append("Pass")
                            break
                        else:
                            pass
                    except Exception as pm:
                        print(pm)
                        pass
                except Exception as ee:
                    print(ee)

            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------
        except Exception as err:
            print(err)
            TestResult.append("PM Notification is not working correctly. Below error found\n"+str(err))
            TestResultStatus.append("Fail")
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


