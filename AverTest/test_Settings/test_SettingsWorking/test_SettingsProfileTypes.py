import datetime
import math
import re
import string
import time
import openpyxl
from datetime import datetime,date
import datetime as datetime
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from sys import platform
import os


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("email").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("password").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_SettingsProfileTypes"
  description = "This test scenario is to verify Profile Type data collected at Settings page"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_SettingsWorking"
  global Exe
  Exe="Yes"
  Directory = 'test_Settings/'
  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'D:/AVER/AverTest/' + Directory

  MachineName = os.getenv('COMPUTERNAME')
  if MachineName == "DESKTOP-JLLTS65":
      path = path.replace('D:', 'C:')

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      if platform == "linux" or platform == "linux2":
          driver = webdriver.Chrome(executable_path="/home/legion/office 1wayit/AVER/AverTest/chrome/chromedriverLinux1")
      elif platform == "win32" or platform == "win64":
          if MachineName == "DESKTOP-JLLTS65":
              driver = webdriver.Chrome(executable_path="C:/AVER/AverTest/chrome/chromedriver.exe")
          else:
              driver = webdriver.Chrome(executable_path="D:/AVER/AverTest/chrome/chromedriver.exe")

      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://averreplica.1wayit.com/login")
      enter_username("admin@averplanning.com")
      enter_password("admin786")
      driver.find_element_by_xpath("//button[@type='submit']").click()

  yield
  if Exe == "Yes":
      time_change = datetime.timedelta(hours=5)
      new_time = datetime.datetime.now() + time_change
      ctReportHeader = new_time.strftime("%d %B %Y %I %M%p")

      ct = new_time.strftime("%d_%B_%Y_%I_%M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 3
        LONG_TIMEOUT = 60
        LOADING_ELEMENT_XPATH = "//body[@class='sidebar-xs loader_overlay']"
        try:
            # ---------------------------Verify Settings icon click-----------------------------
            PageName = "Settings icon"
            Ptitle1 = ""
            try:
                driver.find_element_by_xpath("//i[@class='icon-paragraph-justify3']/parent::a").click()
                time.sleep(2)
                driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[11]/a").click()
                time.sleep(2)

                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break
                time.sleep(2)
                TestResult.append(PageName + " is present in left menu and able to click")
                TestResultStatus.append("Pass")
            except Exception as ee:
                print(ee)
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify working of Profile Types button under system settings-----------------------------
            PageName = "Profile Types button"
            ExpectedDict = {}
            SuccessList = []
            PendingList = []
            try:
                try:
                    driver.find_element_by_xpath("//a[text()='Profile Types']").click()
                    TestResult.append(PageName + " is clickable on settings page")
                    TestResultStatus.append("Pass")
                    for load in range(LONG_TIMEOUT):
                        try:
                            if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                time.sleep(0.5)
                        except Exception:
                            break
                    Rows = driver.find_elements_by_xpath("//table[@id='client_status_data']/tbody/tr")
                    Rows = len(Rows)
                    print(Rows)
                    TestResult.append(
                        "Number of records found on Profile Types settings section is: " + str(Rows))
                    TestResultStatus.append("Pass")
                    for tr in range (Rows):
                        Keys1 = driver.find_element_by_xpath("//table[@id='client_status_data']/tbody/tr["+str(tr+1)+"]/td[3]").text
                        print(Keys1)
                        time.sleep(0.25)
                        Values = driver.find_element_by_xpath("//table[@id='client_status_data']/tbody/tr["+str(tr+1)+"]/td[2]").text
                        print(Values)
                        time.sleep(0.25)
                        if Keys1 == "Success":
                            SuccessList.append(Values)
                            ExpectedDict[Keys1] = SuccessList
                        elif Keys1 == "Pending":
                            PendingList.append(Values)
                            ExpectedDict[Keys1] = PendingList

                    TestResult.append(
                        "Below items found for success status : \n" + str(SuccessList))
                    TestResultStatus.append("Pass")
                    TestResult.append(
                        "Below items found for pending status : \n" + str(PendingList))
                    TestResultStatus.append("Pass")
                except Exception:
                    pass

                try:
                    driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[3]/a/i").click()
                    TestResult.append("Client listing icon is clicked")
                    TestResultStatus.append("Pass")
                    for load in range(LONG_TIMEOUT):
                        try:
                            if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                time.sleep(0.5)
                        except Exception:
                            break
                except Exception:
                    pass


                #-----------------------------------------------------------------------------------------
                try:
                    driver.find_element_by_xpath("//a[text()='Create New Client']").click()
                    TestResult.append(
                        "Clicking on create new client button")
                    TestResultStatus.append("Pass")
                    for load in range(LONG_TIMEOUT):
                        try:
                            if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                time.sleep(0.5)
                        except Exception:
                            break
                except Exception:
                    pass

                #------Fetching dropdown values----------------
                ActSuccessElements=[]
                l = driver.find_element_by_name("type")
                d = Select(l)
                for opt in d.options:
                    print(opt.text)
                    if opt.text !="Select Profile type":
                        ActSuccessElements.append(opt.text)

                #------Comparing results-----------
                print(len(ExpectedDict["Success"]))
                print(len(ActSuccessElements))
                TestResult.append(
                    "Comparing number of items found on Profile type settings section and client Profile type dropdown")
                TestResultStatus.append("Pass")

                if len(ActSuccessElements)==len(ExpectedDict["Success"]):
                    print("Items number matched")
                    print("Items number matched")
                    TestResult.append(
                        "Items number matched")
                    TestResultStatus.append("Pass")
                else:
                    print("Items number does not match")
                    TestResult.append(
                        "Items number does not match")
                    TestResultStatus.append("Fail")


                print(ExpectedDict["Success"])
                print(ActSuccessElements)
                TestResult.append(
                    "Comparing list of items found on Profile type settings section and client Profile type dropdown")
                TestResultStatus.append("Pass")
                if ActSuccessElements==ExpectedDict["Success"]:
                    print("Items list matched")
                    TestResult.append(
                        "Items list matched")
                    TestResultStatus.append("Pass")
                else:
                    print("Items list does not match")
                    TestResult.append(
                        "Items list does not match")
                    TestResultStatus.append("Fail")

            except Exception as wr:
                print(wr)
                TestResult.append("Profile type section is not working correctly")
                TestResultStatus.append("Fail")
            # ---------------------------------------------------------------------------------

            # ---------------------------------------------------------------------------------

        except Exception as err:
            print(err)
            TestResult.append("Settings is not working correctly. Below error found\n"+str(err))
            TestResultStatus.append("Fail")
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


