import datetime
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from sys import platform
import os
from setuptools import glob
from pathlib import Path
import ntpath
import requests as requests

from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("email").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("password").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_FindBrokenLinks"
  description = "This test scenario is to verify Broken links in complete application"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_Working"
  global Exe
  Exe="Yes"
  Directory = 'test_BrokenLinks/'
  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'D:/AVER/AverTest/' + Directory

  MachineName = os.getenv('COMPUTERNAME')
  if MachineName == "DESKTOP-JLLTS65":
      path = path.replace('D:', 'C:')

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      if platform == "linux" or platform == "linux2":
          driver=webdriver.Chrome(executable_path="/home/legion/office 1wayit/AVER/AverTest/chrome/chromedriverLinux")
      elif platform == "win32" or platform == "win64":
          if MachineName == "DESKTOP-JLLTS65":
              driver = webdriver.Chrome(executable_path="C:/AVER/AverTest/chrome/chromedriver.exe")
          else:
              driver = webdriver.Chrome(executable_path="D:/AVER/AverTest/chrome/chromedriver.exe")

      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://averreplica.1wayit.com/login")
      enter_username("admin@averplanning.com")
      enter_password("admin786")
      driver.find_element_by_xpath("//button[@type='submit']").click()

  yield
  if Exe == "Yes":
      time_change = datetime.timedelta(hours=5)
      new_time = datetime.datetime.now() + time_change
      ctReportHeader = new_time.strftime("%d %B %Y %I %M%p")

      ct = new_time.strftime("%d_%B_%Y_%I_%M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 5
        LONG_TIMEOUT = 200
        LOADING_ELEMENT_XPATH = "//body[@class='sidebar-xs loader_overlay']"

        try:
            print()
            Listb = []
            elem = []
            PageName=[]

            for ii in range(14):
                try:
                    if ii!=7:
                        elem.clear()
                        PageName1 = driver.find_element_by_xpath(
                            "//div[@class='card card-sidebar-mobile']/ul/li[" + str(ii) + "]/a").get_attribute('title')
                        print(PageName1)
                        if ii==8:
                            driver.find_element_by_xpath("//i[@class='icon-paragraph-justify3']/parent::a").click()
                            time.sleep(2)
                            driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[8]/a").click()
                            time.sleep(2)
                            driver.find_element_by_xpath(
                                "//div[@class='card card-sidebar-mobile']/ul/li[8]/ul/li/a").click()
                        elif ii == 9:
                            driver.find_element_by_xpath(
                                "//div[@class='card card-sidebar-mobile']/ul/li[" + str(ii) + "]/a").click()
                        else:
                            driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li["+str(ii)+"]/a/i").click()

                        if PageName1 not in PageName:
                            TestResult.append(PageName1 + " opened to collect all available links")
                            TestResultStatus.append("Pass")

                        PageName.append(PageName1)
                        print(PageName)

                        for load in range(LONG_TIMEOUT):
                            try:
                                if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                    time.sleep(0.5)
                            except Exception:
                                break

                        time.sleep(1)
                        elem = driver.find_elements_by_xpath("//a[@href]")
                        for elema in elem:
                            if len(elema.get_attribute("href")) > 1:
                                if elema.get_attribute("href") in Listb:
                                    pass
                                else :
                                    if "logout" in elema.get_attribute("href"):
                                        pass
                                    elif "javascript:void(0)" in elema.get_attribute("href"):
                                        pass
                                    elif "javascript:void(0);" in elema.get_attribute("href"):
                                        pass
                                    elif "javascript:void();" in elema.get_attribute("href"):
                                        pass
                                    else:
                                        Listb.append(elema.get_attribute("href"))
                        print("Total "+str(len(Listb))+" links collected")
                        TestResult.append("Total "+str(len(Listb))+" links collected and verified")
                        TestResultStatus.append("Pass")

                except Exception as dd:
                    pass

            #------------To verify all links are opening correctly with 202 status
            for ad in range (len(Listb)):
                try:
                    print(Listb[ad])
                    driver.get(Listb[ad])

                    try:
                        r = requests.head(Listb[ad])
                        status = r.status_code
                    except Exception as er:
                        status = str(er)
                    if status != 200:
                        print(status)
                        TestResult.append("For link [ "+Listb[ad]+" ] below error found\n" + status)
                        TestResultStatus.append("Fail")

                except Exception as cl:
                    print(cl)
                    pass

            # -------Removing all files from Downloads folder------------
            folder_path = str(Path.home() / "Downloads")
            file_type = r'\*'
            files = glob.glob(folder_path + file_type)

            print(files)
            print(files[0])

            for rem in range(len(files)):
                print(rem)
                filename = ntpath.basename("'r'" + str(files[rem]))
                print(filename)
                if filename == "New folder":
                    print("File found")
                else:
                    os.remove(files[rem])
            TestResult.append("All files in downloads removed successfully")
            TestResultStatus.append("Pass")

        except Exception:
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


