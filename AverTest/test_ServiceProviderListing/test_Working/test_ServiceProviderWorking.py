import datetime
import math
import random
import re
import string
import time
from telnetlib import EC

import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from sys import platform

from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("email").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("password").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path
  global ClickCounter

  TestName = "test_ServiceProviderWorking"
  description = "This test scenario is to verify working of Login Process"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_Working"
  global Exe
  Exe="Yes"
  Directory = 'test_ServiceProviderListing/'
  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'D:/AVER/AverTest/' + Directory

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      if platform == "linux" or platform == "linux2":
          driver=webdriver.Chrome(executable_path="/home/legion/office 1wayit/AVER/AverTest/chrome/chromedriverLinux")
      elif platform == "win32" or platform == "win64":
          driver = webdriver.Chrome(executable_path="D:/AVER/AverTest/chrome/chromedriver.exe")
      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://averreplica.1wayit.com/login")
      enter_username("admin@averplanning.com")
      enter_password("admin786")
      driver.find_element_by_xpath("//button[@type='submit']").click()

  yield
  if Exe == "Yes":
      ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      time_change = datetime.timedelta(hours=5)
      new_time = datetime.datetime.now() + time_change
      ctReportHeader = new_time.strftime("%d %B %Y %I %M%p")
      ct1 = new_time.strftime("%d_%B_%Y_%I_%M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct1 + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 5
        LONG_TIMEOUT = 400
        LOADING_ELEMENT_XPATH = "//div[@id='appian-working-indicator-hidden']"
        UName="admin@averplanning.com"
        PName="admin786"
        try:

            # ---------------------------Verify Service provider listing icon click-----------------------------
            PageName = "Service provider listing icon"
            Ptitle1 = ""
            try:
                driver.find_element_by_xpath("//i[@class='icon-paragraph-justify3']/parent::a").click()
                time.sleep(2)
                driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[8]/a").click()
                time.sleep(2)
                driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[8]/ul/li/a").click()
                time.sleep(2)
                time.sleep(2)
                TestResult.append(PageName + " is present in left menu and able to click")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify working of Back button on Service provider listing page -----------------------------
            PageName = "Back button"
            Ptitle1 = "Rae"
            try:
                driver.find_element_by_xpath("//a[text()='Back']").click()
                time.sleep(2)
                PageTitle1 = driver.find_element_by_xpath("//div[@class='hed_wth_srch']/h2").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is clickable")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not clickable")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ----------------Verify Service provider icon click after verifying back--------
            PageName = "Service provider icon"
            Ptitle1 = ""
            try:
                driver.find_element_by_xpath("//i[@class='icon-paragraph-justify3']/parent::a").click()
                time.sleep(2)
                driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[8]/a").click()
                time.sleep(2)
                driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[8]/ul/li/a").click()
                time.sleep(2)
                TestResult.append(PageName + "  is opened again after verifying back button")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not opened again after verifying back button")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ----------------Verify working of import button-----------------------------------
            PageName = "Import button"
            Ptitle1 = "Import Service Provider"
            try:
                driver.find_element_by_xpath("//a[text()='Import']").click()
                time.sleep(2)
                PageTitle1 = driver.find_element_by_xpath("//h4[text()='Import Service Provider']").text
                time.sleep(2)
                driver.find_element_by_xpath("//a[@class='sbmt_btn close-from']").click()
                time.sleep(2)
                assert PageTitle1 in Ptitle1, PageName + " not able to click"
                TestResult.append(PageName + "  is clickable")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not clickable")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------


            try:
                print()
                # ---------------------------To check we have existing test service provider in excel-----------------------------
                SpPresentxl = "False"
                xcelFileName = "NewSpRefData"
                locx1 = (path + 'SpRefData/' + xcelFileName + '.xlsx')
                wbx1 = openpyxl.load_workbook(locx1)
                sheetx1 = wbx1.active

                for i_ref in range(1, 10):
                    if sheetx1.cell(i_ref, 1).value != None:
                        Namexl = sheetx1.cell(i_ref, 1).value
                        SpPresentxl = "True"
                        break

                    else:
                        SpPresentxl = "False"
                        pass

                if SpPresentxl == "False":
                    print("Service provider is not present in reference sheet, we need to add Service provider first in application")
                    TestResult.append(
                        "Service provider is not present in reference sheet, we need to add Service provider first in application")
                    TestResultStatus.append("Pass")

                    # ---------------------------Verify working of Create new service provider process-----------------------------
                    try:
                        l = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
                        random.shuffle(l)
                        if l[0] == 0:
                            pos = random.choice(range(1, len(l)))
                            l[0], l[pos] = l[pos], l[0]
                        Number = ''.join(map(str, l[0:6]))
                        print(Number)

                        today = datetime.date.today()
                        D1 = today.strftime("%d-%m-%Y")

                        driver.find_element_by_xpath("//a[text()='Create New']").click()
                        time.sleep(2)

                        for rc in range(5):
                            letters = string.ascii_lowercase
                            returna = ''.join(random.choice(letters) for i in range(5))
                            Name = returna
                        print(Name)
                        Data1 = [Name, D1, Number, "@test.com","2132435465"]
                        for ii2 in range(1, 7):
                            driver.find_element_by_xpath(
                                "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(ii2) + "]")
                            if ii2 == 1:
                                for i2 in range(1, 8):
                                    driver.find_element_by_xpath(
                                        "//div[@id='createnewclient']/div/div/div[2]/form/div[1]/div["+str(i2)+"]")
                                    # -------------Name Field--------------------------------------------
                                    if i2 == 1:
                                        time.sleep(1)
                                        driver.find_element_by_xpath(
                                            "//div[@id='createnewclient']/div/div/div[2]/form/div[1]/div["+str(i2)+"]/div/input").send_keys(Data1[0])
                                        TestResult.append("Name is entered successfully")
                                        TestResultStatus.append("Pass")
                                    # -------------Status Dropdown--------------------------------------------
                                    elif i2 == 2:
                                        time.sleep(1)
                                        select = Select(driver.find_element_by_xpath(
                                            "//div[@id='createnewclient']/div/div/div[2]/form/div[1]/div["+str(i2)+"]/div/select"))
                                        select.select_by_visible_text("Active")
                                        TestResult.append("Status is selected successfully")
                                        TestResultStatus.append("Pass")
                                    # -------------Set Up Date Field--------------------------------------------
                                    elif i2 == 3:
                                        time.sleep(1)
                                        driver.find_element_by_xpath(
                                            "//div[@id='createnewclient']/div/div/div[2]/form/div[1]/div["+str(i2)+"]/div/input").clear()
                                        time.sleep(1)
                                        driver.find_element_by_xpath(
                                            "//div[@id='createnewclient']/div/div/div[2]/form/div[1]/div["+str(i2)+"]/div/input").send_keys(Data1[1])
                                        ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                                        time.sleep(2)
                                        TestResult.append("Set up date is entered successfully")
                                        TestResultStatus.append("Pass")
                                    # -------------ABN Field--------------------------------------------
                                    elif i2 == 4:
                                        time.sleep(1)
                                        driver.find_element_by_xpath(
                                            "//div[@id='createnewclient']/div/div/div[2]/form/div[1]/div["+str(i2)+"]/div/input").send_keys(Data1[2])
                                        TestResult.append("ABN is entered successfully")
                                        TestResultStatus.append("Pass")
                                    # -------------Franchise Dropdown--------------------------------------------
                                    elif i2 == 5:
                                        time.sleep(1)
                                        select = Select(driver.find_element_by_xpath(
                                            "//div[@id='createnewclient']/div/div/div[2]/form/div[1]/div["+str(i2)+"]/div/select"))
                                        select.select_by_visible_text("Yes")
                                        TestResult.append("Franchise is selected successfully")
                                        TestResultStatus.append("Pass")
                                    # # -------------Service Type Dropdown--------------------------------------------
                                    # elif i2 == 7:
                                    #     time.sleep(2)
                                    #     select = Select(driver.find_element_by_xpath(
                                    #         "//div[@id='createnewclient']/div/div/div[2]/form/div[1]/div["+str(i2)+"]/div/span/select"))
                                    #     select.select_by_index(1)
                                    #     TestResult.append("Service Type is selected successfully")
                                    #     TestResultStatus.append("Pass")
                            # -------------Office Number Field--------------------------------------------
                            elif ii2 == 2:
                                for ii3 in range(1, 9):
                                    if ii3 == 1:
                                        time.sleep(1)
                                        driver.find_element_by_xpath(
                                            "//div[@id='createnewclient']/div/div/div[2]/form/div[2]/div["+str(ii3)+"]/div/input").send_keys(
                                            Data1[2])
                                        TestResult.append("Office Number is entered successfully")
                                        TestResultStatus.append("Pass")
                                    # -------------Mobile Number Field--------------------------------------------
                                    elif ii3 == 2:
                                        time.sleep(1)
                                        driver.find_element_by_xpath(
                                            "//div[@id='createnewclient']/div/div/div[2]/form/div[2]/div["+str(ii3)+"]/div/input").send_keys(
                                            Data1[4])
                                        TestResult.append("Mobile Number is entered successfully")
                                        TestResultStatus.append("Pass")
                                    # -------------Admin Email Address Field--------------------------------------------
                                    elif ii3 == 3:
                                        time.sleep(1)
                                        driver.find_element_by_xpath(
                                            "//div[@id='createnewclient']/div/div/div[2]/form/div[2]/div["+str(ii3)+"]/div/input").send_keys(Data1[0] + Data1[3])
                                        TestResult.append("Admin Email Address is entered successfully")
                                        TestResultStatus.append("Pass")
                                    # -------------Address Country Field--------------------------------------------
                                    elif ii3 == 4:
                                        time.sleep(1)
                                        select = Select(driver.find_element_by_xpath(
                                            "//div[@id='createnewclient']/div/div/div[2]/form/div[2]/div["+str(ii3)+"]/div/select"))
                                        select.select_by_index(4)
                                        TestResult.append("Address Country is selected successfully")
                                        TestResultStatus.append("Pass")
                                    # -------------Address State Field--------------------------------------------
                                    elif ii3 == 5:
                                        time.sleep(1)
                                        select = Select(driver.find_element_by_xpath(
                                            "//div[@id='createnewclient']/div/div/div[2]/form/div[2]/div["+str(ii3)+"]/div/select"))
                                        select.select_by_index(3)
                                        TestResult.append("Address State is selected successfully")
                                        TestResultStatus.append("Pass")
                                    # -------------Address Street Field--------------------------------------------
                                    elif ii3 == 6:
                                        time.sleep(1)
                                        driver.find_element_by_xpath(
                                            "//div[@id='createnewclient']/div/div/div[2]/form/div[2]/div["+str(ii3)+"]/div/input").send_keys(Data1[0])
                                        TestResult.append("Address Street is entered successfully")
                                        TestResultStatus.append("Pass")
                                    # -------------Post Code Field--------------------------------------------
                                    elif ii3 == 7:
                                        time.sleep(1)
                                        driver.find_element_by_xpath("//div[@id='createnewclient']/div/div/div[2]/form/div[2]/div["+str(ii3)+"]/div/input").send_keys(Data1[2])
                                        TestResult.append("Post Code is entered successfully")
                                        TestResultStatus.append("Pass")
                                    # -------------Suburb Field--------------------------------------------
                                    elif ii3 == 8:
                                        time.sleep(1)
                                        driver.find_element_by_xpath(
                                            "//div[@id='createnewclient']/div/div/div[2]/form/div[2]/div["+str(ii3)+"]/div/input").send_keys(Data1[0])
                                        TestResult.append("Suburb is entered successfully")
                                        TestResultStatus.append("Pass")
                            # -------------Account Name Field--------------------------------------------
                            elif ii2 == 3:
                                for cc in range(1,4):
                                    driver.find_element_by_xpath("//div[@id='createnewclient']/div/div/div[2]/form/div[3]/div["+str(cc)+"]")
                                    if cc==1:
                                        driver.find_element_by_xpath("//div[@id='createnewclient']/div/div/div[2]/form/div[3]/div["+str(cc)+"]/div/input").send_keys(Data1[0])
                                        TestResult.append("Account Name is entered successfully")
                                        TestResultStatus.append("Pass")
                                    # -------------Account BSB Field--------------------------------------------
                                    elif cc==2:
                                        driver.find_element_by_xpath("//div[@id='createnewclient']/div/div/div[2]/form/div[3]/div["+str(cc)+"]/div/input").send_keys(Data1[2])
                                        TestResult.append("Account BSB is entered successfully")
                                        TestResultStatus.append("Pass")
                                    # -------------Account Number Field--------------------------------------------
                                    elif cc == 3:
                                        driver.find_element_by_xpath("//div[@id='createnewclient']/div/div/div[2]/form/div[3]/div["+str(cc)+"]/div/input").send_keys(Data1[2])
                                        TestResult.append("Account Number is entered successfully")
                                        TestResultStatus.append("Pass")
                            # -------------Remittance Email Address(s)--------------------------------------------
                            elif ii2 == 5:
                                time.sleep(1)
                                driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(ii2) + "]/div/input").send_keys(Data1[0] + Data1[3])
                                TestResult.append("Remittance Email Address is entered successfully")
                                TestResultStatus.append("Pass")
                            # -------------Save Button--------------------------------------------
                            elif ii2 == 6:
                                time.sleep(1)
                                driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(ii2) + "]/button").click()
                                TestResult.append("Save Button is clicked successfully")
                                TestResultStatus.append("Pass")
                        TestResult.append("Create new service provider process is working correctly")
                        TestResultStatus.append("Pass")
                    except Exception:
                        pass

                # ---------------------------------------------------------------------------------
                    # --------Saving service provider details in reference sheet------------
                    sheetx1.cell(1, 1).value = Name
                    wbx1.save(locx1)
                elif SpPresentxl == "True":
                    print("Service provider is already present in reference doc. Here is the details")
                    print("First name is: " + Namexl)
                    TestResult.append(
                        "Service provider is already present in reference doc. Here is the details\nFirst name is: " + Namexl)
                    TestResultStatus.append("Pass")

            except Exception:
                TestResult.append("Create new service provider process is not working correctly")
                TestResultStatus.append("Fail")

            # ----------------Verify working of create reimburse button-----------------------------------
            PageName = "Create reimburse button"
            Ptitle1 = "Create Reimburse Client"
            try:
                driver.find_element_by_xpath("//a[text()='Create Reimburse']").click()
                time.sleep(2)
                PageTitle1 = driver.find_element_by_xpath("//h4[text()='Create Reimburse Client']").text
                time.sleep(2)
                driver.find_element_by_xpath("//div[@id='createnewsplatest']/div/div/div[2]/form/div[5]/a").click()
                time.sleep(2)
                assert PageTitle1 in Ptitle1, PageName + " not able to click"
                TestResult.append(PageName + "  is clickable")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not clickable")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # # ---------------------------Verify Pagination clicks-----------------------------
            # PageName = "Service provider listing table"
            # try:
            #     TotalItem = driver.find_element_by_xpath("//div[@id='table_data_wrapper']/div[3]/div[1]").text
            #     print(TotalItem)
            #     substr = "of"
            #     x = TotalItem.split(substr)
            #     string_name = x[0]
            #     TotalItemAfterOf = x[1]
            #     abc = ""
            #     countspace = 0
            #     for element in range(0, len(string_name)):
            #         if string_name[(len(string_name) - 1) - element] == " ":
            #             countspace = countspace + 1
            #             if countspace == 2:
            #                 break
            #         else:
            #             abc = abc + string_name[(len(string_name) - 1) - element]
            #     abc = abc[::-1]
            #     TotalItemBeforeOf = abc
            #     TotalItemAfterOf = TotalItemAfterOf.split(" ")
            #     TotalItemAfterOf = TotalItemAfterOf[1]
            #     TotalItemAfterOf = re.sub('[^A-Za-z0-9]+', '', TotalItemAfterOf)
            #     print(TotalItemAfterOf)
            #
            #     TotalItemAfterOf = int(TotalItemAfterOf)
            #     RecordsPerPage = 50
            #     TotalPages = TotalItemAfterOf / RecordsPerPage
            #     print(TotalPages)
            #     NumberOfPages = math.ceil(float(TotalPages))
            #     print(NumberOfPages)
            #
            #     for i in range(NumberOfPages):
            #         if i < 1:
            #             if i == NumberOfPages - 1:
            #                 TestResult.append("No Pagination found for [ " + str(
            #                     RecordsPerPage) + " ] no. of records under Service Provider Listing table")
            #                 TestResultStatus.append("Pass")
            #                 break
            #         try:
            #             time.sleep(TimeSpeed)
            #             driver.find_element_by_xpath(
            #                 "//div[@id='table_data_paginate']/a[2]").click()
            #             time.sleep(1)
            #             ClickCounter = ClickCounter + 1
            #             TestResult.append("Pagination verified for [ " + str(
            #                 TotalItemAfterOf) + " ] no. of records under Service Provider Listing table")
            #             TestResultStatus.append("Pass")
            #         except Exception:
            #             pass
            #     if i != ClickCounter:
            #         TestResult.append(
            #             "Pagination for [ " + str(RecordsPerPage) + " ] no. of records is not working correctly")
            #         TestResultStatus.append("Fail")
            #     driver.refresh()
            #     try:
            #         WebDriverWait(driver, SHORT_TIMEOUT
            #                       ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            #
            #         WebDriverWait(driver, LONG_TIMEOUT
            #                       ).until(EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            #     except TimeoutException:
            #         pass
            # except Exception as aq:
            #     print(aq)
            #     TestResult.append(PageName + " pagination is not working correctly")
            #     TestResultStatus.append("Fail")
            #     # # ---------------------------------------------------------------------------------
            #
            #     # ----------------Verify working of create franchise button-----------------------------------
            #     PageName = "Create franchise button"
            #     Ptitle1 = "Create Franchise"
            #     try:
            #         driver.find_element_by_xpath("//a[text()='Create Franchise']").click()
            #         time.sleep(2)
            #         PageTitle1 = driver.find_element_by_xpath("//h4[text()='Create Franchise']").text
            #         time.sleep(2)
            #         driver.find_element_by_xpath("//div[@id='createfranchise']/div/div/div/button").click()
            #         time.sleep(2)
            #         assert PageTitle1 in Ptitle1, PageName + " not able to click"
            #         TestResult.append(PageName + "  is clickable")
            #         TestResultStatus.append("Pass")
            #     except Exception:
            #         TestResult.append(PageName + " is not clickable")
            #         TestResultStatus.append("Fail")
            #     print()
            #     time.sleep(TimeSpeed)
            #     # ---------------------------------------------------------------------------------



        except Exception:
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


