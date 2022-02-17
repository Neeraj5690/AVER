import datetime
import math
import re
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from sys import platform

from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("email").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("password").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_InvoiceEntryElements"
  description = "This test scenario is to verify all the Elements present at Invoice Entry of application"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_ElementsPresent"
  global Exe
  Exe="Yes"
  Directory = 'test_InvoiceEntry/'
  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'C:/AVER/AverTest/' + Directory

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      if platform == "linux" or platform == "linux2":
          driver = webdriver.Chrome(executable_path="/home/legion/office 1wayit/AVER/AverTest/chrome/chromedriverLinux")
      elif platform == "win32" or platform == "win64":
          driver = webdriver.Chrome(executable_path="C:/AVER/AverTest/chrome/chromedriver.exe")
      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://averreplica.1wayit.com/login")
      enter_username("admin@averplanning.com")
      enter_password("admin786")
      driver.find_element_by_xpath("//button[@type='submit']").click()

  yield
  if Exe == "Yes":
      ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 2
        LONG_TIMEOUT = 60
        LOADING_ELEMENT_XPATH = "//div[@class='main-loader LoaderImageLogo']"
        try:
            # ---------------------------Verify Invoice Entry icon-----------------------------
            PageName = "Invoice Entry icon"
            Ptitle1 = ""
            try:
                driver.find_element_by_xpath("//i[@class='icon-paragraph-justify3']/parent::a").click()
                time.sleep(2)
                driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[5]/a/i").click()
                time.sleep(2)
                driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[5]/ul/li[1]/a").click()
                time.sleep(2)
                TestResult.append(PageName + " is present in left menu and able to click")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Page title-----------------------------
            PageName = "Page title"
            Ptitle1 = "Invoice Entry "
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//h2[text()='Invoice Entry ']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " (Invoice Entry) is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " (Invoice Entry) is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Total Count tab -------------------------------------
            PageName = "Total Count tab"
            Ptitle1 = "Total Count"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//p[text()='Total Count:- ']").text
                print(PageTitle1)
                if ":-\n" in PageTitle1:
                    PageTitle1 = PageTitle1.split(":-\n")
                    PageTitle1 = PageTitle1[0]
                    print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Total amount tab -------------------------------------
            PageName = "Total amount tab"
            Ptitle1 = "Total Amount"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//p[text()='Total Amount:- ']").text
                print(PageTitle1)
                if ":-\n" in PageTitle1:
                    PageTitle1 = PageTitle1.split(":-\n")
                    PageTitle1 = PageTitle1[0]
                    print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of XERO CSV button-------------------------------------
            PageName = "XERO CSV button"
            Ptitle1 = "XERO CSV"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//button[text()='XERO CSV']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Create New button-------------------------------------
            PageName = "Create New button"
            Ptitle1 = "Create New"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//a[text()='Create New']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Back button-------------------------------------
            PageName = "Back button"
            Ptitle1 = "Back"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//a[text()='Back']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Report button-------------------------------------
            PageName = "Report button"
            Ptitle1 = "Report"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//button[text()='Report']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Action button-------------------------------------
            PageName = "Action button"
            Ptitle1 = "actionBtn"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//button[@id='actionBtn']").get_attribute('id')
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of From date field-------------------------------------
            PageName = "From date field"
            Ptitle1 = "search_date_from"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//input[@id='search_date_from']").send_keys("11-02-2022")
                PageTitle1 = driver.find_element_by_xpath(
                    "//input[@id='search_date_from']").get_attribute('id')
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present and user able to to send inputs")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # -----------------------------------------------------------------------------------------------
            # ---------------------------Verify Presence of To date field-------------------------------------
            PageName = "To date field"
            Ptitle1 = "search_date_to"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//input[@id='search_date_to']").send_keys("17-02-2022")
                PageTitle1 = driver.find_element_by_xpath(
                    "//input[@id='search_date_to']").get_attribute('id')
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present and user able to to send inputs")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Limit dropdown-----------------------------
            PageName = "Limit dropdown"
            Ptitle1 = "50"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//select[@id='invoice_entry_listing_limit']/option[1]").text
                print(PageTitle1)

                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            PageTitle1 = driver.find_element_by_xpath(
                "//select[@id='invoice_entry_listing_limit']").click()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify 50 dropdown value-----------------------------
            PageName = "50 dropdown value"
            Ptitle1 = "50"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//select[@id='invoice_entry_listing_limit']/option[1]").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + " is present in limit dropdown")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present in limit dropdown")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify 100 dropdown value-----------------------------
            PageName = "100 dropdown value"
            Ptitle1 = "100"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//select[@id='invoice_entry_listing_limit']/option[2]").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + " is present in limit dropdown")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present in limit dropdown")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify 300 dropdown value-----------------------------
            PageName = "300 dropdown value"
            Ptitle1 = "300"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//select[@id='invoice_entry_listing_limit']/option[3]").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + " is present in limit dropdown")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present in limit dropdown")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify 500 dropdown value-----------------------------
            PageName = "500 dropdown value"
            Ptitle1 = "500"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//select[@id='invoice_entry_listing_limit']/option[4]").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + " is present in limit dropdown")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present in limit dropdown")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Clear Filter button-----------------------------
            PageName = "Clear Filter button"
            Ptitle1 = "Clear Filter"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//a[text()='Clear Filter']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Exclude Paid Invoices check box-----------------------------
            PageName = "Exclude Paid Invoices check box"
            Ptitle1 = "form-check-label"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//div[@class='row invce_dta']/div[1]/div[1]/div[3]/label").get_attribute('class')
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # # ---------------------------Verify Pagination clicks for Invoice entry table-----------------------------
            # PageName = "Invoice entry table"
            # try:
            #     NumberOfPages = driver.find_element_by_xpath(
            #         "//tbody[@id='invoiceEntryListingAjaxView']/tr[last()]/td/nav/ul/li[14]/a").text
            #     NumberOfPages = int(NumberOfPages)
            #     print(NumberOfPages)
            #     for i in range(NumberOfPages):
            #         if i == NumberOfPages - 1:
            #             TestResult.append("Pagination for " + str(PageName) + " is successfully verified")
            #             TestResultStatus.append("Pass")
            #             break
            #         driver.find_element_by_xpath("//tbody[@id='invoiceEntryListingAjaxView']/tr[last()]/td/nav/ul/li[last()]/a").click()
            #         time.sleep(1)
            #     if i != NumberOfPages - 1:
            #         TestResult.append(
            #             (
            #                 "Pagination for " + str(PageName) + " is not working correctly"))
            #         TestResultStatus.append("Fail")
            #     driver.refresh()
            #     try:
            #         WebDriverWait(driver, SHORT_TIMEOUT
            #                       ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            #
            #         WebDriverWait(driver, LONG_TIMEOUT
            #                       ).until(EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
            #     except TimeoutException:
            #         pass
            # except Exception as aq:
            #     print(aq)
            #     TestResult.append(PageName + " is not present")
            #     TestResultStatus.append("Fail")
            #     # # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of No. of invoices selected text-----------------------------
            PageName = "No. of invoices selected text"
            Ptitle1 = "No. of invoices selected "
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//label[text()='No. of invoices selected ']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Total Amount  $ text-----------------------------
            PageName = "Total Amount  $ text"
            Ptitle1 = "Total Amount  $ "
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//label[text()='Total Amount  $ ']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Today's Total Setup & Monthly Fee text-----------------------------
            PageName = "Today's Total Setup & Monthly Fee text"
            Ptitle1 = "Today's Total Setup & Monthly Fee"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//div[@class='content']/div[2]/div[1]/div[3]/div[2]/div[1]/label").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Setup & Monthly Fee Amount text-----------------------------
            PageName = "Setup & Monthly Fee Amount text"
            Ptitle1 = "Setup & Monthly Fee Amount"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//div[@class='content']/div[2]/div[1]/div[3]/div[2]/div[2]/label").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Today's Total Count text-----------------------------
            PageName = "Today's Total Count text"
            Ptitle1 = "Today's Total Count"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//div[@class='content']/div[2]/div[1]/div[3]/div[3]/div[1]/label").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Today's Total Amount text-----------------------------
            PageName = "Today's Total Amount text"
            Ptitle1 = "Today's Total Amount"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//div[@class='content']/div[2]/div[1]/div[3]/div[3]/div[2]/label").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Draft table header-----------------------------
            PageName = "Draft table header"
            Ptitle1 = "Draft"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//div[@class='content']/div[2]/div[1]/div[4]/div[1]/div/h2").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------


        except Exception as err:
            print(err)
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


