import datetime
import math
import re
import time
import openpyxl
from datetime import datetime,date
import datetime as datetime
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from sys import platform
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import pyperclip
import random
import string


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("email").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("password").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_ElementsWorking"
  description = "This test scenario is to verify all the Working of Elements at Invoice Entry page"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_ElementsWorking"
  global Exe
  Exe="Yes"
  Directory = 'test_InvoiceEntry/'
  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'D:/AVER/AverTest/' + Directory

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      if platform == "linux" or platform == "linux2":
          driver = webdriver.Chrome(executable_path="/home/legion/office 1wayit/AVER/AverTest/chrome/chromedriverLinux1")
      elif platform == "win32" or platform == "win64":
          driver = webdriver.Chrome(executable_path="D:/AVER/AverTest/chrome/chromedriver.exe")

      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://averreplica.1wayit.com/login")
      enter_username("admin@averplanning.com")
      enter_password("admin786")
      driver.find_element_by_xpath("//button[@type='submit']").click()

  yield
  if Exe == "Yes":
      ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 3
        LONG_TIMEOUT = 60
        LOADING_ELEMENT_XPATH = "//div[@class='main-loader LoaderImageLogo']"
        try:
            # ---------------------------Verify Invoice Entry icon click-----------------------------
            PageName = "Invoice Entry icon"
            Ptitle1 = ""
            try:
                driver.find_element_by_xpath("//i[@class='icon-paragraph-justify3']/parent::a").click()
                time.sleep(2)
                driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[5]/a/i").click()
                time.sleep(2)
                driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[5]/ul/li[1]/a").click()
                time.sleep(2)
                TestResult.append(PageName + " is present in left menu and able to click")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify working of Back button on Invoice entry page -----------------------------
            PageName = "Back button"
            Ptitle1 = "Rae"
            try:
                driver.find_element_by_xpath("//a[text()='Back']").click()
                time.sleep(2)
                PageTitle1 = driver.find_element_by_xpath("//div[@class='hed_wth_srch']/h2").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is clickable")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not clickable")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ----------------Verify Invoice entry icon click after verifying back--------
            PageName = "Invoice entry icon"
            Ptitle1 = ""
            try:
                driver.find_element_by_xpath("//i[@class='icon-paragraph-justify3']/parent::a").click()
                time.sleep(2)
                driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[5]/a/i").click()
                time.sleep(2)
                driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[5]/ul/li[1]/a").click()
                time.sleep(2)
                TestResult.append(PageName + "  is opened again after verifying back button")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not opened again after verifying back button")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify working of XERO CSV button-----------------------------
            PageName = "XERO CSV button"
            Ptitle1 = "Generate Invoice Report"
            try:
                driver.find_element_by_xpath("//button[text()='XERO CSV']").click()
                time.sleep(2)
                PageTitle1 = driver.find_element_by_xpath("//div[@id='xeroReport']/div/div/div[1]/h4").text
                time.sleep(2)
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is clickable")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not clickable")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            driver.find_element_by_xpath("// div[ @ id = 'xeroReport'] / div / div / div[1] / button").click()
            #---------------------------------------------------------------------------------

            # ---------------------------Verify working of Create new button on Invoice entry page-----------------------------
            PageName = "Create new button"
            Ptitle1 = "Create New Invoice"
            try:
                driver.find_element_by_xpath("//a[text()='Create New']").click()
                time.sleep(2)
                PageTitle1 = driver.find_element_by_xpath("//h2[text()='Create New Invoice']").text
                time.sleep(2)
                print(PageTitle1)

                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is clickable")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not clickable")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # # ---------------------------Verify working of Create Reimburse Client button on Create new page-----------------------------
            # try:
            #     l = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
            #     random.shuffle(l)
            #     if l[0] == 0:
            #         pos = random.choice(range(1, len(l)))
            #         l[0], l[pos] = l[pos], l[0]
            #     AccountBsb = ''.join(map(str, l[0:6]))
            #     print(AccountBsb)
            #     today = date.today()
            #     D1 = today.strftime("%d-%m-%Y")
            #
            #     driver.find_element_by_xpath("//a[@title='Create Reimburse Client']").click()
            #     time.sleep(2)
            #     for rc in range(5):
            #         letters = string.ascii_lowercase
            #         returna = ''.join(random.choice(letters) for i in range(5))
            #         Name = returna
            #     print(Name)
            #     Data = [Name, D1, AccountBsb,"@test.com"]
            #     for ii in range(1, 6):
            #         driver.find_element_by_xpath("//div[@id='createnewsplatest']/div/div/div[2]/form/div["+str(ii)+"]")
            #         if ii==1:
            #             for i1 in range(1,4):
            #                 driver.find_element_by_xpath("//div[@id='createnewsplatest']/div/div/div[2]/form/div[1]/div["+str(i1)+"]/div")
            #                 # -------------Name Field--------------------------------------------
            #                 if i1 == 1:
            #                     time.sleep(1)
            #                     driver.find_element_by_xpath(
            #                         "//div[@id='createnewsplatest']/div/div/div[2]/form/div[1]/div["+str(i1)+"]/div/input").send_keys(Data[0])
            #
            #                 # -------------Status Dropdown--------------------------------------------
            #                 elif i1 == 2:
            #                     time.sleep(1)
            #                     select = Select(driver.find_element_by_xpath(
            #                         "//div[@id='createnewsplatest']/div/div/div[2]/form/div[1]/div["+str(i1)+"]/div/select"))
            #                     select.select_by_visible_text("Active")
            #                 # -------------Set Up Date Field--------------------------------------------
            #                 elif i1 == 3:
            #                     time.sleep(1)
            #                     driver.find_element_by_xpath(
            #                         "//div[@id='createnewsplatest']/div/div/div[2]/form/div[1]/div[" + str(
            #                             i1) + "]/div/input").clear()
            #                     time.sleep(1)
            #                     driver.find_element_by_xpath(
            #                         "//div[@id='createnewsplatest']/div/div/div[2]/form/div[1]/div[" + str(
            #                             i1) + "]/div/input").send_keys(Data[1])
            #                     ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
            #                     time.sleep(2)
            #         # -------------Account Name Field--------------------------------------------
            #         elif ii == 2:
            #             for ii1 in range(1,4):
            #                 if ii1 ==1:
            #                     time.sleep(1)
            #                     driver.find_element_by_xpath("//div[@id='createnewsplatest']/div/div/div[2]/form/div[2]/div["+str(ii1)+"]/div/input").send_keys(
            #                         Data[0])
            #                 # -------------Account BSB Field--------------------------------------------
            #                 elif ii1==2:
            #                     time.sleep(1)
            #                     driver.find_element_by_xpath(
            #                         "//div[@id='createnewsplatest']/div/div/div[2]/form/div[2]/div[" + str(
            #                             ii1) + "]/div/input").send_keys(
            #                         Data[2])
            #                 # -------------Account Number Field--------------------------------------------
            #                 elif ii1 == 3:
            #                     time.sleep(1)
            #                     driver.find_element_by_xpath(
            #                         "//div[@id='createnewsplatest']/div/div/div[2]/form/div[2]/div[" + str(
            #                             ii1) + "]/div/input").send_keys(
            #                         Data[2])
            #         # -------------Send Remittance Email Check Box--------------------------------------------
            #         elif ii == 3:
            #             print(ii)
            #             time.sleep(1)
            #             driver.find_element_by_xpath(
            #                 "//div[@id='createnewsplatest']/div/div/div[2]/form/div[" + str(
            #                     ii) + "]/div/div/label/input").click()
            #             time.sleep(1)
            #             driver.find_element_by_xpath(
            #                 "//div[@id='createnewsplatest']/div/div/div[2]/form/div[" + str(
            #                     ii) + "]/div/div/label/input").click()
            #             print(ii)
            #         # -------------Remittance Email Address(s)--------------------------------------------
            #         elif ii == 4:
            #             print(ii)
            #             time.sleep(1)
            #             driver.find_element_by_xpath(
            #                 "//div[@id='createnewsplatest']/div/div/div[2]/form/div["+str(ii)+"]/div/input").send_keys(Data[0] + Data[3])
            #             print(ii)
            #         # -------------Save Button--------------------------------------------
            #         elif ii == 5:
            #             print(ii)
            #             time.sleep(1)
            #             driver.find_element_by_xpath(
            #                 "//div[@id='createnewsplatest']/div/div/div[2]/form/div[" + str(ii) + "]/button").click()
            #             print(ii)
            #     TestResult.append("Create Reimburse Client process working is correctly")
            #     TestResultStatus.append("Pass")
            # except Exception:
            #     TestResult.append("Create Reimburse Client process is not working")
            #     TestResultStatus.append("Fail")
            #
            # # ---------------------------------------------------------------------------------

            # # ---------------------------Verify working of Create Reimburse Client button on Create new page-----------------------------
            # l = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
            # random.shuffle(l)
            # if l[0] == 0:
            #     pos = random.choice(range(1, len(l)))
            #     l[0], l[pos] = l[pos], l[0]
            # Number = ''.join(map(str, l[0:6]))
            # print(Number)
            #
            # today = date.today()
            # D1 = today.strftime("%d-%m-%Y")
            #
            # driver.find_element_by_xpath("//a[@title='Create Service Provider']").click()
            # time.sleep(2)
            #
            # for rc in range(5):
            #     letters = string.ascii_lowercase
            #     returna = ''.join(random.choice(letters) for i in range(5))
            #     Name = returna
            # print(Name)
            # Data1 = [Name, D1, Number, "@test.com","2132435465"]
            # for ii2 in range(1, 7):
            #     driver.find_element_by_xpath(
            #         "//div[@id='createnewsp']/div/div/div[2]/form/div[" + str(ii2) + "]")
            #     if ii2 == 1:
            #         for i2 in range(1, 4):
            #             driver.find_element_by_xpath(
            #                 "//div[@id='createnewsp']/div/div/div[2]/form/div[1]/div["+str(i2)+"]")
            #             # -------------Name Field--------------------------------------------
            #             if i2 == 1:
            #                 time.sleep(1)
            #                 driver.find_element_by_xpath(
            #                     "//div[@id='createnewsp']/div/div/div[2]/form/div[1]/div["+str(i2)+"]/div/input").send_keys(Data[0])
            #
            #             # -------------Status Dropdown--------------------------------------------
            #             elif i2 == 2:
            #                 time.sleep(1)
            #                 select = Select(driver.find_element_by_xpath(
            #                     "//div[@id='createnewsp']/div/div/div[2]/form/div[1]/div["+str(i2)+"]/div/select"))
            #                 select.select_by_visible_text("Active")
            #             # -------------Set Up Date Field--------------------------------------------
            #             elif i2 == 3:
            #                 time.sleep(1)
            #                 driver.find_element_by_xpath(
            #                     "//div[@id='createnewsp']/div/div/div[2]/form/div[1]/div["+str(i2)+"]/div/input").clear()
            #                 time.sleep(1)
            #                 driver.find_element_by_xpath(
            #                     "//div[@id='createnewsp']/div/div/div[2]/form/div[1]/div[" + str(i2) + "]/div/input").send_keys(Data[1])
            #                 ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
            #                 time.sleep(2)
            #             # -------------ABN Field--------------------------------------------
            #             elif i2 == 4:
            #                 time.sleep(1)
            #                 driver.find_element_by_xpath(
            #                     "//div[@id='createnewsp']/div/div/div[2]/form/div[1]/div[" + str(
            #                         i2) + "]/div/input").send_keys(Data1[2])
            #             # -------------Franchise Dropdown--------------------------------------------
            #             elif i2 == 5:
            #                 time.sleep(1)
            #                 select = Select(driver.find_element_by_xpath(
            #                     "//div[@id='createnewsp']/div/div/div[2]/form/div[1]/div["+str(i2)+"]/div/select"))
            #                 select.select_by_visible_text("Yes")
            #             # -------------Related Franchise Dropdown--------------------------------------------
            #             elif i2 == 6:
            #                 time.sleep(1)
            #                 select = Select(driver.find_element_by_xpath(
            #                     "//div[@id='createnewsp']/div/div/div[2]/form/div[1]/div[" + str(i2) + "]/div/select"))
            #                 select.select_by_index(0)
            #             # -------------Service Type Dropdown--------------------------------------------
            #             elif i2 == 7:
            #                 time.sleep(1)
            #                 select = Select(driver.find_element_by_xpath(
            #                     "//div[@id='createnewsp']/div/div/div[2]/form/div[1]/div[" + str(i2) + "]/div/span/select"))
            #                 select.select_by_index(3)
            #     # -------------Office Number Field--------------------------------------------
            #     elif ii2 == 2:
            #         for ii3 in range(1, 10):
            #             if ii3 == 1:
            #                 time.sleep(1)
            #                 driver.find_element_by_xpath(
            #                     "//div[@id='createnewsp']/div/div/div[2]/form/div[2]/div["+str(ii3)+"]/div/input").send_keys(
            #                     Data1[2])
            #             # -------------Mobile Number Field--------------------------------------------
            #             elif ii3 == 2:
            #                 time.sleep(1)
            #                 driver.find_element_by_xpath(
            #                     "//div[@id='createnewsp']/div/div/div[2]/form/div[2]/div["+str(ii3)+"]/div/input").send_keys(
            #                     Data1[4])
            #             # -------------Admin Email Address Field--------------------------------------------
            #             elif ii3 == 3:
            #                 time.sleep(1)
            #                 driver.find_element_by_xpath(
            #                     "//div[@id='createnewsp']/div/div/div[2]/form/div[2]/div["+str(ii3)+"]/div/input").send_keys(Data1[0] + Data1[3])
            #             # -------------Address Country Field--------------------------------------------
            #             elif ii3 == 4:
            #                 time.sleep(1)
            #                 select = Select(driver.find_element_by_xpath(
            #                     "//div[@id='createnewsp']/div/div/div[2]/form/div[2]/div["+str(ii3)+"]/div/select"))
            #                 select.select_by_index(4)
            #             # -------------Address State Field--------------------------------------------
            #             elif ii3 == 5:
            #                 time.sleep(1)
            #                 select = Select(driver.find_element_by_xpath(
            #                     "//div[@id='createnewsp']/div/div/div[2]/form/div[2]/div[" + str(ii3) + "]/div/select"))
            #                 select.select_by_index(3)
            #             # -------------Address City Field--------------------------------------------
            #             elif ii3 == 6:
            #                 time.sleep(1)
            #                 select = Select(driver.find_element_by_xpath(
            #                     "//div[@id='createnewsp']/div/div/div[2]/form/div[2]/div[" + str(ii3) + "]/div/select"))
            #                 select.select_by_index(2)
            #             # -------------Address Street Field--------------------------------------------
            #             elif ii3 == 7:
            #                 time.sleep(1)
            #                 driver.find_element_by_xpath(
            #                     "//div[@id='createnewsp']/div/div/div[2]/form/div[2]/div[" + str(ii3) + "]/div/input").send_keys(Data[0])
            #             # -------------Address Street Field--------------------------------------------
            #             elif ii3 == 8:
            #                 time.sleep(1)
            #                 driver.find_element_by_xpath("//div[@id='createnewsp']/div/div/div[2]/form/div[2]/div[" + str(
            #                         ii3) + "]/div/input").send_keys(Data1[2])
            #             # -------------Address Street Field--------------------------------------------
            #             elif ii3 == 9:
            #                 time.sleep(1)
            #                 driver.find_element_by_xpath(
            #                     "//div[@id='createnewsp']/div/div/div[2]/form/div[2]/div[" + str(
            #                         ii3) + "]/div/input").send_keys(Data1[0])
            #     # -------------Account Name Field--------------------------------------------
            #     elif ii2 == 3:
            #         for cc in range(1,4):
            #             driver.find_element_by_xpath("//div[@id='createnewsp']/div/div/div[2]/form/div[3]/div["+str(cc)+"]")
            #             if cc==1:
            #                 driver.find_element_by_xpath("//div[@id='createnewsp']/div/div/div[2]/form/div[3]/div["+str(cc)+"]/div/input").send_keys(Data1[0])
            #             elif cc==2:
            #                 driver.find_element_by_xpath("//div[@id='createnewsp']/div/div/div[2]/form/div[3]/div["+str(cc)+"]/div/input").send_keys(Data1[2])
            #     # # -------------Remittance Email Address(s)--------------------------------------------
            #     # elif ii2 == 4:
            #     #     print(ii)
            #     #     time.sleep(1)
            #     #     driver.find_element_by_xpath(
            #     #         "//div[@id='createnewsplatest']/div/div/div[2]/form/div[" + str(
            #     #             ii) + "]/div/input").send_keys(Data[0] + Data[3])
            #     #     print(ii)
            #     # # -------------Save Button--------------------------------------------
            #     # elif ii2 == 5:
            #     #     print(ii)
            #     #     time.sleep(1)
            #     #     driver.find_element_by_xpath(
            #     #         "//div[@id='createnewsplatest']/div/div/div[2]/form/div[" + str(
            #     #             ii) + "]/button").click()
            #     #     print(ii)
            #
            # # ---------------------------------------------------------------------------------

            # ---------------------------Verify working of Create New Invoice process-----------------------------
            try:
                ClientPresentxl = "True"
                xcelFileName = "RefData"
                locx1 = (path + 'ReferenceData/' + xcelFileName + '.xlsx')
                wbx1 = openpyxl.load_workbook(locx1)
                sheetx1 = wbx1.active
                for i_ref in range(1, 10):
                    if sheetx1.cell(i_ref, 1).value != None:
                        FirstNamexl=sheetx1.cell(i_ref, 1).value
                        LastNamexl = sheetx1.cell(i_ref, 2).value
                        PlanEndDate = sheetx1.cell(i_ref, 3).value
                        break

                    else:
                        ClientPresentxl="False"
                        pass

                if ClientPresentxl=="False":
                    print("Client is not present in reference sheet, Invoice creation process terminated")
                    TestResult.append("Client is not present in reference sheet, Invoice creation process terminated\nFirst run the job for create client")
                    TestResultStatus.append("Fail")
                    driver.close()

                PageName = "Create New Invoice process"
                l = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
                random.shuffle(l)
                if l[0] == 0:
                    pos = random.choice(range(1, len(l)))
                    l[0], l[pos] = l[pos], l[0]
                InvoiceNumber = ''.join(map(str, l[0:4]))
                print(InvoiceNumber)
                try:
                    driver.find_element_by_xpath("//input[@name='search_client_name']").send_keys(FirstNamexl)
                    time.sleep(1)
                    ActionChains(driver).key_down(Keys.DOWN).key_up(Keys.DOWN).perform()
                    time.sleep(1)
                    ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                    time.sleep(1)

                    driver.find_element_by_xpath("//input[@name='provider_invoice_number']").send_keys(InvoiceNumber)
                    time.sleep(2)
                    driver.find_element_by_xpath("//input[@name='search_service_provider_name']").send_keys("Blossom")
                    time.sleep(1)
                    ActionChains(driver).key_down(Keys.DOWN).key_up(Keys.DOWN).perform()
                    time.sleep(1)
                    ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                    time.sleep(1)

                    try:
                        InvoiceError = driver.find_element_by_xpath("//h2[text()='Error']").is_displayed()
                        print(InvoiceError)
                        if InvoiceError == True:
                            driver.find_element_by_xpath("//div[@class='jq-toast-wrap top-right']/div/button").click()
                            time.sleep(2)
                            driver.find_element_by_xpath("//input[@name='provider_invoice_number']").clear()
                            time.sleep(2)
                            driver.find_element_by_xpath("//input[@name='provider_invoice_number']").send_keys(InvoiceNumber)
                        elif InvoiceError == False:
                            pass
                    except Exception:
                        pass

                    today = date.today()
                    D1 = today.strftime("%d-%m-%Y")
                    driver.find_element_by_xpath("//input[@name='provider_invoice_date']").send_keys(D1)
                    time.sleep(1)
                    ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                    time.sleep(1)

                    select = Select(driver.find_element_by_xpath("//select[@id='approvedByClientText']"))
                    select.select_by_index(0)

                    ActionChains(driver).key_down(Keys.PAGE_DOWN).key_up(Keys.PAGE_DOWN).perform()
                    time.sleep(1)

                    time.sleep(2)
                    driver.find_element_by_xpath("//div[@id='search_client_data']/div[2]/div/div[1]/h3/div").click()
                    time.sleep(2)
                    driver.find_element_by_xpath("//div[@id='search_client_data']/div[2]/div/div[1]/div/textarea").send_keys("Test Notes")
                    time.sleep(2)

                    driver.find_element_by_xpath("//div[@id='search_client_data']/div[2]/div/div[3]/h3/div").click()
                    time.sleep(2)
                    driver.find_element_by_xpath("//div[@id='search_client_data']/div[2]/div/div[3]/div/div/div[1]/label/input").click()
                    time.sleep(2)

                    ActionChains(driver).key_down(Keys.PAGE_DOWN).key_up(Keys.PAGE_DOWN).perform()
                    time.sleep(1)
                    try:
                        PlanEndDate = PlanEndDate.strftime("%d-%m-%Y")
                    except Exception:
                        print(PlanEndDate)
                    print(PlanEndDate)
                    driver.find_element_by_xpath("//input[@name='service_detail[2][delivered_date]']").send_keys(PlanEndDate)
                    time.sleep(1)
                    ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                    time.sleep(1)
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until(
                            EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                    time.sleep(2)

                    select = Select(driver.find_element_by_xpath(
                                            "//select[@name='service_detail[2][category]']"))
                    select.select_by_index(1)
                    time.sleep(1)

                    forZeroBreak=0
                    for cs in range(1,10):
                        print("------------cs is "+str(cs))
                        try:
                            select = Select(driver.find_element_by_xpath(
                                "//select[@name='service_detail[2][category]']"))
                            select.select_by_index(cs)
                        except Exception:
                            if forZeroBreak == 0:
                                print("No Category with some amount found. Terminating the invoice creation process")
                                TestResult.append("No Category with some amount found. Terminating the invoice creation process")
                                TestResultStatus.append("Fail")
                                driver.close()
                        time.sleep(2)
                        RemAmount = driver.find_element_by_xpath("//div[@id='category_remaining_div_2']/span").text
                        print(RemAmount)
                        for char in RemAmount:
                            RemAmount = RemAmount.replace(',', "")
                        temp = re.findall(r'\d+', RemAmount)
                        res = list(map(int, temp))
                        try:
                            amountFound = res[0]
                            amountFound = float(amountFound)
                        except Exception:
                            amountFound = 0.0

                        print(amountFound)
                        if amountFound>0.0:
                            forZeroBreak=1
                            break
                        print(amountFound)

                    if forZeroBreak==1:
                        InvoiceRate = (amountFound / 100) * 10
                        print(InvoiceRate)

                    select = Select(driver.find_element_by_xpath(
                        "//select[@name='service_detail[2][line_item]']"))
                    select.select_by_index(1)
                    time.sleep(1)

                    driver.find_element_by_xpath("//input[@name='service_detail[2][qty]']").send_keys("1")
                    time.sleep(2)

                    driver.find_element_by_xpath("//input[@name='service_detail[2][price]']").send_keys(InvoiceRate)
                    time.sleep(2)

                    driver.find_element_by_xpath("//button[@id='submitButton']").click()
                    TestResult.append("Create New Invoice process working is correctly")
                    TestResultStatus.append("Pass")
                except Exception:
                    TestResult.append("Create New Invoice process is not working")
                    TestResultStatus.append("Fail")
                print()
                time.sleep(TimeSpeed)
            except Exception:
                pass
            # ---------------------------------------------------------------------------------

        except Exception as err:
            print(err)
            TestResult.append("Invoice entry is not working correctly. Below error found\n"+str(err))
            TestResultStatus.append("Fail")
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


