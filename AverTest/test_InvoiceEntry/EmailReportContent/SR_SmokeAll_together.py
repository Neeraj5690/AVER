import os
import smtplib
import ssl
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl
import datetime
import pytest
from sys import platform

@pytest.mark.smoke
def test_ReportSendSmokeAll():
    print()
    DirectoryName = []
    PDFName1=[]
    TestName=[]
    TestDescription = []
    TestStatus = []
    SendStatus = []
    AttachmntAdded= []
    Directory= 'test_InvoiceEntry/'
    if platform == "linux" or platform == "linux2":
        PDFpath = '/home/legion/office 1wayit/AVER/AverTest/' + Directory
    elif platform == "win32" or platform == "win64":
        PDFpath = 'D:/AVER/AverTest/' + Directory

    #-------------------To read content to send in e-Mail--------------------
    ExcelFileName = "FileName"
    loc = (PDFpath+'PDFFileNameData/' + ExcelFileName + '.xlsx')
    wb=openpyxl.load_workbook(loc)
    sheet = wb.active
    for i in range(1, 100):
        if sheet.cell(i, 1).value == None:
            break
        else:
            PDFName1.append(sheet.cell(i, 2).value)
            DirectoryName.append(sheet.cell(i, 3).value)
            TestName.append(sheet.cell(i, 1).value)
            TestDescription.append(sheet.cell(i, 4).value)
            TestStatus.append(sheet.cell(i, 5).value)
            SendStatus.append(sheet.cell(i, 6).value)

        B = ""
        for io in range(len(TestName)):
            try:
                B = B + "<br /><br />"+str(io+1)+") " + "".join(TestName[io])+" => "+"".join(TestDescription[io])+" => "+"".join(TestStatus[io])
            except Exception:
                print("No attachment details to add in email description")
    #print(B)

    ##############################################################
    html = '''
        <html>
            <body>
                <p>Hi Team <br />Here is the test summary report of Test Suite 5 ( Invoice Entry ) <br />Below test scenarios are covered </p>
                <p></p>
                <p>'''+B+'''</p
                <p></p>
                <img src='cid:myimageid' width="500" align="center">
                <p>Please find attached PDFs of test scenarios results<br />Note: Attachments are only for FAILED test cases<br /></p>
                <p>Many Thanks <br/>Rae automation team</p>
            </body>
        </html>
        '''
    ##############################################################

    def attach_file_to_email(msg, filename, extra_headers=None):
        with open(filename, "rb") as f:
            file_attachment = MIMEApplication(f.read())
        file_attachment.add_header(
            "Content-Disposition",
            f"attachment; filename= {filename}",
        )
        if extra_headers is not None:
            for name, value in extra_headers.items():
                file_attachment.add_header(name, value)
        msg.attach(file_attachment)


    email_from = 'Rae CRM Test Automation Team'
    email_to = ['avneet.kumar@bitsinglass.com', 'sumreet.kaur@bitsinglass.com', 'reilly@averplanmanagers.com.au',
                'scott@scnsa.com.au']
    #email_to =['neeraj1wayitsol@gmail.com','avneet.kumar@bitsinglass.com','sumreet.kaur@bitsinglass.com','gagandeep.singh@bitsinglass.com','reilly@averplanmanagers.com.au','scott@scnsa.com.au']
    #email_to = ['neeraj1wayitsol@gmail.com', 'avneet.kumar@bitsinglass.com', 'sumreet.kaur@bitsinglass.com', 'gagandeep.singh@bitsinglass.com']
    #email_to =['gagandeep.singh@bitsinglass.com','neeraj1wayitsol@gmail.com']

    time_change = datetime.timedelta(hours=5)
    new_time = datetime.datetime.now() + time_change
    ctDate = new_time.strftime("%d %B %Y")
    ct = new_time.strftime("%p")
    if ct == "PM":
        ct = "E"
    elif ct == "AM":
        ct = "M"

    SenderEmail = "Raeautomationbig@gmail.com"
    RandmStr = "Raeautomationbig@786"
    msg = MIMEMultipart()
    msg['Subject']=ctDate+" ["+ct+"]"+'-[Test Suite 5 (Invoice Entry)]-Rae CRM Test Automation Report'
    msg['From'] = email_from
    msg['To'] = ','.join(email_to)
    msg.attach(MIMEText(html, "html"))

    #------------------------------------------------------------------------
    try:
        attach_file_to_email(msg, PDFpath+'TestPieResult.png',
                         {'Content-ID': '<myimageid>'})
    except Exception:
        print("No Pie File to attach")
    # ------------------To add attachments in the report email--------------
    i=0
    for file in PDFName1:
        print()
        try:
            #print(file)
            if SendStatus[i]=="Send Only when Fail=Yes" and  TestStatus[i]=="Fail":
                    attach_file_to_email(msg,PDFpath+PDFName1[i])
                    AttachmntAdded.append("Yes")
            if SendStatus[i] == "Send Only when Fail=No":
                    print("Inside Send Only when Fail=No")
                    attach_file_to_email(msg, PDFpath + PDFName1[i])
                    AttachmntAdded.append("Yes")
        except Exception as e1:
            print("No Attachment found to Add")
            #print(e1)
        i = i + 1
    #-----------------------------------------------------------------------

    # ------------------------To attach all in e-Mail-----------------------
    email_string = msg.as_string()
    context = ssl.create_default_context()
    # -----------------------------------------------------------------------

    # ----------------------------SMTP setup--------------------------------
    server=smtplib.SMTP_SSL('smtp.gmail.com',465)
    server.login(SenderEmail,RandmStr)
    #-----------------------------------------------------------------------

    #---------------------------------Sending email-------------------------
    for io1 in range(len(AttachmntAdded)):
        if AttachmntAdded[io1] == "Yes":
            print("Inside AttachmntAdded=Yes ")
            server.sendmail(email_from, email_to, email_string)
            print("Test Report sent")
            break
    #-----------------------------------------------------------------------

    #-----------------To delete pdf report files----------------------------
    ii=0
    for ii in range(0,len(PDFName1)):
        print()
        try:
            os.remove(PDFpath+PDFName1[ii])
        except Exception:
            print("No Attachment found to delete")
    try:
        os.remove(PDFpath+'TestPieResult.png')
    except Exception:
        print("No Attachment found to delete")
    #-----------------------------------------------------------------------
    server.quit()
