import datetime
import math
import re
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("email").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("password").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_CommunicationLogElements"
  description = "This test scenario is to verify all the Elements present at Communication log of application"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_ElementsPresent"
  global Exe
  Exe="Yes"
  Directory = 'test_CommunicationLog/'
  path = 'C:/AVER/AverTest/' + Directory

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      driver=webdriver.Chrome(executable_path="C:/AVER/AverTest/chrome/chromedriver.exe")
      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://averreplica.1wayit.com/login")
      enter_username("admin@averplanning.com")
      enter_password("admin786")
      driver.find_element_by_xpath("//button[@type='submit']").click()

  yield
  if Exe == "Yes":
      ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      #driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        SHORT_TIMEOUT = 2
        LONG_TIMEOUT = 60
        LOADING_ELEMENT_XPATH = "//div[@class='main-loader LoaderImageLogo']"
        try:
            #---------------------------Verify Communication Log elements-----------------------------
            PageName = "Communication log icon"
            Ptitle1 = ""
            try:
                driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[2]/a/i").click()
                time.sleep(2)
                TestResult.append(PageName + "  is present in left menu and able to click")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            #---------------------------------------------------------------------------------

            #---------------------------Verify Page title-----------------------------
            PageName = "Page title"
            Ptitle1 = "Add new communication log"
            try:
                PageTitle1 = driver.find_element_by_xpath("//div[@class='content yellow_color']/div[1]/div/form/div/div/div/label").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + "(Add new communication log)  is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + "(Add new communication log)  is not present")
                TestResultStatus.append("Fail")
            print()
            #---------------------------------------------------------------------------------

            #---------------------------Verify Presence of select entry dropdown-----------------------------
            PageName = "Select entry dropdown"
            Ptitle1 = "Select New Entry"
            try:
                PageTitle1 = driver.find_element_by_xpath("//div[@class='content yellow_color']/div[1]/div/form/div/div/div/select/option").text
                print(PageTitle1)

                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + "  is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            PageTitle1 = driver.find_element_by_xpath(
                "//div[@class='content yellow_color']/div[1]/div/form/div/div/div/select").click()
            #---------------------------------------------------------------------------------
            #
            # ---------------------------Verify General dropdown Label-----------------------------
            PageName = "General dropdown Label"
            Ptitle1 = "General"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//div[@class='content yellow_color']/div[1]/div/form/div/div/div/select/optgroup[1]").get_attribute('label')
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + "  is present in select entry dropdown")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present in select entry dropdown")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------
            # ---------------------------Verify Phone Call dropdown value-----------------------------
            PageName = "Phone Call dropdown value"
            Ptitle1 = "Phone Call"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//div[@class='content yellow_color']/div[1]/div/form/div/div/div/select/optgroup[1]/option[1]").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + "  is present in select entry dropdown ")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present in select entry dropdown")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Email dropdown value-----------------------------
            PageName = "Email dropdown value"
            Ptitle1 = "Email"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//div[@class='content yellow_color']/div[1]/div/form/div/div/div/select/optgroup[1]/option[2]").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + "  is present in select entry dropdown")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present in select entry dropdown")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Letter dropdown value-----------------------------
            PageName = "Letter dropdown value"
            Ptitle1 = "Letter"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//div[@class='content yellow_color']/div[1]/div/form/div/div/div/select/optgroup[1]/option[3]").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + "  is present in select entry dropdown")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present in select entry dropdown")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify SMS dropdown value-----------------------------
            PageName = "SMS dropdown value"
            Ptitle1 = "SMS"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//div[@class='content yellow_color']/div[1]/div/form/div/div/div/select/optgroup[1]/option[4]").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + "  is present in select entry dropdown")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present in select entry dropdown")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify New Plan Form dropdown label-----------------------------
            PageName = "New Plan Form dropdown label"
            Ptitle1 = "New Plan Form"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//div[@class='content yellow_color']/div[1]/div/form/div/div/div/select/optgroup[2]/option").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + "  is present in select entry dropdown")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present in select entry dropdown")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify New Plan Form dropdown value-----------------------------
            PageName = "New Plan Form dropdown value"
            Ptitle1 = "New Plan Form"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//div[@class='content yellow_color']/div[1]/div/form/div/div/div/select/optgroup[2]/option").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + "  is present in select entry dropdown")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present in select entry dropdown")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Header of Choose Template dropdown-----------------------------
            PageName = "Header of Choose Template dropdown"
            Ptitle1 = "Choose Template"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//div[@class='content yellow_color']/div[2]/div[1]/div/div/form/div/div/label").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + "  is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Choose Template dropdown-----------------------------
            PageName = "Choose Template dropdown"
            Ptitle1 = "selectTemplate"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//div[@class='content yellow_color']/div[2]/div[1]/div/div/form/div/div/select").get_attribute('id')
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + "  is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            PageTitle1 = driver.find_element_by_xpath(
                "//div[@class='content yellow_color']/div[2]/div[1]/div/div/form/div/div/select").click()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify General dropdown value-----------------------------
            PageName = "General dropdown value"
            Ptitle1 = "General"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//div[@class='content yellow_color']/div[2]/div[1]/div/div/form/div/div/select/option[1]").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + "  is present in Choose Template dropdown")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present in Choose Template dropdown")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify New Plan Form dropdown value-----------------------------
            PageName = "New Plan Form dropdown value"
            Ptitle1 = "New Plan Form"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//div[@class='content yellow_color']/div[2]/div[1]/div/div/form/div/div/select/option[2]").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + "  is present in Choose Template dropdown")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present in Choose Template dropdown")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Communication Log Listing-General table header-----------------------------
            PageName = "Communication Log Listing-General table header"
            Ptitle1 = "Communication Log Listing - General"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//div[@class='content yellow_color']/div[2]/div[1]/div/div/h2").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + "  is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------
            #
            # ---------------------------Verify Filter search for Communication Log Listing-General table-----------------------------
            PageName = "Filter search for Communication Log Listing-General table"
            Ptitle1 = "search"
            try:
                PageTitle1 = driver.find_element_by_xpath("//div[@id='comm-log-blank_filter']/label/input").get_attribute('type')
                assert PageTitle1 in Ptitle1, PageName + " not able to open"
                driver.find_element_by_xpath("//div[@id='comm-log-blank_filter']/label/input").send_keys("testing")
                TestResult.append(PageName + "  is present and user is able to send inputs")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # # ---------------------------Verify Presence of elements in Communication Log Listing-General table-----------------------------
            inside = "Communication Log Listing - General"
            # ---------------loop for Columns in table for Funds View----------
            ItemList = ["#", "Date & Time", "Incoming / Outgoing", "Communication Type", "Main Contact Name", "Other Contacts", "Staff Name", "View Note"]
            ItemPresent = []
            ItemNotPresent = []
            for ii in range(len(ItemList)):
                Text1 = ItemList[ii]
                try:
                    Element1 = driver.find_element_by_xpath(
                        "//table[@id='comm-log-blank']/thead/tr/th[" + str(
                            ii + 1) + "]").text
                except Exception:
                    pass
                try:
                    assert Text1 in Element1, Text1 + " column under " + inside + " table is not present"
                    ItemPresent.append(Text1)
                except Exception as e1:
                    ItemNotPresent.append(Text1)
            if ItemPresent:
                print("ItemPresent list is not empty")
                ListC = ', '.join(ItemPresent)
                TestResult.append("Below columns are present under " + inside + " table\n" + ListC)
                TestResultStatus.append("Pass")
            if ItemNotPresent:
                print("ItemNotPresent list is not empty")
                ListD = ', '.join(ItemNotPresent)
                TestResult.append("Below columns are not present under " + inside + " table\n" + ListD)
                TestResultStatus.append("Fail")
            # # ---------------------------------------------------------------------------------

            # ---------------------------Verify Communication Log Listing - Draft General table header-----------------------------
            PageName = "Communication Log Listing - Draft General table header"
            Ptitle1 = "Communication Log Listing - Draft General"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//div[@class='content yellow_color']/div[2]/div[3]/div/div/h2").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + "  is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Filter search for Communication Log Listing - Draft General table-----------------------------
            PageName = "Filter search for Communication Log Listing - Draft General table"
            Ptitle1 = "search"
            try:
                PageTitle1 = driver.find_element_by_xpath("//div[@class='content yellow_color']/div[2]/div[4]/div/div/div/label/input").get_attribute('type')
                assert PageTitle1 in Ptitle1, PageName + " not able to open"
                driver.find_element_by_xpath("//div[@class='content yellow_color']/div[2]/div[4]/div/div/div/label/input").send_keys("testabc")
                TestResult.append(PageName + "  is present and user is able to send inputs")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # # ---------------------------Verify Presence of elements in Communication Log Listing - Draft General-----------------------------
            inside = "Communication Log Listing - Draft General"
            # ---------------loop for Columns in table for Funds View----------
            ItemList = ["#", "Date & Time", "Incoming / Outgoing", "Communication Type", "Main Contact Name",
                        "Other Contacts", "Staff Name"]
            ItemPresent = []
            ItemNotPresent = []
            for ii1 in range(len(ItemList)):
                Text1 = ItemList[ii1]
                try:
                    Element1 = driver.find_element_by_xpath(
                        "//table[@id='comm-log-draft-blank']/thead/tr/th[" + str(
                            ii1 + 1) + "]").text
                except Exception:
                    pass
                try:
                    assert Text1 in Element1, Text1 + " column under " + inside + " table is not present"
                    ItemPresent.append(Text1)
                except Exception as e1:
                    ItemNotPresent.append(Text1)
            if ItemPresent:
                print("ItemPresent list is not empty")
                ListC = ', '.join(ItemPresent)
                TestResult.append("Below columns are present under " + inside + " table\n" + ListC)
                TestResultStatus.append("Pass")
            if ItemNotPresent:
                print("ItemNotPresent list is not empty")
                ListD = ', '.join(ItemNotPresent)
                TestResult.append("Below columns are not present under " + inside + " table\n" + ListD)
                TestResultStatus.append("Fail")
            # # ---------------------------------------------------------------------------------
            # ---------------------------Verify Presence of back button on communication log page-----------------------------
            PageName = "Back button"
            Ptitle1 = "Back"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//div[@class='content yellow_color']/div/div[2]/div/a").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + "  is present on communication log page")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present on communication log page")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # # ---------------------------------------------------------------------------------

        except Exception as err:
            print(err)
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


