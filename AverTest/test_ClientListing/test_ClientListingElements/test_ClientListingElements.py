import datetime
import math
import os
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from sys import platform
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("email").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("password").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_ClientListingElements"
  description = "This test scenario is to verify all the Elements present at Client Listing of application"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_ClientListingElements"
  global Exe
  Exe="Yes"
  Directory = 'test_ClientListing/'

  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'D:/AVER/AverTest/' + Directory

  MachineName = os.getenv('COMPUTERNAME')
  print(MachineName)
  if MachineName == "DESKTOP-JLLTS65":
      path = path.replace('D:', 'C:')

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      if platform == "linux" or platform == "linux2":
          driver = webdriver.Chrome(executable_path="/home/legion/office 1wayit/AVER/AverTest/chrome/chromedriverLinux")
      elif platform == "win32" or platform == "win64":
          if MachineName == "DESKTOP-JLLTS65":
              driver = webdriver.Chrome(executable_path="C:/AVER/AverTest/chrome/chromedriver.exe")
          else:
              driver = webdriver.Chrome(executable_path="D:/AVER/AverTest/chrome/chromedriver.exe")

      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://averreplica.1wayit.com/login")
      enter_username("admin@averplanning.com")
      enter_password("admin786")
      driver.find_element_by_xpath("//button[@type='submit']").click()

  yield
  if Exe == "Yes":
      ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      time_change = datetime.timedelta(hours=5)
      new_time = datetime.datetime.now() + time_change
      ctReportHeader = new_time.strftime("%d %B %Y %I %M%p")
      ct1 = new_time.strftime("%d_%B_%Y_%I_%M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct1 + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 2
        LONG_TIMEOUT = 60
        LOADING_ELEMENT_XPATH = "//div[@class='main-loader LoaderImageLogo']"
        try:
            # ---------------------------Verify Client Listing elements-----------------------------
            PageName = "Client Listing icon"
            Ptitle1 = ""
            try:
                driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[3]/a/i").click()
                time.sleep(2)
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed()==True:
                            time.sleep(0.5)
                    except Exception:
                        break
                TestResult.append(PageName + " is present in left menu and able to click")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Page title-----------------------------
            PageName = "Page title"
            Ptitle1 = "Client Listing"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//h2[text()='Client Listing']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " (Client Listing) is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " (Client Listing) is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Total active participants tab -------------------------------------
            PageName = "Total active participants tab"
            Ptitle1 = "Total active participants "
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//p[text()='Total active participants ']").text
                print(PageTitle1)
                if "\n" in PageTitle1:
                    PageTitle1 = PageTitle1.split("\n")
                    PageTitle1 = PageTitle1[0]
                    print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Participants added today tab -------------------------------------
            PageName = "Participants added today tab"
            Ptitle1 = "Participants added Today"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//p[text()='Participants added Today ']").text
                print(PageTitle1)
                if "\n" in PageTitle1:
                    PageTitle1 = PageTitle1.split("\n")
                    PageTitle1 = PageTitle1[0]
                    print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Back button -------------------------------------
            PageName = "Back button"
            Ptitle1 = "Back"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//a[text()='Back']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Create New Client button-------------------------------------
            PageName = "Create New Client button"
            Ptitle1 = "Create New Client"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//a[text()='Create New Client']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Import button-------------------------------------
            PageName = "Import button"
            Ptitle1 = "Import"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//a[text()='Import']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Download to CSV button-------------------------------------
            PageName = "Download to CSV button"
            Ptitle1 = "Download to CSV"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//a[text()='Download to CSV']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Create Report button-------------------------------------
            PageName = "Create Report button"
            Ptitle1 = "Create Report"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//a[text()='Create Report']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of search filter for client table-------------------------------------
            PageName = "Search filter for client table"
            Ptitle1 = ""
            try:
                driver.find_element_by_xpath(
                    "//input[@id='searchFilter']").clear()
                driver.find_element_by_xpath(
                    "//input[@id='searchFilter']").send_keys("Test search")
                TestResult.append(PageName + " is present and user is able to send inputs")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------



        except Exception as err:
            print(err)
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


