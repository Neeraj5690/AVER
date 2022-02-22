import datetime
import math
import re
from selenium.webdriver.support.select import Select
import time
import openpyxl
from datetime import datetime,date
import datetime as datetime
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from sys import platform
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import pyperclip
import random
import string


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("email").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("password").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_CreateClientWorking"
  description = "This test scenario is to verify the Working of Create new client process"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_CreateClient"
  global Exe
  Exe="Yes"
  Directory = 'test_InvoiceEntry/'
  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'C:/AVER/AverTest/' + Directory

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      if platform == "linux" or platform == "linux2":
          driver = webdriver.Chrome(executable_path="/home/legion/office 1wayit/AVER/AverTest/chrome/chromedriverLinux1")
      elif platform == "win32" or platform == "win64":
          driver = webdriver.Chrome(executable_path="C:/AVER/AverTest/chrome/chromedriver.exe")

      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://averreplica.1wayit.com/login")
      enter_username("admin@averplanning.com")
      enter_password("admin786")
      driver.find_element_by_xpath("//button[@type='submit']").click()

  yield
  if Exe == "Yes":
      ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      #driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    global select
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 3
        LONG_TIMEOUT = 60
        LOADING_ELEMENT_XPATH = "//div[@class='main-loader LoaderImageLogo']"
        try:
            # ---------------------------Verify Invoice Entry icon click-----------------------------
            today = date.today()
            D1 = today.strftime("%d-%m-%Y")
            ClientListing = driver.find_element_by_xpath(
                "//div[@class='card card-sidebar-mobile']/ul/li[3]/a/i").click()
            time.sleep(2)
            Records = driver.find_elements_by_xpath("//div[@class='datatable-scroll']/table/tbody/tr")
            RowsLength = len(Records)
            for i in range(1,RowsLength):
                FirstName = driver.find_element_by_xpath("//div[@class='datatable-scroll']/table/tbody/tr["+str(i)+"]/td[2]").text
                print(FirstName)

                if FirstName == "BitsInGlass":
                    driver.find_element_by_xpath("//table[@class='table datatable-sorting dataTable']/tbody/tr[1]/td[2]").click()
                    break
                elif FirstName != "BitsInGlass":
                    driver.find_element_by_xpath("//a[text()='Create New Client']").click()
                    break
            for aa in range(5):
                letters = string.ascii_lowercase
                returna = ''.join(random.choice(letters) for i in range(5))
                FName = returna
            print(FName)

            Data=[FName,"Lname","TReferTo","01-02-1990","2456",D1,"1122334455","TStreet","123","TSuburb","@test.com","213243","1000","TestSupport","TestCommunication"]
            for i2 in range(1, 29):
                if i2==1:
                    time.sleep(1)
                    select = Select(driver.find_element_by_xpath("//div[@id='createnewclient']/div/div/div[2]/form/div[1]/div/select"))
                    select.select_by_visible_text("Active")
                elif i2 == 2:
                    time.sleep(1)
                    driver.find_element_by_xpath(
                        "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/input").send_keys(
                        Data[0])
                elif i2 == 3:
                    time.sleep(1)
                    driver.find_element_by_xpath(
                        "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/input").send_keys(
                        Data[1])
                elif i2 == 4:
                    time.sleep(1)
                    driver.find_element_by_xpath(
                        "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/input").send_keys(
                        Data[2])
                elif i2 == 5:
                    time.sleep(1)
                    driver.find_element_by_xpath(
                        "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/input").send_keys(
                        Data[3])
                    ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                elif i2==6:
                    time.sleep(1)
                    select = Select(driver.find_element_by_xpath("//div[@id='createnewclient']/div/div/div[2]/form/div[6]/div/select"))
                    select.select_by_visible_text("Male")
                elif i2 == 9:
                    time.sleep(1)
                    NDISNumToUSe = int(Data[4])
                    for NDISNum in range(10):
                        driver.find_element_by_xpath(
                            "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/input").send_keys(
                            str(NDISNumToUSe))
                        ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                        try:
                            WebDriverWait(driver, SHORT_TIMEOUT
                                          ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                            WebDriverWait(driver, LONG_TIMEOUT
                                          ).until(EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                        except TimeoutException:
                            pass
                        time.sleep(1)
                        NdisError = driver.find_element_by_xpath(
                            "//span[@id='error_ndIs']").is_displayed()
                        if NdisError == True:
                            driver.find_element_by_xpath(
                                "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/input").clear()
                            pass
                        elif  NdisError == False:
                            break
                        NDISNumToUSe=NDISNumToUSe+1

                elif i2 == 10:
                    time.sleep(1)
                    driver.find_element_by_xpath(
                        "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/input").send_keys(
                        Data[5])
                    ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                #------------asasasasd------------------
                elif i2 == 11:
                    time.sleep(1)
                    driver.find_element_by_xpath(
                        "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/input").send_keys(
                        Data[6])
                # ------------asasasasd------------------
                elif i2 == 12:
                    time.sleep(1)
                    driver.find_element_by_xpath(
                        "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/input").send_keys(
                        Data[7])
                # ------------asasasasd------------------
                elif i2 == 13:
                    time.sleep(1)
                    driver.find_element_by_xpath(
                        "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/input").send_keys(
                        Data[8])
                elif i2 == 14:
                    time.sleep(1)
                    driver.find_element_by_xpath(
                        "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/input").send_keys(
                        Data[9])
                elif i2 == 15:
                    time.sleep(1)
                    driver.find_element_by_xpath(
                        "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/input").send_keys(Data[0]+
                        Data[10])
                elif i2==16:
                    time.sleep(1)
                    select = Select(driver.find_element_by_xpath("//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/select"))
                    select.select_by_visible_text("SA")
                elif i2==17:
                    time.sleep(1)
                    select = Select(driver.find_element_by_xpath("//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/select"))
                    select.select_by_visible_text("Yes")
                elif i2 == 18:
                    time.sleep(1)
                    driver.find_element_by_xpath(
                        "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/input").send_keys(
                        Data[11])
                elif i2==19:
                    time.sleep(1)
                    select = Select(driver.find_element_by_xpath("//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/select"))
                    select.select_by_visible_text("Plan Managed")
                elif i2 == 20:
                    time.sleep(1)
                    driver.find_element_by_xpath(
                        "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/input").send_keys(
                        Data[12])
                elif i2 == 21:
                    time.sleep(1)
                    driver.find_element_by_xpath(
                        "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/input").send_keys(
                        Data[13])
                elif i2==22:
                    time.sleep(1)
                    select = Select(driver.find_element_by_xpath("//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/select"))
                    select.select_by_visible_text("Yes")
                elif i2==23:
                    time.sleep(1)
                    select = Select(driver.find_element_by_xpath("//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/select"))
                    select.select_by_visible_text("Yes")
                elif i2==24:
                    time.sleep(1)
                    select = Select(driver.find_element_by_xpath("//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/select"))
                    select.select_by_visible_text("No")
                elif i2 == 26:
                    time.sleep(1)
                    driver.find_element_by_xpath(
                        "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/textarea").send_keys(
                        Data[14])
                elif i2==27:
                    time.sleep(1)
                    select = Select(driver.find_element_by_xpath("//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/select"))
                    select.select_by_visible_text("National Remote")
                elif i2 == 28:
                    time.sleep(1)
                    driver.find_element_by_xpath(
                        "//div[@id='createnewclient']/div/div/div[2]/form/div[28]/button").click()
                    time.sleep(2)

                    EmailError = driver.find_element_by_xpath(
                        "//span[@id='error_email_address']").is_displayed()
                    if EmailError == True:
                        time.sleep(1)
                        driver.find_element_by_xpath(
                            "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(15) + "]/div/input").clear()
                        driver.find_element_by_xpath(
                            "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(15) + "]/div/input").send_keys(Data[0]+Data[1] +Data[10])


            # ---------------------------------------------------------------------------------


        except Exception as err:
            print(err)
            TestResult.append("Invoice entry is not working correctly. Below error found\n"+str(err))
            TestResultStatus.append("Fail")
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


