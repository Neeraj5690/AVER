import datetime
import math
import re
import time
import openpyxl
from datetime import datetime,date
import datetime as datetime
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from sys import platform
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import pyperclip
import random
import os


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("email").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("password").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_ClientDataVerify"
  description = "This test scenario is to verify Client Data at client listing section of the Portal"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_PortalWorking"
  global Exe
  Exe="Yes"
  Directory = 'test_Portal/'

  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'D:/AVER/AverTest/' + Directory

  MachineName = os.getenv('COMPUTERNAME')
  if MachineName == "DESKTOP-JLLTS65":
      path = path.replace('D:', 'C:')

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      if platform == "linux" or platform == "linux2":
          driver = webdriver.Chrome(executable_path="/home/legion/office 1wayit/AVER/AverTest/chrome/chromedriverLinux1")
      elif platform == "win32" or platform == "win64":
          if MachineName == "DESKTOP-JLLTS65":
              driver = webdriver.Chrome(executable_path="C:/AVER/AverTest/chrome/chromedriver.exe")
          else:
              driver = webdriver.Chrome(executable_path="D:/AVER/AverTest/chrome/chromedriver.exe")

      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://averreplica.1wayit.com/login")
      enter_username("admin@averplanning.com")
      enter_password("admin786")
      driver.find_element_by_xpath("//button[@type='submit']").click()

  yield
  if Exe == "Yes":
      time_change = datetime.timedelta(hours=5)
      new_time = datetime.datetime.now() + time_change
      ctReportHeader = new_time.strftime("%d %B %Y %I %M%p")

      ct = new_time.strftime("%d_%B_%Y_%I_%M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 3
        LONG_TIMEOUT = 60
        LOADING_ELEMENT_XPATH = "//body[@class='sidebar-xs loader_overlay']"
        try:
            print()
            #----------------Fecthing Client name from the ref Data sheet--------------------
            ExcelFileName2 = "RefData"
            locx2 = (path + 'Ref/' + ExcelFileName2 + '.xlsx')
            wbx2 = openpyxl.load_workbook(locx2)
            sheetx2 = wbx2.active

            try:
                FNameXL=sheetx2.cell(1, 1).value
                print(FNameXL)
                LNameXL = sheetx2.cell(1, 2).value
                print(LNameXL)
                NDISXL = sheetx2.cell(1, 3).value
                print(NDISXL)

                if FNameXL==None or LNameXL==None or NDISXL==None:
                    print("Either First name, Last name or NDIS number not found in ref sheet")
                    TestResult.append("Either First name, Last name or NDIS number not found in ref sheet")
                    TestResultStatus.append("Fail")
                    driver.close()
            except Exception as sd:
                print(sd)
                print("Ref sheet is not able to read, please check the ref doc sheet")
                TestResult.append("Ref sheet is not able to read, please check the ref doc sheet")
                TestResultStatus.append("Fail")
                driver.close()

            # ---------------------------Verify Portal icon click-----------------------------
            PageName = "Client listing page"
            try:
                driver.find_element_by_xpath("//i[@class='icon-paragraph-justify3']/parent::a").click()
                time.sleep(2)
                driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[3]/a").click()
                time.sleep(2)

                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break

                time.sleep(2)
                TestResult.append(PageName + " opened successfully")
                TestResultStatus.append("Pass")
            except Exception as ee:
                print(ee)
                TestResult.append(PageName + " is not able to open")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # -----------------Searching client on Client listing page -----------------------
            driver.find_element_by_xpath("//input[@id='searchFilter']").send_keys(FNameXL)
            TestResult.append("Searching for client [ "+FNameXL+" ] at client listing")
            TestResultStatus.append("Pass")
            time.sleep(1)
            ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
            driver.find_element_by_xpath("//table/tbody/tr[1]/td[text()='"+FNameXL+"']/following-sibling::td[text()='"+LNameXL+"']/following-sibling::td[text()='"+str(NDISXL)+"']").click()
            TestResult.append("Client is able to search")
            TestResultStatus.append("Pass")

            time.sleep(2)
            try:
                button = driver.find_element_by_xpath("//div[@id='alert_modal']/div/div/div/button")
                driver.execute_script("arguments[0].click();", button)
            except Exception as dd:
                print(dd)
                pass
            # ---------------------------------------------------------------------------------

            ProfileDataDic={}
            # ---------------------Storing client data in Dictionary --------------------------
            UsernamewithNDIS=driver.find_element_by_xpath("//div[@class='src_sec_hed']/a[text()='Back']/parent::div/parent::div/h2").text
            print(UsernamewithNDIS)
            NameDataList=UsernamewithNDIS.split()
            Firstname=NameDataList[0]
            print(Firstname)
            TestResult.append("Client first name [ "+Firstname+" ] stored in List")
            TestResultStatus.append("Pass")

            Lastname = NameDataList[1]
            print(Lastname)
            TestResult.append("Client last name [ "+Lastname+" ] stored in List")
            TestResultStatus.append("Pass")
            ProfileDataDic["User Name"]=Firstname+" "+Lastname

            NDIS = NameDataList[3]
            print(NDIS)
            TestResult.append("Client NDIS [ "+NDIS+" ] stored in List")
            TestResultStatus.append("Pass")
            ProfileDataDic["NDIS Number"] = NDIS

            ContactNumber=driver.find_element_by_xpath("//div/label[text()='Mobile Number']/following-sibling::span").text
            print(ContactNumber)
            if ContactNumber=="":
                ContactNumber="N/A"
            TestResult.append("Client Contact Number [ " + ContactNumber + " ] stored in List")
            TestResultStatus.append("Pass")
            ProfileDataDic["Contact Number"] = ContactNumber

            EmailAddress = driver.find_element_by_xpath("//div/label[text()='Email']/following-sibling::span").text
            print(EmailAddress)
            TestResult.append("Client Email Address [ " + EmailAddress + " ] stored in List")
            TestResultStatus.append("Pass")
            ProfileDataDic["Email Address"] = EmailAddress

            UserAddress = driver.find_element_by_xpath("//div/label[text()='Address']/following-sibling::span").text
            print(UserAddress)
            TestResult.append("Client User Address [ " + UserAddress + " ] stored in List")
            TestResultStatus.append("Pass")
            ProfileDataDic["User Address"] = UserAddress

            time.sleep(2)
            PlanTablePresence=driver.find_element_by_xpath("//table[@class='table datatable-sorting']/thead").is_displayed()
            print(PlanTablePresence)
            if PlanTablePresence==True:
                print("Service Booking Plan Table is Present")
                TestResult.append("Service Booking Plan Table is Present")
                TestResultStatus.append("Pass")
                PlanStatusCount = driver.find_elements_by_xpath("//table[@class='table datatable-sorting']/tbody/tr/td[@class='sub_tbl_icn']")
                print(len(PlanStatusCount))
                ProfileDataDic["Plan Status Count"] = len(PlanStatusCount)
                TestResult.append("Plans Count is "+str(len(PlanStatusCount)))
                TestResultStatus.append("Pass")

                if len(PlanStatusCount)==0:
                    PlanStatusCount="NA"
                    ProfileDataDic["Plan Status"] = PlanStatusCount
                    TestResult.append("Plan Status found "+PlanStatusCount)
                    TestResultStatus.append("Pass")
                elif len(PlanStatusCount)>0:
                    FirstPlanName=driver.find_element_by_xpath("//table[@class='table datatable-sorting']/tbody/tr[1]/td[@class='sub_tbl_icn']/preceding-sibling::td[6]").text
                    print(FirstPlanName)
                    TestResult.append("Plan at first position " + FirstPlanName)
                    TestResultStatus.append("Pass")

                    ActivePlan = driver.find_element_by_xpath(
                        "//table[@class='table datatable-sorting']/tbody/tr[1]/td[@class='sub_tbl_icn']/preceding-sibling::td[5]").text
                    print(ActivePlan)

                    if ActivePlan=="Active":
                        ProfileDataDic["Plan Status"] = FirstPlanName
                        TestResult.append("Plan status found " + ActivePlan)
                        TestResultStatus.append("Pass")
                    else:
                        ProfileDataDic["Plan Status"] = "NA"
                        TestResult.append("Plan status found Inactive")
                        TestResultStatus.append("Pass")
            else:
                print("Service Booking Plan Table is not Present")
                TestResult.append("Service Booking Plan Table is not Present")
                TestResultStatus.append("Fail")


            # ----------------------------Storing data in the excel-------------------------------
            sheetx2.cell(3, 1).value=ProfileDataDic["User Name"]
            sheetx2.cell(3, 2).value = str(ProfileDataDic["NDIS Number"])
            sheetx2.cell(3, 3).value = ProfileDataDic["Contact Number"]
            sheetx2.cell(3, 4).value = ProfileDataDic["Email Address"]
            sheetx2.cell(3, 5).value = ProfileDataDic["User Address"]
            sheetx2.cell(3, 6).value = ProfileDataDic["Plan Status"]
            wbx2.save(locx2)

            # ----------------------Fetching Additional Contact details-------------------------------
            driver.find_element_by_xpath("//a[text()='Additional Contacts']").click()
            for load in range(LONG_TIMEOUT):
                try:
                    if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                        time.sleep(0.5)
                except Exception:
                    break

            ACFirstNameList=[]
            ACLastNameList = []
            ACRelationList = []

            AddContCount=driver.find_elements_by_xpath("//table[@id='table_data']/tbody/tr")
            if len(AddContCount)==1:
                print("No additional Contact present")
                TestResult.append("No additional Contact present")
                TestResultStatus.append("Pass")
            elif len(AddContCount)>1:
                print("Additional Contact present")
                print(len(AddContCount))
                TestResult.append("Additional Contact present [ "+str(len(AddContCount))+" ]")
                TestResultStatus.append("Pass")
                for c in range(len(AddContCount)):
                    print()
                    print("c is "+str(c))
                    ACFirstName=driver.find_element_by_xpath("//table[@id='table_data']/tbody/tr["+str(c+1)+"]/td[2]/a").text
                    print(ACFirstName)
                    ACFirstNameList.append(ACFirstName)
                    ProfileDataDic["AC First Name"] = ACFirstNameList

                    ACLastName = driver.find_element_by_xpath(
                        "//table[@id='table_data']/tbody/tr[" + str(c+1) + "]/td[3]").text
                    print(ACLastName)
                    ACLastNameList.append(ACLastName)
                    ProfileDataDic["AC Last Name"] = ACLastNameList

                    ACRelation = driver.find_element_by_xpath(
                        "//table[@id='table_data']/tbody/tr[" + str(c+1) + "]/td[4]").text
                    print(ACRelation)
                    ACRelationList.append(ACRelation)
                    ProfileDataDic["AC Relation"] = ACRelationList

            print(ProfileDataDic)
            for c1 in range(len(ACFirstNameList)):
                print("c1 is "+str(c1))
                sheetx2.cell(c1+5, 1).value = ACFirstNameList[c1]
                sheetx2.cell(c1 + 5, 2).value = ACLastNameList[c1]
                sheetx2.cell(c1 + 5, 3).value = ACRelationList[c1]
                TestResult.append(ACFirstNameList[c1]+", "+ACLastNameList[c1]+", "+ACRelationList[c1])
                TestResultStatus.append("Pass")

            if len(ACFirstNameList)==0:
                sheetx2.cell(4, 2).value ="No Contact found"
            else:
                sheetx2.cell(4, 2).value = len(ACFirstNameList)

            wbx2.save(locx2)

        except Exception as err:
            print(err)
            TestResult.append("Portal is not working correctly. Below error found\n"+str(err))
            TestResultStatus.append("Fail")
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


