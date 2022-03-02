import datetime
import math
import re
import time
from telnetlib import EC

import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from sys import platform

from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("email").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("password").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path
  global ClickCounter

  TestName = "test_ServiceProviderElements"
  description = "This test scenario is to verify all the Elements present at Login Page"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_ElementsPresent"
  global Exe
  Exe="Yes"
  Directory = 'test_ServiceProviderListing/'
  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'D:/AVER/AverTest/' + Directory

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      if platform == "linux" or platform == "linux2":
          driver=webdriver.Chrome(executable_path="/home/legion/office 1wayit/AVER/AverTest/chrome/chromedriverLinux")
      elif platform == "win32" or platform == "win64":
          driver = webdriver.Chrome(executable_path="D:/AVER/AverTest/chrome/chromedriver.exe")
      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://averreplica.1wayit.com/login")
      enter_username("admin@averplanning.com")
      enter_password("admin786")
      driver.find_element_by_xpath("//button[@type='submit']").click()

  yield
  if Exe == "Yes":
      ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        TimeSpeed = 1
        SHORT_TIMEOUT = 5
        LONG_TIMEOUT = 400
        LOADING_ELEMENT_XPATH = "//div[@id='appian-working-indicator-hidden']"
        try:
            #---------------------------Verify Service provider listing icon-----------------------------
            PageName="Servive provider listing icon"
            Ptitle1=""
            try:
                driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[7]/a").click()
                time.sleep(2)
                TestResult.append(PageName + " is present in left menu and able to click")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            #---------------------------------------------------------------------------------

            # ---------------------------Verify Page title-----------------------------
            PageName = "Page title"
            Ptitle1 = "Service Provider Listing"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//h2[text()='Service Provider Listing']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " (Service provider listing) is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " (Service provider listing) is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of back button on service provider listing page-----------------------------
            PageName = "Back button"
            Ptitle1 = "Back"
            try:
                PageTitle1 = driver.find_element_by_xpath("//a[text()='Back']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not able to open"
                TestResult.append(PageName + "  is present on service provider listing page")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present on service provider listing page")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of import button on service provider listing page-----------------------------
            PageName = "Import button"
            Ptitle1 = "Import"
            try:
                PageTitle1 = driver.find_element_by_xpath("//a[text()='Import']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not able to open"
                TestResult.append(PageName + "  is present on service provider listing page")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present on service provider listing page")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of download csv button on service provider listing page-----------------------------
            PageName = "Download CSV button"
            Ptitle1 = "Download CSV"
            try:
                PageTitle1 = driver.find_element_by_xpath("//a[text()='Download CSV']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not able to open"
                TestResult.append(PageName + "  is present on service provider listing page")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present on service provider listing page")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Create New button on service provider listing page-----------------------------
            PageName = "Create New button"
            Ptitle1 = "Create New"
            try:
                PageTitle1 = driver.find_element_by_xpath("//a[text()='Create New']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not able to open"
                TestResult.append(PageName + "  is present on service provider listing page")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present on service provider listing page")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Create Reimburse button on service provider listing page-----------------------------
            PageName = "Create Reimburse button"
            Ptitle1 = "Create Reimburse"
            try:
                PageTitle1 = driver.find_element_by_xpath("//a[text()='Create Reimburse']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not able to open"
                TestResult.append(PageName + "  is present on service provider listing page")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present on service provider listing page")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Search filter under service provider listing table-----------------------------
            PageName = "Search filter"
            Ptitle1 = "search"
            try:
                PageTitle1 = driver.find_element_by_xpath("//div[@id='table_data_filter']/label/input").get_attribute('type')
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not able to open"
                PageTitle1 = driver.find_element_by_xpath("//div[@id='table_data_filter']/label/input").send_keys("Test service")
                TestResult.append(PageName + "  is present under service provider listing table and user is able to send inputs ")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present under service provider listing table")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of elements in Service provider table-----------------------------
            inside = "Service provider listing"
            # ---------------loop for Columns in table for Funds View----------
            ItemList = ["#", "Name", "Service Type", "Franchise", "ABN",
                        "Account Name", "BSB", "Account Number","Action"]
            ItemPresent = []
            ItemNotPresent = []
            for ii in range(len(ItemList)):
                Text1 = ItemList[ii]
                try:
                    Element1 = driver.find_element_by_xpath(
                        "//table[@id='table_data']/thead/tr/th[" + str(
                            ii + 1) + "]").text
                except Exception:
                    pass
                try:
                    assert Text1 in Element1, Text1 + " column under " + inside + " table is not present"
                    ItemPresent.append(Text1)
                except Exception as e1:
                    ItemNotPresent.append(Text1)
            if ItemPresent:
                print("ItemPresent list is not empty")
                ListC = ', '.join(ItemPresent)
                TestResult.append("Below columns are present under [ " + inside + " ] table\n" + ListC)
                TestResultStatus.append("Pass")
            if ItemNotPresent:
                print("ItemNotPresent list is not empty")
                ListD = ', '.join(ItemNotPresent)
                TestResult.append("Below columns are not present under [ " + inside + " ] table\n" + ListD)
                TestResultStatus.append("Fail")
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Pagination clicks-----------------------------
            PageName = "Service provider listing table"
            try:
                TotalItem = driver.find_element_by_xpath("//div[@id='table_data_wrapper']/div[3]/div[1]").text
                substr = "of"
                x = TotalItem.split(substr)
                string_name = x[0]
                TotalItemAfterOf = x[1]
                abc = ""
                countspace = 0
                for element in range(0, len(string_name)):
                    if string_name[(len(string_name) - 1) - element] == " ":
                        countspace = countspace + 1
                        if countspace == 2:
                            break
                    else:
                        abc = abc + string_name[(len(string_name) - 1) - element]
                abc = abc[::-1]
                TotalItemBeforeOf = abc
                TotalItemAfterOf = TotalItemAfterOf.split(" ")
                TotalItemAfterOf = TotalItemAfterOf[1]
                TotalItemAfterOf = re.sub('[^A-Za-z0-9]+', '', TotalItemAfterOf)

                TotalItemAfterOf = int(TotalItemAfterOf)
                RecordsPerPage = 50
                TotalPages = TotalItemAfterOf / RecordsPerPage
                NumberOfPages = math.ceil(float(TotalPages))

                for i in range(NumberOfPages):
                    if i < 1:
                        if i == NumberOfPages - 1:
                            TestResult.append("No Pagination found for [ " + str(
                                RecordsPerPage) + " ] no. of records under Service Provider Listing table")
                            TestResultStatus.append("Pass")
                            break
                    try:
                        time.sleep(TimeSpeed)
                        driver.find_element_by_xpath(
                            "//div[@id='table_data_paginate']/a[2]").click()
                        time.sleep(1)
                        ClickCounter = ClickCounter + 1
                        TestResult.append("Pagination verified for [ " + str(
                            TotalItemAfterOf) + " ] no. of records under Service Provider Listing table")
                        TestResultStatus.append("Pass")
                    except Exception as cc:
                        pass
                if i != ClickCounter:
                    TestResult.append(
                        "Pagination for [ " + str(RecordsPerPage) + " ] no. of records is not working correctly")
                    TestResultStatus.append("Fail")
                driver.refresh()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until(EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
            except Exception as aq:
                print(aq)
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
                # # ---------------------------------------------------------------------------------

            # ---------------------------Verify Franchise Listing table header-----------------------------
            PageName = "Franchise Listing table header"
            Ptitle1 = "Franchise Listing"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//h2[text()='Franchise Listing']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Create franchise button-----------------------------
            PageName = "Create franchise button"
            Ptitle1 = "Create Franchise"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//a[text()='Create Franchise']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Search filter under franchise listing table-----------------------------
            PageName = "Search filter"
            Ptitle1 = "search"
            try:
                PageTitle1 = driver.find_element_by_xpath("//div[@id='franchise_table_data_filter']/label/input").get_attribute(
                    'type')
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not able to open"
                PageTitle1 = driver.find_element_by_xpath("//div[@id='franchise_table_data_filter']/label/input").send_keys(
                    "Test franchise")
                TestResult.append(PageName + "  is present under franchise listing table and user is able to send inputs ")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present under franchise listing table")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of elements in Service provider table-----------------------------
            inside = "Franchise listing"
            # ---------------loop for Columns in table for Funds View----------
            ItemList = ["#", "Name", "Status", "Set Up Date", "Action"]
            ItemPresent = []
            ItemNotPresent = []
            for ii in range(len(ItemList)):
                Text1 = ItemList[ii]
                try:
                    Element1 = driver.find_element_by_xpath(
                        "//table[@id='franchise_table_data']/thead/tr/th[" + str(
                            ii + 1) + "]").text
                except Exception:
                    pass
                try:
                    assert Text1 in Element1, Text1 + " column under " + inside + " table is not present"
                    ItemPresent.append(Text1)
                except Exception as e1:
                    ItemNotPresent.append(Text1)
            if ItemPresent:
                print("ItemPresent list is not empty")
                ListC = ', '.join(ItemPresent)
                TestResult.append("Below columns are present under [ " + inside + " ] table\n" + ListC)
                TestResultStatus.append("Pass")
            if ItemNotPresent:
                print("ItemNotPresent list is not empty")
                ListD = ', '.join(ItemNotPresent)
                TestResult.append("Below columns are not present under [ " + inside + " ] table\n" + ListD)
                TestResultStatus.append("Fail")
            # ---------------------------------------------------------------------------------

            # ---------------------------------------------------------------------------------

        except Exception:
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


