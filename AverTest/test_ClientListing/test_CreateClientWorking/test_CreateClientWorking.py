import datetime
import math
import re
from selenium.webdriver.support.select import Select
import time
import openpyxl
from datetime import datetime,date
import datetime as datetime
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from sys import platform
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import os
import random
import string


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("email").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("password").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_CreateClientWorking"
  description = "This test scenario is to verify the Working of Create new client process"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_CreateClientWorking"
  global Exe
  Exe="Yes"
  Directory = 'test_ClientListing/'



  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'D:/AVER/AverTest/' + Directory

  MachineName = os.getenv('COMPUTERNAME')
  print(MachineName)
  if MachineName=="DESKTOP-JLLTS65":
      path=path.replace('D:', 'C:')

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      if platform == "linux" or platform == "linux2":
          driver = webdriver.Chrome(executable_path="/home/legion/office 1wayit/AVER/AverTest/chrome/chromedriverLinux1")
      elif platform == "win32" or platform == "win64":
          if MachineName == "DESKTOP-JLLTS65":
              driver = webdriver.Chrome(executable_path="C:/AVER/AverTest/chrome/chromedriver.exe")
          else:
            driver = webdriver.Chrome(executable_path="D:/AVER/AverTest/chrome/chromedriver.exe")

      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://averreplica.1wayit.com/login")
      enter_username("admin@averplanning.com")
      enter_password("admin786")
      driver.find_element_by_xpath("//button[@type='submit']").click()

  yield
  if Exe == "Yes":
      time_change = datetime.timedelta(hours=5)
      new_time = datetime.datetime.now() + time_change
      ctReportHeader = new_time.strftime("%d %B %Y %I %M%p")

      ct = new_time.strftime("%d_%B_%Y_%I_%M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    global select
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 3
        LONG_TIMEOUT = 60
        LOADING_ELEMENT_XPATH = "//div[@class='main-loader LoaderImageLogo']"
        today = date.today()
        D1 = today.strftime("%d-%m-%Y")


        try:
            print()
            # ---------------------------To check we have existing test client in excel-----------------------------
            ClientPresentxl = "False"
            xcelFileName = "RefData"

            InvoicePath=path
            InvoicePath = os.path.abspath(os.path.join(InvoicePath, '..'))
            InvoicePath = InvoicePath.replace('\\', '/')
            InvoicePath = InvoicePath + "/test_InvoiceEntry/"
            print(InvoicePath)

            locx1 = (InvoicePath + 'ReferenceData/' + xcelFileName + '.xlsx')
            wbx1 = openpyxl.load_workbook(locx1)
            sheetx1 = wbx1.active

            for i_ref in range(1, 10):
                if sheetx1.cell(i_ref, 1).value != None:
                    FirstNamexl=sheetx1.cell(i_ref, 1).value
                    LastNamexl = sheetx1.cell(i_ref, 2).value
                    ClientPresentxl = "True"
                    break

                else:
                    ClientPresentxl="False"
                    pass

            if ClientPresentxl=="False":
                print("Client is not present in reference sheet, we need to add client first in application")
                TestResult.append(
                    "Client is not present in reference sheet, we need to add client first in application")
                TestResultStatus.append("Pass")
                # ---------------------------adding new client in application-----------------------------
                driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[3]/a/i").click()
                time.sleep(2)
                Records = driver.find_elements_by_xpath("//div[@class='datatable-scroll']/table/tbody/tr")
                RowsLength = len(Records)
                if RowsLength>50:
                    print("We need to add code for Pagination")

                for aa in range(5):
                    letters = string.ascii_lowercase
                    returna = ''.join(random.choice(letters) for i in range(5))
                    FName = returna
                print(FName)

                LName="test"
                print(LName)

                driver.find_element_by_xpath("//a[@data-target='#createnewclient']").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until(
                        EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                try:
                    print()
                    Data = [FName, LName, "TReferTo", "01-02-1990", "2480", D1, "1122334455", "TStreet", "123", "TSuburb",
                            "@test.com", "213243", "1000", "TestSupport", "TestCommunication"]
                    for i2 in range(1, 29):
                        # -------------Client Status dropdown--------------------------------------------
                        if i2 == 1:
                            time.sleep(1)
                            select = Select(driver.find_element_by_xpath(
                                "//div[@id='createnewclient']/div/div/div[2]/form/div[1]/div/select"))
                            select.select_by_visible_text("Active")
                            TestResult.append(
                                "Client 'Status' is selected from dropdown")
                            TestResultStatus.append("Pass")
                        # -------------First Name--------------------------------------------
                        elif i2 == 2:
                            time.sleep(1)
                            driver.find_element_by_xpath(
                                "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                    i2) + "]/div/input").send_keys(
                                Data[0])
                            TestResult.append(
                                "Client 'first name' is entered successfully")
                            TestResultStatus.append("Pass")
                        # -------------Last Name--------------------------------------------
                        elif i2 == 3:
                            time.sleep(1)
                            driver.find_element_by_xpath(
                                "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                    i2) + "]/div/input").send_keys(
                                Data[1])
                            TestResult.append(
                                "Client 'last name' is entered successfully")
                            TestResultStatus.append("Pass")
                        # -------------Referred To By--------------------------------------------
                        elif i2 == 4:
                            time.sleep(1)
                            driver.find_element_by_xpath(
                                "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                    i2) + "]/div/input").send_keys(
                                Data[2])
                            TestResult.append(
                                "'Referred to by' is entered successfully")
                            TestResultStatus.append("Pass")
                        # -------------DOB--------------------------------------------
                        elif i2 == 5:
                            time.sleep(1)
                            driver.find_element_by_xpath(
                                "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                    i2) + "]/div/input").send_keys(
                                Data[3])
                            ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                            TestResult.append("'DOB' is entered successfully")
                            TestResultStatus.append("Pass")
                        # -------------Gender--------------------------------------------
                        elif i2 == 6:
                            time.sleep(1)
                            select = Select(driver.find_element_by_xpath(
                                "//div[@id='createnewclient']/div/div/div[2]/form/div[6]/div/select"))
                            select.select_by_visible_text("Male")
                            TestResult.append(
                                "'Gender' is selected successfully")
                            TestResultStatus.append("Pass")
                        # -------------NDIS Number--------------------------------------------
                        elif i2 == 9:
                            time.sleep(1)
                            NDISNumToUSe = int(Data[4])
                            for NDISNum in range(10):
                                driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                        i2) + "]/div/input").send_keys(
                                    str(NDISNumToUSe))
                                ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                                try:
                                    WebDriverWait(driver, SHORT_TIMEOUT
                                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                                    WebDriverWait(driver, LONG_TIMEOUT
                                                  ).until(
                                        EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                                except TimeoutException:
                                    pass
                                time.sleep(1)
                                NdisError = driver.find_element_by_xpath(
                                    "//span[@id='error_ndIs']").is_displayed()
                                if NdisError == True:
                                    driver.find_element_by_xpath(
                                        "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                            i2) + "]/div/input").clear()
                                    pass
                                elif NdisError == False:
                                    break
                                NDISNumToUSe = NDISNumToUSe + 1
                            TestResult.append(
                                "'NDIS number' is entered successfully")
                            TestResultStatus.append("Pass")
                        # -------------Sign Up Date--------------------------------------------
                        elif i2 == 10:
                            time.sleep(1)
                            driver.find_element_by_xpath(
                                "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                    i2) + "]/div/input").send_keys(
                                Data[5])
                            ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                            TestResult.append(
                                "'Sign up date' is entered successfully")
                            TestResultStatus.append("Pass")
                        # -------------Mobile Number--------------------------------------------
                        elif i2 == 11:
                            time.sleep(1)
                            driver.find_element_by_xpath(
                                "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                    i2) + "]/div/input").send_keys(
                                Data[6])
                            TestResult.append(
                                "'Mobile number' is entered successfully")
                            TestResultStatus.append("Pass")
                        # -------------Street Address--------------------------------------------
                        elif i2 == 12:
                            time.sleep(1)
                            driver.find_element_by_xpath(
                                "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                    i2) + "]/div/input").send_keys(
                                Data[7])
                            TestResult.append(
                                "'Street address' is entered successfully")
                            TestResultStatus.append("Pass")
                        # -------------Home Number--------------------------------------------
                        elif i2 == 13:
                            time.sleep(1)
                            driver.find_element_by_xpath(
                                "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                    i2) + "]/div/input").send_keys(
                                Data[8])
                            TestResult.append(
                                "'Home Number' is entered successfully")
                            TestResultStatus.append("Pass")
                        # -------------Suburb--------------------------------------------
                        elif i2 == 14:
                            time.sleep(1)
                            driver.find_element_by_xpath(
                                "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                    i2) + "]/div/input").send_keys(
                                Data[9])
                            TestResult.append(
                                "'Suburb' is entered successfully")
                            TestResultStatus.append("Pass")
                        # -------------Email Address--------------------------------------------
                        elif i2 == 15:
                            time.sleep(1)
                            driver.find_element_by_xpath(
                                "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                    i2) + "]/div/input").send_keys(Data[0] +
                                                                   Data[10])
                            TestResult.append(
                                "'Email Address' is entered successfully")
                            TestResultStatus.append("Pass")
                        # -------------State--------------------------------------------
                        elif i2 == 16:
                            time.sleep(1)
                            select = Select(driver.find_element_by_xpath(
                                "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/select"))
                            select.select_by_visible_text("SA")
                            TestResult.append(
                                "'State' is selected successfully")
                            TestResultStatus.append("Pass")
                        # -------------Access to App--------------------------------------------
                        elif i2 == 17:
                            time.sleep(1)
                            select = Select(driver.find_element_by_xpath(
                                "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/select"))
                            select.select_by_visible_text("Yes")
                            TestResult.append(
                                "'Access to app' is selected successfully")
                            TestResultStatus.append("Pass")
                        # -------------Postcode--------------------------------------------
                        elif i2 == 18:
                            time.sleep(1)
                            driver.find_element_by_xpath(
                                "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                    i2) + "]/div/input").send_keys(
                                Data[11])
                            TestResult.append(
                                "'Postcode' is entered successfully")
                            TestResultStatus.append("Pass")
                        # -------------Profile Type--------------------------------------------
                        elif i2 == 19:
                            time.sleep(1)
                            select = Select(driver.find_element_by_xpath(
                                "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/select"))
                            select.select_by_visible_text("Plan Managed")
                            TestResult.append(
                                "'Profile Type' is selected successfully")
                            TestResultStatus.append("Pass")
                        # -------------Monthly Fee Rate ($)--------------------------------------------
                        elif i2 == 20:
                            time.sleep(1)
                            driver.find_element_by_xpath(
                                "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                    i2) + "]/div/input").send_keys(
                                Data[12])
                            TestResult.append(
                                "'Monthly Fee Rate ($)' is entered successfully")
                            TestResultStatus.append("Pass")
                        # -------------Support Coordinator--------------------------------------------
                        elif i2 == 21:
                            time.sleep(1)
                            driver.find_element_by_xpath(
                                "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                    i2) + "]/div/input").send_keys(
                                Data[13])
                            TestResult.append(
                                "'Support Coordinator' is entered successfully")
                            TestResultStatus.append("Pass")
                        # -------------Main Profile Contact--------------------------------------------
                        elif i2 == 22:
                            time.sleep(1)
                            select = Select(driver.find_element_by_xpath(
                                "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/select"))
                            select.select_by_visible_text("Yes")
                            TestResult.append(
                                "'Main Profile Contact' is selected successfully")
                            TestResultStatus.append("Pass")
                        # -------------Receive Payment Updates--------------------------------------------
                        elif i2 == 23:
                            time.sleep(1)
                            select = Select(driver.find_element_by_xpath(
                                "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/select"))
                            select.select_by_visible_text("Yes")
                            TestResult.append(
                                "'Receive Payment Updates' is selected successfully")
                            TestResultStatus.append("Pass")
                        # -------------Statement Preference--------------------------------------------
                        elif i2 == 24:
                            time.sleep(1)
                            select = Select(driver.find_element_by_xpath(
                                "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/select"))
                            select.select_by_visible_text("No")
                            TestResult.append(
                                "'Statement Preference' is selected successfully")
                            TestResultStatus.append("Pass")
                        # -------------Communication Preferences--------------------------------------------
                        elif i2 == 26:
                            time.sleep(1)
                            driver.find_element_by_xpath(
                                "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                    i2) + "]/div/textarea").send_keys(
                                Data[14])
                            TestResult.append(
                                "'Communication Preferences' is entered successfully")
                            TestResultStatus.append("Pass")
                        # -------------NDIS Rate--------------------------------------------
                        elif i2 == 27:
                            time.sleep(1)
                            select = Select(driver.find_element_by_xpath(
                                "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/select"))
                            select.select_by_visible_text("National Remote")
                            TestResult.append(
                                "'NDIS Rate' is selected successfully")
                            TestResultStatus.append("Pass")
                        # -------------Save button--------------------------------------------
                        elif i2 == 28:
                            time.sleep(1)
                            driver.find_element_by_xpath(
                                "//div[@id='createnewclient']/div/div/div[2]/form/div[28]/button").click()
                            TestResult.append(
                                "Save button is clicked")
                            TestResultStatus.append("Pass")
                            try:
                                WebDriverWait(driver, SHORT_TIMEOUT
                                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                                WebDriverWait(driver, LONG_TIMEOUT
                                              ).until(
                                    EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                            except TimeoutException:
                                pass
                            time.sleep(2)
                            try:
                                EmailError = driver.find_element_by_xpath(
                                    "//span[@id='error_email_address']").is_displayed()
                                if EmailError == True:
                                    time.sleep(1)
                                    driver.find_element_by_xpath(
                                        "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                            15) + "]/div/input").clear()
                                    driver.find_element_by_xpath(
                                        "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                            15) + "]/div/input").send_keys(Data[0] + Data[1] + Data[10])
                            except Exception as ec:
                                pass
                    TestResult.append("Create Client process working correctly")
                    TestResultStatus.append("Pass")
                except Exception:
                    TestResult.append("Create Client process is not working")
                    TestResultStatus.append("Fail")
                # --------Saving client details in reference sheet------------
                sheetx1.cell(1, 1).value = FName
                sheetx1.cell(1, 2).value = LName
                wbx1.save(locx1)

                #------------Going back to client listing after creating client---------
                driver.find_element_by_xpath("//a[text()='Back']").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until(
                        EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass

                driver.find_element_by_xpath("//td[text()='"+FName+"']/a").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until(
                        EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass

            elif ClientPresentxl=="True":
                print("Client is already present in reference doc. Here is the details")
                print("First name is: "+FirstNamexl)
                print("Last name is: " + LastNamexl)
                TestResult.append("Client is already present in reference doc. Here is the details\nFirst name is: "+FirstNamexl+", Last name is: "+LastNamexl)
                TestResultStatus.append("Pass")
                try:
                    driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[3]/a/i").click()
                    time.sleep(2)
                    driver.find_element_by_xpath("//td[text()='"+FirstNamexl+"']/a").click()
                except Exception:
                    print("Client is present in reference excel but not found in application. We need to add new client")
                    TestResult.append(
                        "Client is present in reference excel but not found in application. We need to add new client")
                    TestResultStatus.append("Pass")

                    #------------Adding new client in application------------
                    try:
                        driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[3]/a/i").click()
                        time.sleep(2)
                        Records = driver.find_elements_by_xpath("//div[@class='datatable-scroll']/table/tbody/tr")
                        RowsLength = len(Records)
                        if RowsLength > 50:
                            print("We need to add code for Pagination")

                        for aa in range(5):
                            letters = string.ascii_lowercase
                            returna = ''.join(random.choice(letters) for i in range(5))
                            FName = returna
                        print(FName)

                        LName = "test"
                        print(LName)

                        driver.find_element_by_xpath("//a[@data-target='#createnewclient']").click()
                        try:
                            WebDriverWait(driver, SHORT_TIMEOUT
                                          ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                            WebDriverWait(driver, LONG_TIMEOUT
                                          ).until(
                                EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                        except TimeoutException:
                            pass

                        Data = [FName, LName, "TReferTo", "01-02-1990", "2480", D1, "1122334455", "TStreet", "123",
                                "TSuburb",
                                "@test.com", "213243", "1000", "TestSupport", "TestCommunication"]
                        for i2 in range(1, 29):
                            # -------------Client Status dropdown--------------------------------------------
                            if i2 == 1:
                                time.sleep(1)
                                select = Select(driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[1]/div/select"))
                                select.select_by_visible_text("Active")
                                TestResult.append(
                                    "Client 'status' is selected successfully")
                                TestResultStatus.append("Pass")
                            # -------------First Name--------------------------------------------
                            elif i2 == 2:
                                time.sleep(1)
                                driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                        i2) + "]/div/input").send_keys(
                                    Data[0])
                                TestResult.append(
                                    "Client 'first name' is entered successfully")
                                TestResultStatus.append("Pass")
                            # -------------Last Name--------------------------------------------
                            elif i2 == 3:
                                time.sleep(1)
                                driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                        i2) + "]/div/input").send_keys(
                                    Data[1])
                                TestResult.append(
                                    "Client 'last name' is entered successfully")
                                TestResultStatus.append("Pass")
                            # -------------Referred To By--------------------------------------------
                            elif i2 == 4:
                                time.sleep(1)
                                driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                        i2) + "]/div/input").send_keys(
                                    Data[2])
                                TestResult.append(
                                    "'Referred to by' is entered successfully")
                                TestResultStatus.append("Pass")
                            # -------------DOB--------------------------------------------
                            elif i2 == 5:
                                time.sleep(1)
                                driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                        i2) + "]/div/input").send_keys(
                                    Data[3])
                                ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                                TestResult.append("'DOB' is entered successfully")
                                TestResultStatus.append("Pass")
                            # -------------Gender--------------------------------------------
                            elif i2 == 6:
                                time.sleep(1)
                                select = Select(driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[6]/div/select"))
                                select.select_by_visible_text("Male")
                                TestResult.append(
                                    "'Gender' is selected successfully")
                                TestResultStatus.append("Pass")
                            # -------------NDIS Number--------------------------------------------
                            elif i2 == 9:
                                time.sleep(1)
                                NDISNumToUSe = int(Data[4])
                                for NDISNum in range(1000):
                                    driver.find_element_by_xpath(
                                        "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                            i2) + "]/div/input").send_keys(
                                        str(NDISNumToUSe))
                                    ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                                    try:
                                        WebDriverWait(driver, SHORT_TIMEOUT
                                                      ).until(
                                            EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                                        WebDriverWait(driver, LONG_TIMEOUT
                                                      ).until(
                                            EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                                    except TimeoutException:
                                        pass
                                    time.sleep(1)
                                    NdisError = driver.find_element_by_xpath(
                                        "//span[@id='error_ndIs']").is_displayed()
                                    if NdisError == True:
                                        driver.find_element_by_xpath(
                                            "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                                i2) + "]/div/input").clear()
                                        pass
                                    elif NdisError == False:
                                        break
                                    NDISNumToUSe = NDISNumToUSe + 1
                                TestResult.append(
                                    "'NDIS number' is entered successfully")
                                TestResultStatus.append("Pass")
                            # -------------Sign Up Date--------------------------------------------
                            elif i2 == 10:
                                time.sleep(1)
                                driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                        i2) + "]/div/input").send_keys(
                                    Data[5])
                                ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                                TestResult.append(
                                    "'Sign up date' is entered successfully")
                                TestResultStatus.append("Pass")
                            # -------------Mobile Number--------------------------------------------
                            elif i2 == 11:
                                time.sleep(1)
                                driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                        i2) + "]/div/input").send_keys(
                                    Data[6])
                                TestResult.append(
                                    "'Mobile number' is entered successfully")
                                TestResultStatus.append("Pass")
                            # -------------Street Address--------------------------------------------
                            elif i2 == 12:
                                time.sleep(1)
                                driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                        i2) + "]/div/input").send_keys(
                                    Data[7])
                                TestResult.append(
                                    "'Street address' is entered successfully")
                                TestResultStatus.append("Pass")
                            # -------------Home Number--------------------------------------------
                            elif i2 == 13:
                                time.sleep(1)
                                driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                        i2) + "]/div/input").send_keys(
                                    Data[8])
                                TestResult.append(
                                    "'Home Number' is entered successfully")
                                TestResultStatus.append("Pass")
                            # -------------Suburb--------------------------------------------
                            elif i2 == 14:
                                time.sleep(1)
                                driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                        i2) + "]/div/input").send_keys(
                                    Data[9])
                                TestResult.append(
                                    "'Suburb' is entered successfully")
                                TestResultStatus.append("Pass")
                            # -------------Email Address--------------------------------------------
                            elif i2 == 15:
                                time.sleep(1)
                                driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                        i2) + "]/div/input").send_keys(Data[0] +
                                                                       Data[10])
                                TestResult.append(
                                    "'Email Address' is entered successfully")
                                TestResultStatus.append("Pass")
                            # -------------State--------------------------------------------
                            elif i2 == 16:
                                time.sleep(1)
                                select = Select(driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/select"))
                                select.select_by_visible_text("SA")
                                TestResult.append(
                                    "'State' is selected successfully")
                                TestResultStatus.append("Pass")
                            # -------------Access to App--------------------------------------------
                            elif i2 == 17:
                                time.sleep(1)
                                select = Select(driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/select"))
                                select.select_by_visible_text("Yes")
                                TestResult.append(
                                    "'Access to app' is selected successfully")
                                TestResultStatus.append("Pass")
                            # -------------Postcode--------------------------------------------
                            elif i2 == 18:
                                time.sleep(1)
                                driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                        i2) + "]/div/input").send_keys(
                                    Data[11])
                                TestResult.append(
                                    "'Postcode' is entered successfully")
                                TestResultStatus.append("Pass")
                            # -------------Profile Type--------------------------------------------
                            elif i2 == 19:
                                time.sleep(1)
                                select = Select(driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/select"))
                                select.select_by_visible_text("Plan Managed")
                                TestResult.append(
                                    "'Profile Type' is selected successfully")
                                TestResultStatus.append("Pass")
                            # -------------Monthly Fee Rate ($)--------------------------------------------
                            elif i2 == 20:
                                time.sleep(1)
                                driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                        i2) + "]/div/input").send_keys(
                                    Data[12])
                                TestResult.append(
                                    "'Monthly Fee Rate ($)' is entered successfully")
                                TestResultStatus.append("Pass")
                            # -------------Support Coordinator--------------------------------------------
                            elif i2 == 21:
                                time.sleep(1)
                                driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                        i2) + "]/div/input").send_keys(
                                    Data[13])
                                TestResult.append(
                                    "'Support Coordinator' is entered successfully")
                                TestResultStatus.append("Pass")
                            # -------------Main Profile Contact--------------------------------------------
                            elif i2 == 22:
                                time.sleep(1)
                                select = Select(driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/select"))
                                select.select_by_visible_text("Yes")
                                TestResult.append(
                                    "'Main Profile Contact' is selected successfully")
                                TestResultStatus.append("Pass")
                            # -------------Receive Payment Updates--------------------------------------------
                            elif i2 == 23:
                                time.sleep(1)
                                select = Select(driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/select"))
                                select.select_by_visible_text("Yes")
                                TestResult.append(
                                    "'Receive Payment Updates' is selected successfully")
                                TestResultStatus.append("Pass")
                            # -------------Statement Preference--------------------------------------------
                            elif i2 == 24:
                                time.sleep(1)
                                select = Select(driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/select"))
                                select.select_by_visible_text("No")
                                TestResult.append(
                                    "'Statement Preference' is selected successfully")
                                TestResultStatus.append("Pass")
                            # -------------Communication Preferences--------------------------------------------
                            elif i2 == 26:
                                time.sleep(1)
                                driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                        i2) + "]/div/textarea").send_keys(
                                    Data[14])
                                TestResult.append(
                                    "'Communication Preferences' is entered successfully")
                                TestResultStatus.append("Pass")
                            # -------------NDIS Rate--------------------------------------------
                            elif i2 == 27:
                                time.sleep(1)
                                select = Select(driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(i2) + "]/div/select"))
                                select.select_by_visible_text("National Remote")
                                TestResult.append(
                                    "'NDIS Rate' is selected successfully")
                                TestResultStatus.append("Pass")
                            # -------------Save button--------------------------------------------
                            elif i2 == 28:
                                time.sleep(1)
                                driver.find_element_by_xpath(
                                    "//div[@id='createnewclient']/div/div/div[2]/form/div[28]/button").click()
                                TestResult.append(
                                    "Save button clicked")
                                TestResultStatus.append("Pass")
                                try:
                                    WebDriverWait(driver, SHORT_TIMEOUT
                                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                                    WebDriverWait(driver, LONG_TIMEOUT
                                                  ).until(
                                        EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                                except TimeoutException:
                                    pass
                                time.sleep(2)
                                try:
                                    EmailError = driver.find_element_by_xpath(
                                        "//span[@id='error_email_address']").is_displayed()
                                    if EmailError == True:
                                        time.sleep(1)
                                        driver.find_element_by_xpath(
                                            "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                                15) + "]/div/input").clear()
                                        driver.find_element_by_xpath(
                                            "//div[@id='createnewclient']/div/div/div[2]/form/div[" + str(
                                                15) + "]/div/input").send_keys(Data[0] + Data[1] + Data[10])
                                except Exception as ec:
                                    pass
                        TestResult.append("Create new client process is working correctly")
                        TestResultStatus.append("Pass")
                    except Exception:
                        TestResult.append("Create new client process is not working")
                        TestResultStatus.append("Fail")

                    # --------Saving client details in reference sheet------------
                    sheetx1.cell(1, 1).value = FName
                    sheetx1.cell(1, 2).value = LName
                    wbx1.save(locx1)

                    print("New client First name is: " + FName)
                    print("New client Last name is: " + LName)

                    # ------------Going back to client listing after creating client---------
                    try:
                        driver.find_element_by_xpath("//a[text()='Back']").click()
                        try:
                            WebDriverWait(driver, SHORT_TIMEOUT
                                          ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                            WebDriverWait(driver, LONG_TIMEOUT
                                          ).until(
                                EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                        except TimeoutException:
                            pass

                        driver.find_element_by_xpath("//td[text()='" + FName + "']/a").click()
                        try:
                            WebDriverWait(driver, SHORT_TIMEOUT
                                          ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                            WebDriverWait(driver, LONG_TIMEOUT
                                          ).until(
                                EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                        except TimeoutException:
                            pass
                    except Exception:
                        pass

            #------------Checking plan details once Client found in application--------------
            RemainingAmountLimit = 500
            PlanStartDate = "07-02-2022"
            PlanEndDate = "17-02-2022"

            try:
                PlanStatus = driver.find_element_by_xpath("//tbody/tr[@class='MagentaColorTR']/td[2]").text
            except Exception:
                PlanStatus = "No Plan Found"

            if PlanStatus == "No Plan Found":
                print("Inside " + PlanStatus)
                TestResult.append("No Plan Found, creating new plan")
                TestResultStatus.append("Pass")
                driver.find_element_by_xpath("//a[@id='addNewServicePlan']").click()
                try:
                    for np in range(1, 9):
                        driver.find_element_by_xpath(
                            "//div[@id='UploadNewPlan']/div/div/div[2]/form/div[" + str(np) + "]")
                        # -----------------Plan start date-------------------------------------------
                        if np == 5:
                            time.sleep(1)
                            driver.find_element_by_xpath("//div[@id='UploadNewPlan']/div/div/div[2]/form/div[" + str(
                                np) + "]/div/input[1]").send_keys(PlanStartDate)
                            ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                            TestResult.append(
                                "Plan start date is selected")
                            TestResultStatus.append("Pass")
                        # -----------------Plan end date-------------------------------------------
                        elif np == 6:
                            time.sleep(1)
                            driver.find_element_by_xpath("//div[@id='UploadNewPlan']/div/div/div[2]/form/div[" + str(
                                np) + "]/div/input[1]").send_keys(PlanEndDate)
                            ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                            TestResult.append(
                                "Plan end date is selected")
                            TestResultStatus.append("Pass")
                        # -----------------Status dropdown-------------------------------------------
                        elif np == 7:
                            time.sleep(1)
                            select = Select(driver.find_element_by_xpath(
                                "//div[@id='UploadNewPlan']/div/div/div[2]/form/div[" + str(np) + "]/div/select"))
                            select.select_by_visible_text("Active")
                            TestResult.append(
                                "Plan status is selected")
                            TestResultStatus.append("Pass")
                        # -----------------Create button-------------------------------------------
                        elif np == 8:
                            time.sleep(1)
                            driver.find_element_by_xpath(
                                "//div[@id='UploadNewPlan']/div/div/div[2]/form/div[" + str(np) + "]/button").click()
                            TestResult.append(
                                "Create button is clicked")
                            TestResultStatus.append("Pass")
                            try:
                                WebDriverWait(driver, SHORT_TIMEOUT
                                              ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                                WebDriverWait(driver, LONG_TIMEOUT
                                              ).until(
                                    EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                            except TimeoutException:
                                pass
                    TestResult.append("Adding new plan process is working correctly")
                    TestResultStatus.append("Pass")
                except Exception:
                    TestResult.append("Adding new plan process is not working")
                    TestResultStatus.append("Fail")


                PlanStatus = "Active"

            if PlanStatus == "InActive":
                print("Inside " + PlanStatus)
                print("We need to edit existing plan")
                TestResult.append(
                    "Plan Found inactive, activating the plan")
                TestResultStatus.append("Pass")
                try:
                    driver.find_element_by_xpath("//tbody/tr/td[@class='text-right']/a[2]").click()
                    select = Select(driver.find_element_by_xpath(
                        "//select[@class='servicePlanStatus']"))
                    select.select_by_visible_text("Active")
                    driver.find_element_by_xpath(
                        "//div[@id='UploadNewPlan']/div/div/div[2]/form/div[10]/button").click()
                    TestResult.append("Plan is activated successfully")
                    TestResultStatus.append("Pass")
                except Exception:
                    TestResult.append("Not able to activate plan")
                    TestResultStatus.append("Fail")
                PlanStatus = "Active"

            print("Once plan status is set / found Active, now checking plan remianing amount")
            TestResult.append("Once plan status is set / found Active, now checking plan remianing amount")
            TestResultStatus.append("Pass")
            if PlanStatus == "Active":
                TestResult.append("Plan status found active, No need to add new plan")
                TestResultStatus.append("Pass")
                try:
                    PlanPresent = driver.find_element_by_xpath(
                        "//td[@class='ServiceBookingTdCol']/table/tbody/tr/td[1]/p/a").text
                    print(PlanPresent)
                except Exception:
                    try:
                        # --------------clicking on Add Plan Managed Service Booking-----------
                        l = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
                        random.shuffle(l)
                        if l[0] == 0:
                            pos = random.choice(range(1, len(l)))
                            l[0], l[pos] = l[pos], l[0]
                        BookingNumber = ''.join(map(str, l[0:4]))
                        AllocatedAmount = "1000"

                        driver.find_element_by_xpath("//tbody/tr/td[@class='TrButtonAdd']/button[2]").click()
                        TestResult.append("Add Plan Managed Service Booking is clicked")
                        TestResultStatus.append("Pass")
                        for pm in range(1, 8):
                            # ----------Booking number field on plan managed service booking page--------------------
                            if pm == 1:
                                time.sleep(1)
                                driver.find_element_by_xpath(
                                    "//div[@id='AddServiceBooking']/div/div/div[2]/form[1]/div[" + str(
                                        pm) + "]/div/div[2]/ul/li/span[1]/input").send_keys(BookingNumber)
                                TestResult.append("Booking number is entered")
                                TestResultStatus.append("Pass")
                            # ----------Support Budget dropdown on plan managed service booking page--------------------
                            elif pm == 3:
                                time.sleep(1)
                                select = Select(driver.find_element_by_xpath(
                                    "//div[@id='AddServiceBooking']/div/div/div[2]/form[1]/div[" + str(
                                        pm) + "]/div[1]/div/select"))
                                select.select_by_index(4)
                                TestResult.append("Support Budget is selected")
                                TestResultStatus.append("Pass")
                            # ----------Allocated Amount (Unit Price) field on plan managed service booking page--------------------
                            elif pm == 5:
                                time.sleep(1)
                                driver.find_element_by_xpath(
                                    "//div[@id='AddServiceBooking']/div/div/div[2]/form[1]/div[" + str(
                                        pm) + "]/div/input").send_keys(AllocatedAmount)
                                TestResult.append("Allocated Amount (Unit Price) is entered")
                                TestResultStatus.append("Pass")

                        for pm1 in range(1, 10):
                            # ----------Add button on plan managed service booking page--------------------
                            driver.find_element_by_xpath("//button[text()='Add']").click()
                            TestResult.append("Add button is clicked")
                            TestResultStatus.append("Pass")
                            try:
                                WebDriverWait(driver, SHORT_TIMEOUT
                                              ).until(
                                    EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                                WebDriverWait(driver, LONG_TIMEOUT
                                              ).until(
                                    EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                            except TimeoutException:
                                pass

                            try:
                                BookingNumberError = driver.find_element_by_xpath(
                                    "//span[@id='error_booking_number']").is_displayed()
                                if BookingNumberError == True:
                                    time.sleep(1)
                                    BookingNumber = BookingNumber + 1
                                    driver.find_element_by_xpath(
                                        "//div[@id='AddServiceBooking']/div/div/div[2]/form[1]/div[1]/div/div[2]/ul/li/span[1]/input").clear()
                                    driver.find_element_by_xpath(
                                        "//div[@id='AddServiceBooking']/div/div/div[2]/form[1]/div[1]/div/div[2]/ul/li/span[1]/input").send_keys(
                                        BookingNumber)
                                elif BookingNumberError == False:
                                    break
                            except Exception:
                                pass

                        driver.find_element_by_xpath("//button[@class='upload_btn_plan btn_clr_gr']").click()
                        try:
                            WebDriverWait(driver, SHORT_TIMEOUT
                                          ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                            WebDriverWait(driver, LONG_TIMEOUT
                                          ).until(
                                EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                        except TimeoutException:
                            pass
                        TestResult.append("Add Plan Managed Service Booking process is working correctly")
                        TestResultStatus.append("Pass")
                    except Exception:
                        TestResult.append("Add Plan Managed Service Booking process is not working")
                        TestResultStatus.append("Fail")
                try:
                    Remaining = driver.find_element_by_xpath(
                        "//td[@class='ServiceBookingTdCol']/table/tbody/tr/td[3]").text
                    print(Remaining)

                    BalanceAmt = Remaining
                    rem = 0
                    try:
                        ind = BalanceAmt.index('.')
                        try:
                            if BalanceAmt[ind + 1] and BalanceAmt[ind + 2] == "0":
                                rem = 1
                        except Exception as qq:
                            print(qq)
                            if BalanceAmt[ind + 1] == "0":
                                rem = 1
                            pass
                        ab = re.findall('[^A-Za-z0-9]+', BalanceAmt)
                        ab = int(len(ab))
                        ab = ab - 1
                        BalanceAmt = re.sub('[^A-Za-z0-9]+', '', BalanceAmt)
                        if BalanceAmt=="000":
                            BalanceAmt="0"
                        else:
                            BalanceAmt = BalanceAmt[:ind - ab] + "." + BalanceAmt[ind - ab:]
                            if rem == 1:
                                BalanceAmt = BalanceAmt.strip('.').strip('0').strip('0')
                                BalanceAmt = BalanceAmt.strip('.')
                                BalanceAmt = int(BalanceAmt)
                            print(BalanceAmt)
                    except Exception as rr:
                        BalanceAmt = re.sub('[^A-Za-z0-9]+', '', BalanceAmt)
                        BalanceAmt = int(BalanceAmt)
                        print(BalanceAmt)
                except Exception:
                    pass

            print(float(BalanceAmt))
            if float(BalanceAmt) < float(RemainingAmountLimit):
                print("Inside " + PlanStatus+ " and Balance Amt less than Remaining Amount Limit")
                TestResult.append("Remaining amount found less than remaining amount limit\n Need to add some amount in allocated amount to maintain remaining amount")
                TestResultStatus.append("Pass")
                try:
                    driver.find_element_by_xpath("//td[@class='ServiceBookingTHwidth']/p/a").click()
                    try:
                        WebDriverWait(driver, SHORT_TIMEOUT
                                      ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                        WebDriverWait(driver, LONG_TIMEOUT
                                      ).until(
                            EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                    except TimeoutException:
                        pass
                    try:
                        driver.find_element_by_xpath("//div[@class='src_sec_hed src_sec_hed_cus']/a[1]").click()
                        time.sleep(2)
                        AmtToEdit = driver.find_element_by_xpath(
                            "//input[@name='booking[0][amount]']").get_attribute('value')
                        AmtToEdit = float(AmtToEdit)
                        print(AmtToEdit)

                        AmtNeedToAdd = float(RemainingAmountLimit) - float(BalanceAmt)
                        AmtNeedToAdd = float(AmtNeedToAdd)
                        print(AmtNeedToAdd)

                        NewAmount = AmtToEdit + AmtNeedToAdd
                        NewAmount = float(NewAmount)
                        print(NewAmount)
                        driver.find_element_by_xpath("//input[@name='booking[0][amount]']").clear()
                        time.sleep(2)
                        driver.find_element_by_xpath("//input[@name='booking[0][amount]']").send_keys(NewAmount)
                        driver.find_element_by_xpath("//button[@id='updateAllocationSubmitBtn']").click()
                        try:
                            WebDriverWait(driver, SHORT_TIMEOUT
                                          ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                            WebDriverWait(driver, LONG_TIMEOUT
                                          ).until(
                                EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                        except TimeoutException:
                            pass
                        driver.find_element_by_xpath("//button[@class='btn btn-primary checkOverBudgetBtnEvent']").click()
                        try:
                            WebDriverWait(driver, SHORT_TIMEOUT
                                          ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                            WebDriverWait(driver, LONG_TIMEOUT
                                          ).until(
                                EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                        except TimeoutException:
                            pass
                        TestResult.append("Amount added in allocated amount to maintain remaining amount")
                        TestResultStatus.append("Pass")
                    except Exception:
                        TestResult.append("Not able to add amount in allocated amount to maintain remaining amount")
                        TestResultStatus.append("Fail")

                except Exception:
                    pass

            sheetx1.cell(1, 3).value = PlanEndDate
            wbx1.save(locx1)
            # -------------------------------------------------------------------------------------------

        except Exception as err:
            print(err)
            TestResult.append("Invoice entry is not working correctly. Below error found\n"+str(err))
            TestResultStatus.append("Fail")
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


