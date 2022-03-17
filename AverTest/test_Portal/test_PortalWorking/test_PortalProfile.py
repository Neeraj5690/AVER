import datetime
import math
import re
import time
import openpyxl
from datetime import datetime,date
import datetime as datetime
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from sys import platform
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import pyperclip
import random
import os


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("email").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("password").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_PortalProfile"
  description = "This test scenario is to verify content at Profile page of Client Portal"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_PortalWorking"
  global Exe
  Exe="Yes"
  Directory = 'test_Portal/'

  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'D:/AVER/AverTest/' + Directory

  MachineName = os.getenv('COMPUTERNAME')
  if MachineName == "DESKTOP-JLLTS65":
      path = path.replace('D:', 'C:')

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      if platform == "linux" or platform == "linux2":
          driver = webdriver.Chrome(executable_path="/home/legion/office 1wayit/AVER/AverTest/chrome/chromedriverLinux1")
      elif platform == "win32" or platform == "win64":
          if MachineName == "DESKTOP-JLLTS65":
              driver = webdriver.Chrome(executable_path="C:/AVER/AverTest/chrome/chromedriver.exe")
          else:
              driver = webdriver.Chrome(executable_path="D:/AVER/AverTest/chrome/chromedriver.exe")

      driver.implicitly_wait(10)

  yield
  if Exe == "Yes":
      time_change = datetime.timedelta(hours=5)
      new_time = datetime.datetime.now() + time_change
      ctReportHeader = new_time.strftime("%d %B %Y %I %M%p")

      ct = new_time.strftime("%d_%B_%Y_%I_%M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      #driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 3
        LONG_TIMEOUT = 60
        LOADING_ELEMENT_XPATH = "//body[@class='sidebar-xs loader_overlay']"
        try:
            print()
            # ----------------Fecthing Client name from the ref Data sheet--------------------
            ExcelFileName2 = "RefData"
            locx2 = (path + 'Ref/' + ExcelFileName2 + '.xlsx')
            wbx2 = openpyxl.load_workbook(locx2)
            sheetx2 = wbx2.active

            try:
                UsernameNameXL = sheetx2.cell(1, 4).value
                print(UsernameNameXL)
                PasswordXL = sheetx2.cell(1,5).value
                print(PasswordXL)
                if UsernameNameXL == None or PasswordXL == None:
                    print("Username and / or Password not found in ref sheet")
                    driver.close()
                else:
                    # ------------Login to Client Portal----------------
                    driver.maximize_window()
                    driver.get("https://averreplica.1wayit.com/login")
                    enter_username(UsernameNameXL)
                    enter_password(PasswordXL)
                    driver.find_element_by_xpath("//button[@type='submit']").click()
                    time.sleep(2)
                    try:
                        LoginError=driver.find_element_by_xpath("//span[@class='invalid-feedback']/strong").text
                        print("User is not able to login. Below error found\n"+LoginError)

                    except Exception:
                        pass
            except Exception:
                print("Ref sheet is not able to read, please check the ref doc sheet")
                driver.close()

            try:
                FLNameXL = sheetx2.cell(3, 1).value
                print(FLNameXL)
                NDISXL = sheetx2.cell(3,2).value
                print(NDISXL)
                PhoneXL = sheetx2.cell(3, 3).value
                print(PhoneXL)
                EmailXL = sheetx2.cell(3, 4).value
                print(EmailXL)
                AddressXL = sheetx2.cell(3, 5).value
                print(AddressXL)
                PlanNameXL = sheetx2.cell(3, 6).value
                print(PlanNameXL)

                AddContactCountXL = sheetx2.cell(4, 2).value
                FirstNameListXL = []
                LastNameListXL = []
                RelationListXL = []
                try:
                    AddContactCountXL=int(AddContactCountXL)
                    print(AddContactCountXL)
                    if AddContactCountXL>0:
                        for ac in range(AddContactCountXL):
                            print("ac is "+str(ac))

                            FirstNameXL = sheetx2.cell(ac+5, 1).value
                            FirstNameListXL.append(FirstNameXL)

                            LastNameXL = sheetx2.cell(ac + 5, 2).value
                            LastNameListXL.append(LastNameXL)

                            RelationXL = sheetx2.cell(ac + 5, 3).value
                            RelationListXL.append(RelationXL)

                except Exception:
                    AddContactCountXL=0
                    print("Ref sheet has invalid Additional contact count, please check the ref doc sheet")
                    pass

                if FLNameXL == None or NDISXL == None or PhoneXL == None or EmailXL == None or AddressXL == None or PlanNameXL == None:
                    print("Client details - First name, Last name, NDIS, Email, or Phone number does not found in ref sheet")
                    driver.close()
            except Exception:
                print("Ref sheet is not able to read, please check the ref doc sheet")
                driver.close()

            # ---------------------------Verify Client Portal Profile Data-----------------------------
            PageName = "Client Portal Profile page"
            TitleExpected="Profile"
            try:
                driver.find_element_by_xpath("//div[@class='hed_wth_srch']/a").click()
                time.sleep(1)
                driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[2]").click()

                TitleFound=driver.find_element_by_xpath("//h2[text()='Profile']").text
                time.sleep(2)
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break

                time.sleep(2)
                if TitleFound==TitleExpected:
                    print(PageName + " opened successfully")
                    TestResult.append(PageName + " opened successfully")
                    TestResultStatus.append("Pass")
                else:
                    TestResult.append(PageName + " is not able to open")
                    TestResultStatus.append("Fail")
            except Exception as ee:
                print(ee)
                TestResult.append(PageName + " is not able to open")
                TestResultStatus.append("Fail")

            # ------------------------Fetching Data present at Profile page --------------
            FoundUserName=driver.find_element_by_xpath("//label[text()='User Name']/parent::div/p").text
            print(FoundUserName)

            FoundNDIS = driver.find_element_by_xpath("//label[text()='NDIS Number']/parent::div/p").text
            print(FoundNDIS)

            FoundEmailAddress = driver.find_element_by_xpath("//label[text()='Email Address']/parent::div/p").text
            print(FoundEmailAddress)

            FoundPlanStatus = driver.find_element_by_xpath("//label[text()='Plan Status ']/parent::div/p/span/a").text
            print(FoundPlanStatus)

            FoundContactNumber = driver.find_element_by_xpath("//label[text()='Contact Number']/parent::div/p").text
            print(FoundContactNumber)

            FoundUserAddress = driver.find_element_by_xpath("//label[text()='User Address']/parent::div/p").text
            print(FoundUserAddress)

            # ------------------------Verify Data present at Profile page --------------
            if FLNameXL!=FoundUserName:
                print("Client name at client portal (Profile) does not match with client name at admin portal")
            else:
                print("Client name at client portal (Profile) matched with client name at admin portal")

            if NDISXL!=FoundNDIS:
                print("NDIS at client portal (Profile) does not match with NDIS at admin portal")
            else:
                print("NDIS at client portal (Profile) matched with NDIS at admin portal")

            if PhoneXL!=FoundContactNumber:
                print("Mobile Number at client portal (Profile) does not match with Mobile Number at admin portal")
            else:
                print("Mobile Number at client portal (Profile) matched with Mobile Number at admin portal")

            if EmailXL!=FoundEmailAddress:
                print("Email at client portal (Profile) does not match with Email at admin portal")
            else:
                print("Email at client portal (Profile) matched with Email at admin portal")

            if AddressXL!=FoundUserAddress:
                print("User Address at client portal (Profile) does not match with User Address at admin portal")
            else:
                print("User Address at client portal (Profile) matched with User Address at admin portal")

            if PlanNameXL not in FoundPlanStatus:
                print("Plan Status at client portal (Profile) does not match with Plan Status at admin portal")
            else:
                print("Plan Status at client portal (Profile) matched with Plan Status at admin portal")

            # -----------Clearing Client data from the ref sheet-------------------
            sheetx2.cell(3, 1).value = None
            sheetx2.cell(3, 2).value = None
            sheetx2.cell(3, 3).value = None
            sheetx2.cell(3, 4).value = None
            sheetx2.cell(3, 5).value = None
            sheetx2.cell(3, 6).value = None
            wbx2.save(locx2)

            #---------------------------Fetching Additonal contact details-----------------------------
            if AddContactCountXL>0:
                ACFirstNameList = []
                ACLastNameList = []
                ACRelationList = []

                AddContCount = driver.find_elements_by_xpath("//table[@id='contacts-list']/tbody/tr")
                if len(AddContCount) == 1:
                    print("No additional Contact present")
                elif len(AddContCount) > 1:
                    print("Additional Contact present")
                    print(len(AddContCount))
                    for c in range(len(AddContCount)):
                        print()
                        print("c is " + str(c))
                        ACFirstName = driver.find_element_by_xpath(
                            "//table[@id='contacts-list']/tbody/tr[" + str(c + 1) + "]/td[2]").text
                        print(ACFirstName)
                        ACFirstNameList.append(ACFirstName)

                        ACLastName = driver.find_element_by_xpath(
                            "//table[@id='contacts-list']/tbody/tr[" + str(c + 1) + "]/td[3]").text
                        print(ACLastName)
                        ACLastNameList.append(ACLastName)

                        ACRelation = driver.find_element_by_xpath(
                            "//table[@id='contacts-list']/tbody/tr[" + str(c + 1) + "]/td[4]").text
                        print(ACRelation)
                        ACRelationList.append(ACRelation)
                print(ACFirstNameList)
                print(ACLastNameList)
                print(ACRelationList)

                print(FirstNameListXL)
                print(LastNameListXL)
                print(RelationListXL)

                # ------------------------Verify Additional contact data at Profile page --------------
                if FirstNameListXL != ACFirstNameList:
                    print("Additional contact first name at client portal (Profile) does not match with Additional contact first name at admin portal")
                else:
                    print("Additional contact first name at client portal (Profile) matched with Additional contact first name at admin portal")

                if LastNameListXL != ACLastNameList:
                    print("Additional contact last name at client portal (Profile) does not match with Additional contact last name at admin portal")
                else:
                    print("Additional contact last name at client portal (Profile) matched with Additional contact last name at admin portal")

                if RelationListXL != ACRelationList:
                    print("Additional contact relation at client portal (Profile) does not match with Additional contact relation at admin portal")
                else:
                    print("Additional contact relation at client portal (Profile) matched with Additional contact relation at admin portal")

                #-----------Clearing additional contact data from the ref sheet-------------------
                for ac1 in range(AddContactCountXL):
                    sheetx2.cell(ac1 + 5, 1).value=None
                    sheetx2.cell(ac1 + 5, 2).value=None
                    sheetx2.cell(ac1 + 5, 3).value=None

                #-------Clearing additional contact count data from the ref sheet-------------------
                sheetx2.cell(4, 2).value = None
                wbx2.save(locx2)


            else:
                print("No additional contact present in client profile page")


        except Exception as err:
            print(err)
            TestResult.append("Portal is not working correctly. Below error found\n"+str(err))
            TestResultStatus.append("Fail")
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


