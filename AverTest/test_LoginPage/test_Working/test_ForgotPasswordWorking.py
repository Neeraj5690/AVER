import datetime
import os
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
import imaplib
from sys import platform

@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("email").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("password").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_ForgotPasswordWorking"
  description = "This test scenario is to verify working of Forgot Password"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_Working"
  global Exe
  Exe="Yes"

  p = os.path.dirname(os.path.realpath(__file__))
  p = os.path.abspath(os.path.join(p, '..'))
  print(p)
  p = p.replace('\\', '/')
  p = p + "/"
  print(p)

  path = p
  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      if platform == "linux" or platform == "linux2":
          driver = webdriver.Chrome(executable_path="/home/legion/office 1wayit/AVER/AverTest/chrome/chromedriverLinux")
      elif platform == "win32" or platform == "win64":
          driver = webdriver.Chrome(executable_path=path+"C:/AVER/AverTest/chrome/chromedriver.exe")

      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://averreplica.1wayit.com/login")

  yield
  if Exe == "Yes":
      ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
      ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 5
        LONG_TIMEOUT = 400
        LOADING_ELEMENT_XPATH = "//div[@id='appian-working-indicator-hidden']"
        host = 'imap.gmail.com'
        username = 'teamqa59@gmail.com'
        password = 'Gagan@0309'
        ForgotPasswordEmail = username
        try:
            try:
               driver.find_element_by_xpath("//div[@class='card-body']/div[5]/a").click()
               TestResult.append("Forgot Password link text clicked")
               TestResultStatus.append("Pass")
            except Exception:
               TestResult.append("Forgot Password link text not able to click")
               TestResultStatus.append("Fail")
               driver.close()
            time.sleep(TimeSpeed)
            #---------------------------Enter Username Data-----------------------------
            PageName=ForgotPasswordEmail
            try:
                driver.find_element_by_xpath("//div[@class='card-body']/div[2]/input").send_keys(PageName)
                TestResult.append("Email address [ "+PageName + " ] entered successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append("Email address [ "+PageName + " ] not able to enter")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            #---------------------------------------------------------------------------------

            # ---------------------------Verify Submit Button-----------------------------
            PageName = "Submit Button"
            try:
                driver.find_element_by_xpath("//div[@class='card-body']/div[3]/button").click()
                TestResult.append(PageName + " clicked successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not click")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # -------------------Receiving forgot password email from Aver Planning--------------
            PageName="Forgot password email"
            try:
                # ------------------------Get verification code from Gmail---------------------------------
                # -------------Function to get email content part i.e its body part
                def get_body(msg):
                    if msg.is_multipart():
                        return get_body(msg.get_payload(0))
                    else:
                        return msg.get_payload(None, True)

                # -----------Function to search for a key value pair
                def search(key, value, con):
                    result, data = con.search(None, key, '"{}"'.format(value))
                    return data

                # ---------------Function to get the list of emails under this label
                def get_emails(result_bytes):
                    msgs = []  # all the email data are pushed inside an array
                    for num in result_bytes[0].split():
                        typ, data = con.fetch(num, '(RFC822)')
                        msgs.append(data)
                    return msgs

                con = imaplib.IMAP4_SSL(host)
                con.login(username, password)
                con.select('Inbox')

                # --------------fetching emails from a user
                msgs = get_emails(search('FROM', 'sohia.1wayit@gmail.com', con))
                Code = ""
                for msg in msgs[::-1]:
                    for sent in msg:
                        if Code != "":
                            break
                        else:
                            if type(sent) is tuple:
                                content = str(sent[1], 'utf-8')
                                data = str(content)
                                try:
                                    indexstart = data.find("ltr")
                                    data2 = data[indexstart + 5: len(data)]
                                    indexend = data2.find("</div>")
                                    indx = data2.find('AverPlanning')
                                    #print(data2)
                                    Code = data2[indx] +data2[indx + 1] + data2[indx + 2]+data2[indx + 3] + data2[indx + 4]+data2[indx + 5] + data2[indx + 6]+data2[indx + 7] + data2[indx + 8]+ data2[indx + 9]+ data2[indx + 10] + data2[indx + 11]
                                    print(Code)
                                    break
                                except UnicodeEncodeError as e:
                                    pass

                TestResult.append(PageName + " is received")
                TestResultStatus.append("Pass")
            except Exception as ww:
                print(ww)
                TestResult.append(PageName + " is not able to receive. Below error found\n"+ww)
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # -------------------Reading forgot password email from Aver Planning--------------
            PageName = "Forgot password email"
            try:

                if "AverPlanning" in Code:
                    TestResult.append(PageName + " is able to read")
                    TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not able to read")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

        except Exception:
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


