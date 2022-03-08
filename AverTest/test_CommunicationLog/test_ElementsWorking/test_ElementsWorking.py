import datetime
import math
import re
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from sys import platform
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import pyperclip


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("email").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("password").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_ElementsWorking"
  description = "This test scenario is to verify all the Working of Elements at Client Listing page"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_ElementsWorking"
  global Exe
  Exe="Yes"
  Directory = 'test_CommunicationLog/'
  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'D:/AVER/AverTest/' + Directory

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe=="Yes":
      if platform == "linux" or platform == "linux2":
          driver = webdriver.Chrome(executable_path="/home/legion/office 1wayit/AVER/AverTest/chrome/chromedriverLinux1")
      elif platform == "win32" or platform == "win64":
          driver = webdriver.Chrome(executable_path="D:/AVER/AverTest/chrome/chromedriver.exe")

      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://averreplica.1wayit.com/login")
      enter_username("admin@averplanning.com")
      enter_password("admin786")
      driver.find_element_by_xpath("//button[@type='submit']").click()

  yield
  if Exe == "Yes":
      time_change = datetime.timedelta(hours=5)
      new_time = datetime.datetime.now() + time_change
      ctReportHeader = new_time.strftime("%d %B %Y %I %M%p")

      ct = new_time.strftime("%d_%B_%Y_%I_%M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 3
        LONG_TIMEOUT = 60
        LOADING_ELEMENT_XPATH = "//div[@class='main-loader LoaderImageLogo']"
        try:
            # ---------------------------Verify Communication Log icon click-----------------------------
            PageName = "Communication log icon"
            Ptitle1 = ""
            try:
                driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[2]/a/i").click()
                time.sleep(TimeSpeed)
                TestResult.append(PageName + "  is present in left menu and able to click")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Working of Back button on Communication log page-----------------------------
            PageName = "Back button on Communication log page"
            Ptitle1 = "Rae"
            try:
                driver.find_element_by_xpath("//div[@class='content yellow_color']/div/div[2]/div/a").click()
                time.sleep(TimeSpeed)
                PageTitle1 = driver.find_element_by_xpath("//div[@class='hed_wth_srch']/h2").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not "
                TestResult.append(PageName + "  is clickable")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not clickable")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ----------------Verify Communication Log icon click after verifying back--------
            PageName = "Communication log icon"
            Ptitle1 = ""
            try:
                driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[2]/a/i").click()
                time.sleep(TimeSpeed)
                TestResult.append(PageName + "  is opened again after verifying back button")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not opened again after verifying back button")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------
            # ---------------------------Verify Select entry dropdown working-----------------------------
            for cv in range (5):
                DropdownValues = {"Email": "//select[@name='communication_type']/option","Letter": "//select[@name='communication_type']/option","SMS": "//select[@name='communication_type']/option    ","Phone Call": "//select[@name='communication_type']/option","New Plan Form": "//ul[@class='GeneralClientDetails']/li[1]"}
                select = Select(driver.find_element_by_xpath(
                    "//select[@id='selectTemplateOption']"))
                Selector=['Email','Letter','SMS','Phone Call','New Plan Form']
                if cv==0:
                    select.select_by_visible_text(Selector[3])
                    time.sleep(TimeSpeed)
                    TextCheck=Selector[3]
                    path1=DropdownValues[Selector[3]]
                elif cv==1:
                    select.select_by_visible_text(Selector[0])
                    time.sleep(TimeSpeed)
                    TextCheck = Selector[0]
                    path1 =DropdownValues[Selector[0]]
                elif cv==2:
                    select.select_by_visible_text(Selector[1])
                    time.sleep(TimeSpeed)
                    TextCheck = Selector[1]
                    path1 =DropdownValues[Selector[1]]
                elif cv==3:
                    select.select_by_visible_text(Selector[2])
                    time.sleep(TimeSpeed)
                    TextCheck = Selector[2]
                    path1 =DropdownValues[Selector[2]]
                elif cv==4:
                    select.select_by_visible_text(Selector[4])
                    time.sleep(TimeSpeed)
                    TextCheck = Selector[4]
                    path1 =DropdownValues[Selector[4]]


                textFound=driver.find_element_by_xpath(path1).text
                if ":" in textFound:
                    textFound=textFound.split(":")
                    textFound=textFound[1]

                textFound=textFound.strip()
                if textFound==TextCheck:
                    TestResult.append(TextCheck + " dropdown value inside select new entry dropdown is able to click and open")
                    TestResultStatus.append("Pass")
                else:
                    TestResult.append(TextCheck + " dropdown value inside select new entry dropdown is not able to click")
                    TestResultStatus.append("Fail")
                driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[2]/a/i").click()
                time.sleep(TimeSpeed)
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Choose template dropdown working-----------------------------
            for dct in range(2):
                DropdownCT = {"General": "//div[@class='content yellow_color']/div[2]/div[1]/div/div/h2/small",
                                  "New Plan Form": "//div[@class='content-wrapper']/div[2]/div[3]/div[1]/div/div/h2/small"}
                select = Select(driver.find_element_by_xpath(
                    "//div[@class='content yellow_color']/div[2]/div/div/div/form/div/div/select"))
                Selector = ['General', 'New Plan Form']
                if dct == 0:
                    select.select_by_index(1)
                    time.sleep(TimeSpeed)
                    TextCheck = Selector[1]
                    path1 = DropdownCT[Selector[1]]
                elif dct == 1:
                    select.select_by_index(0)
                    time.sleep(TimeSpeed)
                    TextCheck = Selector[0]
                    path1 = DropdownCT[Selector[0]]
                textFound = driver.find_element_by_xpath(path1).text

                if textFound == TextCheck:
                    TestResult.append(TextCheck + " dropdown value inside Choose Template dropdown is able to click")
                    TestResultStatus.append("Pass")
                else:
                    TestResult.append(TextCheck + " dropdown value inside Choose Template dropdown is not able to click and open")
                    TestResultStatus.append("Fail")
                driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[2]/a/i").click()
                time.sleep(TimeSpeed)
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Communication Log Listing - General pagination working-----------------------------
            try:
                TotalItem = driver.find_element_by_xpath("//div[@id='comm-log-blank_info']").text
                print(TotalItem)
                substr = "of"
                x = TotalItem.split(substr)
                string_name = x[0]
                TotalItemAfterOf = x[1]
                abc = ""
                countspace = 0
                for element in range(0, len(string_name)):
                    if string_name[(len(string_name) - 1) - element] == " ":
                        countspace = countspace + 1
                        if countspace == 2:
                            break
                    else:
                        abc = abc + string_name[(len(string_name) - 1) - element]
                abc = abc[::-1]
                TotalItemBeforeOf = abc
                TotalItemAfterOf = TotalItemAfterOf.split(" ")
                TotalItemAfterOf=TotalItemAfterOf[1]
                TotalItemAfterOf = re.sub('[^A-Za-z0-9]+', '', TotalItemAfterOf)

                TotalItemAfterOf = int(TotalItemAfterOf)
                RecordsPerPage=50
                TotalPages = TotalItemAfterOf/RecordsPerPage
                NumberOfPages = math.ceil(float(TotalPages))
                ClickCounter=0
                for i in range(NumberOfPages):
                    if i<1:
                        if i==NumberOfPages-1:
                            TestResult.append("No Pagination found for [ "+str(RecordsPerPage)+" ] no. of records under Communication Log Listing - General table")
                            TestResultStatus.append("Pass")
                            break
                    try:
                        time.sleep(TimeSpeed)
                        driver.find_element_by_xpath("//div[@class='dataTables_paginate paging_simple_numbers']/a[2]").click()
                        time.sleep(1)
                        ClickCounter=ClickCounter+1
                        TestResult.append("Pagination verified for [ " + str(
                            TotalItemAfterOf) + " ] no. of records under Communication Log Listing - General table")
                        TestResultStatus.append("Pass")
                    except Exception as cc:
                        pass
                if i != ClickCounter:
                    TestResult.append(
                        "Pagination for [ " + str(RecordsPerPage) + " ] no. of records is not working correctly")
                    TestResultStatus.append("Fail")
                driver.refresh()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until(EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
            except Exception as aq:
                TestResult.append("Pagination for [ " + str(RecordsPerPage) + " ] no. of records is not working correctly")
                TestResultStatus.append("Fail")

            # ---------------------------Verify Add new process by all Select entry dropdowns-----
            DropdownValues = {"Email": "","Letter": "","SMS": "","Phone Call": "","New Plan Form": ""}
            Selector=['Email','Letter','SMS','Phone Call','New Plan Form']

            #-------Verifying add new Email process-------------------------------------------------
            try:
                select = Select(driver.find_element_by_xpath(
                    "//div[@class='content yellow_color']/div[1]/div/form/div/div/div/select"))
                select.select_by_visible_text(Selector[0])
                driver.find_element_by_xpath("//input[@id='autocompleteSystemContactSearch']").send_keys("sumreet client")
                time.sleep(TimeSpeed)
                ActionChains(driver).key_down(Keys.DOWN).key_up(Keys.DOWN).perform()
                time.sleep(TimeSpeed)
                ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                time.sleep(TimeSpeed)
                driver.find_element_by_xpath("//div[@id='communication-log-form']/form/div[3]/div/div[2]/div/div[3]/div[1]/div[2]/div/input").send_keys("Test Subject")
                time.sleep(TimeSpeed)
                select = Select(driver.find_element_by_xpath("//div[@id='communication-log-form']/form/div[3]/div/div[5]/div/div[2]/div/div[1]/div/select"))
                select.select_by_index(1)
                time.sleep(TimeSpeed)
                driver.find_element_by_xpath("//button[text()='Save']").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until(EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                time.sleep(TimeSpeed)
                TestResult.append("Add new Email process successfully verified")
                TestResultStatus.append("Pass")
            except Exception as er1:
                TestResult.append("Below error found while verifying Add new Email process\n"+str(er1))
                TestResultStatus.append("Fail")
            # ----------------------------------------------------------------------------------------------

            # -------Verifying add new Letter process--------------------------------------------------------
            try:
                select = Select(driver.find_element_by_xpath(
                    "//div[@class='content yellow_color']/div[1]/div/form/div/div/div/select"))
                select.select_by_visible_text(Selector[1])
                driver.find_element_by_xpath("//input[@id='autocompleteSystemContactSearch']").send_keys("sumreet client")
                time.sleep(TimeSpeed)
                ActionChains(driver).key_down(Keys.DOWN).key_up(Keys.DOWN).perform()
                time.sleep(TimeSpeed)
                ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                time.sleep(TimeSpeed)
                driver.find_element_by_xpath(
                    "//div[@id='communication-log-form']/form/div[3]/div/div[2]/div/div[3]/div[1]/div[2]/div/input").send_keys(
                    "Test Subject")
                time.sleep(TimeSpeed)
                ActionChains(driver).key_down(Keys.TAB).key_up(Keys.TAB).perform()
                time.sleep(TimeSpeed)
                ActionChains(driver).key_down(Keys.TAB).key_up(Keys.TAB).perform()
                TEXT = "Test Content"
                pyperclip.copy(TEXT)
                ActionChains(driver).key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
                time.sleep(TimeSpeed)
                select = Select(driver.find_element_by_xpath(
                    "//div[@id='communication-log-form']/form/div[3]/div/div[5]/div/div[2]/div/div[1]/div/select"))
                select.select_by_index(1)
                time.sleep(TimeSpeed)
                driver.find_element_by_xpath("//button[text()='Save']").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until(EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                time.sleep(TimeSpeed)
                TestResult.append("Add new Letter process successfully verified")
                TestResultStatus.append("Pass")
            except Exception as er2:
                TestResult.append("Below error found while verifying Add new Letter process\n" + str(er2))
                TestResultStatus.append("Fail")

            # -------Verifying add new Phone call process-------------------------------------------------
            try:
                select = Select(driver.find_element_by_xpath(
                    "//div[@class='content yellow_color']/div[1]/div/form/div/div/div/select"))
                select.select_by_visible_text(Selector[3])
                driver.find_element_by_xpath("//input[@id='autocompleteCallerNameSearch']").send_keys("sumreet client")
                time.sleep(TimeSpeed)
                ActionChains(driver).key_down(Keys.DOWN).key_up(Keys.DOWN).perform()
                time.sleep(TimeSpeed)
                ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                time.sleep(TimeSpeed)
                driver.find_element_by_xpath(
                    "//input[@id='autocompleteSystemContactSearch']").send_keys(
                    "Test System Contact")
                time.sleep(TimeSpeed)
                driver.find_element_by_xpath(
                    "//h3[@class='NoteAlertSectionHed NoteSectionHed']").click()
                time.sleep(TimeSpeed)
                driver.find_element_by_xpath(
                    "//input[@name='note_heading']").send_keys(
                    "Test call Note")
                time.sleep(TimeSpeed)
                ActionChains(driver).key_down(Keys.TAB).key_up(Keys.TAB).perform()
                time.sleep(TimeSpeed)
                TEXT = "Test Content"
                pyperclip.copy(TEXT)
                ActionChains(driver).key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
                time.sleep(TimeSpeed)
                select = Select(driver.find_element_by_xpath(
                    "//select[@name='communication_method']"))
                select.select_by_index(1)
                time.sleep(TimeSpeed)
                driver.find_element_by_xpath("//button[text()='Save']").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until(EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                time.sleep(TimeSpeed)
                TestResult.append("Add new Phone call process successfully verified")
                TestResultStatus.append("Pass")
            except Exception as er3:
                TestResult.append("Below error found while verifying Add new Phone call process\n" + str(er3))
                TestResultStatus.append("Fail")
            # ----------------------------------------------------------------------------------------------

            # -------Verifying add SMS process-------------------------------------------------
            try:
                select = Select(driver.find_element_by_xpath(
                    "//div[@class='content yellow_color']/div[1]/div/form/div/div/div/select"))
                select.select_by_visible_text(Selector[2])
                driver.find_element_by_xpath("//input[@name='system_contact']").send_keys("sumreet client")
                time.sleep(TimeSpeed)
                ActionChains(driver).key_down(Keys.DOWN).key_up(Keys.DOWN).perform()
                time.sleep(TimeSpeed)
                ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                time.sleep(TimeSpeed)
                driver.find_element_by_xpath(
                    "//input[@name='note_heading']").send_keys(
                    "Test SMS")
                time.sleep(TimeSpeed)
                driver.find_element_by_xpath(
                    "//textarea[@id='editor-full1']").send_keys(
                    "This is a test SMS")
                time.sleep(TimeSpeed)
                select = Select(driver.find_element_by_xpath(
                    "//select[@name='communication_method']"))
                select.select_by_index(1)
                time.sleep(TimeSpeed)
                driver.find_element_by_xpath("//button[text()='Save']").click()
                try:
                    WebDriverWait(driver, SHORT_TIMEOUT
                                  ).until(EC.presence_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))

                    WebDriverWait(driver, LONG_TIMEOUT
                                  ).until(EC.invisibility_of_element_located((By.XPATH, LOADING_ELEMENT_XPATH)))
                except TimeoutException:
                    pass
                time.sleep(TimeSpeed)
                driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[2]/a/i").click()
                time.sleep(TimeSpeed)
                TestResult.append("Add SMS process successfully verified")
                TestResultStatus.append("Pass")
            except Exception as er4:
                TestResult.append("Below error found while verifying Add SMS process\n" + str(er4))
                TestResultStatus.append("Fail")
            # ----------------------------------------------------------------------------------------------

            # -------Verifying add New plan form process-------------------------------------------------
            PageName = "Communication log"
            try:
                try:
                    print()
                    select = Select(driver.find_element_by_xpath(
                        "//div[@class='content yellow_color']/div[1]/div/form/div/div/div/select"))
                    select.select_by_visible_text(Selector[4])
                    driver.find_element_by_xpath("//div[@id='communication-log-form']/form/div[2]/div[1]/div/div/label[1]/input").click()
                    time.sleep(TimeSpeed)
                    driver.find_element_by_xpath(
                        "//div[@id='communication-log-form']/form/div[2]/div[2]/div/input").send_keys("Sumreet Client2")
                    time.sleep(TimeSpeed)
                    ActionChains(driver).key_down(Keys.DOWN).key_up(Keys.DOWN).perform()
                    time.sleep(TimeSpeed)
                    ActionChains(driver).key_down(Keys.ENTER).key_up(Keys.ENTER).perform()
                    time.sleep(TimeSpeed)
                except Exception:
                    pass

                # # # -------Participant information table-------------------------------------------------
                # # ParticipantInfo = driver.find_elements_by_xpath("//div[@id='client-div']/div[2]/div/table/tbody/tr")
                # # RowsLength = len(ParticipantInfo)
                # # for a in range(1,RowsLength, 2):
                # #     driver.find_element_by_xpath(
                # #         "//div[@id='client-div']/div[2]/div/table/tbody/tr["+str(a)+"]/td[3]/div/div/label[1]/input").click()
                # #     time.sleep(TimeSpeed)
                #
                # # -------Additional contacts table-------------------------------------------------
                #
                # print()
                # for scrolldown in range(1, 10):
                #     time.sleep(2)
                #     try:
                #         driver.find_element_by_xpath(
                #             "//div[@id='add-cont-div']/div/div/table/tbody/tr[1]/td")
                #         break
                #     except Exception:
                #         # ActionChains(driver).key_down(Keys.).perform()
                #         print("Inside Excep")
                #         ActionChains(driver).key_down(Keys.PAGE_DOWN).perform()
                #         print("Page Down")
                #         pass
                # AdditionalText = driver.find_elements_by_xpath(
                #     "//div[@id='add-cont-div']/div/div/table/tbody/tr[1]/td").text
                # print(AdditionalText)
                # AdditionalCont = driver.find_elements_by_xpath("//div[@id='add-cont-div']/div/div/table/tbody/tr")
                # AddConLength = len(AdditionalCont)
                # print(AddConLength)
                # for i in range(1,AddConLength):
                #     if "No Contacts Available" not in AdditionalText:
                #         driver.find_element_by_xpath(
                #             "//div[@id='add-cont-div']/div/div/table/tbody/tr["+str(i)+"]/td[8]/div/div/label[1]/input").click()
                #         time.sleep(TimeSpeed)
                #     if "No Contacts Available" in AdditionalText:
                #         driver.find_element_by_xpath(
                #             "//div[@id='add-cont-div']/div/div/table/tbody/tr[last()]/td/a").click()
                #         time.sleep(TimeSpeed)
                #         driver.find_element_by_xpath(
                #             "//div[@id='AddAnotherContact']/div/div/div[2]/form/div[1]/div/input").send_keys(
                #             "Test First Name")
                #         time.sleep(TimeSpeed)
                #         driver.find_element_by_xpath(
                #             "//div[@id='AddAnotherContact']/div/div/div[2]/form/div[2]/div/input").send_keys(
                #             "Test Last Name")
                #         time.sleep(TimeSpeed)
                #         select = Select(driver.find_element_by_xpath(
                #             "//div[@id='AddAnotherContact']/div/div/div[2]/form/div[3]/div/select"))
                #         select.select_by_visible_text("Grandma")
                #         time.sleep(TimeSpeed)
                #         select = Select(driver.find_element_by_xpath(
                #             "//div[@id='AddAnotherContact']/div/div/div[2]/form/div[4]/div/select"))
                #         select.select_by_visible_text("sumreet SP")
                #         time.sleep(TimeSpeed)
                #         driver.find_element_by_xpath(
                #             "//div[@id='AddAnotherContact']/div/div/div[2]/form/div[5]/div/input").send_keys(
                #             "abc@test.com")
                #         time.sleep(TimeSpeed)
                #         select = Select(driver.find_element_by_xpath(
                #             "//div[@id='AddAnotherContact']/div/div/div[2]/form/div[6]/div/select"))
                #         select.select_by_visible_text("Account Nominee")
                #         driver.find_element_by_xpath(
                #             "//div[@id='AddAnotherContact']/div/div/div[2]/form/div[last()]/button").click()
                #         time.sleep(TimeSpeed)
                #         print("Additional contact added successfully")
                #
                # # ----------------------------------------------------------------------------------------------

                # -------App access table-------------------------------------------------
                try:
                    AppAccess = driver.find_elements_by_xpath("//div[@id='add-access-div']/div/div/table/tbody/tr")
                    AppText =  driver.find_elements_by_xpath("//div[@id='add-access-div']/div/div/table/tbody/tr[1]/td").text
                    AppAccLength = len(AppAccess)
                    print(AppAccLength)
                    print(AppText)
                    for ac in range(1,AppAccLength):
                        if "No Contacts Available" not in AppText:
                            driver.find_element_by_xpath("//div[@id='add-access-div']/div/div/table/tbody//tr[" + str(i) + "]/td[8]/div/div/label[1]/input").click()
                            time.sleep(TimeSpeed)
                        if "No Contacts Available" in AppText:
                            driver.find_element_by_xpath(
                                "//div[@id='add-access-div']/div/div/table/tbody/tr[2]/td/a").click()
                            time.sleep(TimeSpeed)
                            driver.find_element_by_xpath(
                                "//div[@id='AddClientAccess']/div/div/div[2]/form/div[1]/div/input").send_keys(
                                "Test First Name")
                            time.sleep(TimeSpeed)
                            driver.find_element_by_xpath(
                                "//div[@id='AddClientAccess']/div/div/div[2]/form/div[2]/div/input").send_keys(
                                "Test Last Name")
                            time.sleep(TimeSpeed)
                            select = Select(driver.find_element_by_xpath(
                                "//div[@id='AddClientAccess']/div/div/div[2]/form/div[3]/div/select"))
                            select.select_by_visible_text("Grandma")
                            time.sleep(TimeSpeed)
                            select = Select(driver.find_element_by_xpath(
                                "//div[@id='AddClientAccess']/div/div/div[2]/form/div[4]/div/select"))
                            select.select_by_visible_text("sumreet SP")
                            time.sleep(TimeSpeed)
                            driver.find_element_by_xpath(
                                "//div[@id='AddClientAccess']/div/div/div[2]/form/div[5]/div/input").send_keys(
                                "abc@test.com")
                            time.sleep(TimeSpeed)
                            select = Select(driver.find_element_by_xpath(
                                "//div[@id='AddClientAccess']/div/div/div[2]/form/div[6]/div/select"))
                            select.select_by_visible_text("Account Nominee")
                            driver.find_element_by_xpath(
                                "//div[@id='AddClientAccess']/div/div/div[2]/form/div[last()]/button").click()
                            time.sleep(TimeSpeed)
                            print("App access added successfully")
                except Exception:
                 pass

                # # -------Add Additional Contacts-------------------------------------------------
                # # driver.find_element_by_xpath(
                # #     "//div[@id='add-cont-div']/div/div/table/tbody/tr[last()]/td/a").click()
                # # time.sleep(TimeSpeed)
                # # driver.find_element_by_xpath(
                # #     "//div[@id='AddAnotherContact']/div/div/div[2]/form/div[1]/div/input").send_keys("Test First Name")
                # # time.sleep(TimeSpeed)
                # # driver.find_element_by_xpath(
                # #     "//div[@id='AddAnotherContact']/div/div/div[2]/form/div[2]/div/input").send_keys("Test Last Name")
                # # time.sleep(TimeSpeed)
                # # select = Select(driver.find_element_by_xpath(
                # #     "//div[@id='AddAnotherContact']/div/div/div[2]/form/div[3]/div/select"))
                # # select.select_by_visible_text("Grandma")
                # # time.sleep(TimeSpeed)
                # # select = Select(driver.find_element_by_xpath(
                # #     "//div[@id='AddAnotherContact']/div/div/div[2]/form/div[4]/div/select"))
                # # select.select_by_visible_text("sumreet SP")
                # # time.sleep(TimeSpeed)
                # # driver.find_element_by_xpath(
                # #     "//div[@id='AddAnotherContact']/div/div/div[2]/form/div[5]/div/input").send_keys("abc@test.com")
                # # time.sleep(TimeSpeed)
                # # select = Select(driver.find_element_by_xpath(
                # #     "//div[@id='AddAnotherContact']/div/div/div[2]/form/div[6]/div/select"))
                # # select.select_by_visible_text("Account Nominee")
                # # driver.find_element_by_xpath(
                # #     "//div[@id='AddAnotherContact']/div/div/div[2]/form/div[last()]/button").click()
                # # time.sleep(TimeSpeed)
                # # print("Additional contact added successfully")
                # # driver.find_element_by_xpath(
                # #     "//div[@id='add-access-div']/div/div/table/tbody/tr[1]/td[8]/div/div/label[1]/input").click()
                # # time.sleep(TimeSpeed)
                # #
                # # # -------Add App Access-------------------------------------------------
                # # driver.find_element_by_xpath(
                # #         "//div[@id='add-access-div']/div/div/table/tbody/tr[2]/td/a").click()
                # # time.sleep(TimeSpeed)
                # # driver.find_element_by_xpath(
                # #     "//div[@id='AddClientAccess']/div/div/div[2]/form/div[1]/div/input").send_keys("Test First Name")
                # # time.sleep(TimeSpeed)
                # # driver.find_element_by_xpath(
                # #     "//div[@id='AddClientAccess']/div/div/div[2]/form/div[2]/div/input").send_keys("Test Last Name")
                # # time.sleep(TimeSpeed)
                # # select = Select(driver.find_element_by_xpath(
                # #     "//div[@id='AddClientAccess']/div/div/div[2]/form/div[3]/div/select"))
                # # select.select_by_visible_text("Grandma")
                # # time.sleep(TimeSpeed)
                # # select = Select(driver.find_element_by_xpath(
                # #     "//div[@id='AddClientAccess']/div/div/div[2]/form/div[4]/div/select"))
                # # select.select_by_visible_text("sumreet SP")
                # # time.sleep(TimeSpeed)
                # # driver.find_element_by_xpath(
                # #     "//div[@id='AddClientAccess']/div/div/div[2]/form/div[5]/div/input").send_keys("abc@test.com")
                # # time.sleep(TimeSpeed)
                # # select = Select(driver.find_element_by_xpath(
                # #     "//div[@id='AddClientAccess']/div/div/div[2]/form/div[6]/div/select"))
                # # select.select_by_visible_text("Account Nominee")
                # # driver.find_element_by_xpath(
                # #     "//div[@id='AddClientAccess']/div/div/div[2]/form/div[last()]/button").click()
                # # time.sleep(TimeSpeed)
                # # print("App access added successfully")
                #
                # # -------Communication log New plan form-------------------------------------------------
                # driver.find_element_by_xpath(
                #     "//div[@id='communication-log-form']/form/div[2]/div[6]/div/textarea").send_keys("Test feedback")
                # time.sleep(TimeSpeed)
                # driver.find_element_by_xpath(
                #     "//div[@id='communication-log-form']/form/div[2]/div[7]/div/div/label[1]/input").click()
                # time.sleep(TimeSpeed)
                # driver.find_element_by_xpath(
                #     "//div[@id='communication-log-form']/form/div[2]/div[8]/div/textarea").send_keys("Test regular check in's")
                # time.sleep(TimeSpeed)
                # driver.find_element_by_xpath(
                #     "//div[@id='communication-log-form']/form/div[2]/div[9]/div/textarea").send_keys(
                #     "Test action")
                # time.sleep(TimeSpeed)
                # driver.find_element_by_xpath("//button[text()='Save']").click()
                # time.sleep(TimeSpeed)
                # SuccessText = driver.find_element_by_xpath(
                #     "//div[@class='content-wrapper']/div/p").text
                # print(SuccessText)
                # if SuccessText=="Communication Log data is saved successfully.":
                #     TestResult.append(SuccessText)
                #     TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " data is not saved")
                TestResultStatus.append("Fail")
            # ----------------------------------------------------------------------------------------------

        except Exception as err:
            print(err)
            TestResult.append("Communication Log is not working correctly. Below error found\n"+str(err))
            TestResultStatus.append("Fail")
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


